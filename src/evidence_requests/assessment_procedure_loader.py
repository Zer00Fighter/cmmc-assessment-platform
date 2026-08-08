from __future__ import annotations

import re

from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.evidence_requests.catalog_compiler import (
    AssessmentProcedureRow,
)


class AssessmentProcedureLoaderError(ValueError):
    """Raised when assessment procedure source data cannot be loaded."""


@dataclass(frozen=True, slots=True)
class AssessmentProcedureDataset:
    """
    Normalized assessment-procedure dataset.

    The dataset preserves source metadata while exposing
    AssessmentProcedureRow records to downstream compilers.
    """

    framework_id: str

    framework_name: str = ""

    framework_version: str = ""

    source_document: str = ""

    source_revision: str = ""

    source_file: str = ""

    source_sheet: str = ""

    rows: Tuple[
        AssessmentProcedureRow,
        ...
    ] = field(
        default_factory=tuple
    )

    def __post_init__(self) -> None:
        framework_id = self.framework_id.strip()

        if not framework_id:
            raise AssessmentProcedureLoaderError(
                "AssessmentProcedureDataset.framework_id "
                "cannot be blank."
            )

        object.__setattr__(
            self,
            "framework_id",
            framework_id,
        )

        for field_name in (
            "framework_name",
            "framework_version",
            "source_document",
            "source_revision",
            "source_file",
            "source_sheet",
        ):
            object.__setattr__(
                self,
                field_name,
                str(
                    getattr(
                        self,
                        field_name,
                    )
                ).strip(),
            )

        object.__setattr__(
            self,
            "rows",
            tuple(
                self.rows
            ),
        )

    @property
    def row_count(self) -> int:
        return len(
            self.rows
        )

    @property
    def requirement_ids(
        self,
    ) -> Tuple[str, ...]:
        values: List[str] = []
        seen = set()

        for row in self.rows:
            key = (
                row.requirement_id.casefold()
            )

            if key in seen:
                continue

            seen.add(
                key
            )

            values.append(
                row.requirement_id
            )

        return tuple(
            values
        )

    @property
    def requirement_count(self) -> int:
        return len(
            self.requirement_ids
        )

    @property
    def objective_count(self) -> int:
        return sum(
            bool(
                row.objective_id
            )
            for row in self.rows
        )

    def for_requirement(
        self,
        requirement_id: str,
    ) -> Tuple[
        AssessmentProcedureRow,
        ...
    ]:
        key = (
            requirement_id
            .strip()
            .casefold()
        )

        return tuple(
            row
            for row in self.rows
            if (
                row.requirement_id
                .casefold()
                == key
            )
        )


@dataclass(frozen=True, slots=True)
class _RequirementContext:
    """
    Internal context retained while reading the
    objective rows belonging to a requirement.
    """

    family: str

    source_requirement_id: str

    requirement_id: str

    requirement_title: str

    requirement_text: str

    examine: str

    interview: str

    test: str

    source_row: int


class AssessmentProcedureLoader:
    """
    Load NIST SP 800-171A style assessment procedures.

    For CMMC Level 2, control identifiers are normalized into
    the same canonical IDs used by the Assessment workbook:

        AC + 3.1.1
            ->
        AC.L2-3.1.1

        IA + 3.5.3
            ->
        IA.L2-3.5.3

    Assessment objectives remain separate:

        3.1.1[a]
            ->
        requirement_id = AC.L2-3.1.1
        objective_id   = a
    """

    DEFAULT_SHEET_NAME = "SP800-171A"

    FAMILY_CODES = {
        "access control": "AC",
        "awareness and training": "AT",
        "audit and accountability": "AU",
        "configuration management": "CM",
        "identification and authentication": "IA",
        "incident response": "IR",
        "maintenance": "MA",
        "media protection": "MP",
        "personnel security": "PS",
        "physical protection": "PE",
        "risk assessment": "RA",
        "security assessment": "CA",
        "system and communications protection": "SC",
        "system and information integrity": "SI",
    }

    REQUIRED_HEADERS = (
        "Family",
        "Identifier",
        "Security Requirement",
        "Assessment Objective",
        (
            "Potential Assessment Method "
            "and Objects: Examine"
        ),
        (
            "Potential Assessment Method "
            "and Objects: Interview"
        ),
        (
            "Potential Assessment Method "
            "and Objects: Test"
        ),
    )

    _OBJECTIVE_PATTERN = re.compile(
        r"^(?P<requirement>.+?)"
        r"\[(?P<objective>[^\]]+)\]$"
    )

    def __init__(
        self,
        *,
        framework_id: str,
        framework_name: str = "",
        framework_version: str = "",
        source_document: str = "",
        source_revision: str = "",
        sheet_name: str = DEFAULT_SHEET_NAME,
        requirement_text_provider: Optional[
            Callable[
                [str],
                Optional[str],
            ]
        ] = None,
        requirement_title_mapper: Optional[
            Callable[
                [str, str, str],
                str,
            ]
        ] = None,
        sprs_weight_provider: Optional[
            Callable[
                [str],
                Optional[int],
            ]
        ] = None,
    ) -> None:
        framework_id = framework_id.strip()
        sheet_name = sheet_name.strip()

        if not framework_id:
            raise AssessmentProcedureLoaderError(
                "framework_id cannot be blank."
            )

        if not sheet_name:
            raise AssessmentProcedureLoaderError(
                "sheet_name cannot be blank."
            )

        self.framework_id = framework_id

        self.framework_name = (
            framework_name.strip()
        )

        self.framework_version = (
            framework_version.strip()
        )

        self.source_document = (
            source_document.strip()
        )

        self.source_revision = (
            source_revision.strip()
        )

        self.sheet_name = sheet_name

        self.requirement_text_provider = (
            requirement_text_provider
        )

        self.requirement_title_mapper = (
            requirement_title_mapper
        )

        self.sprs_weight_provider = (
            sprs_weight_provider
        )

    def load(
        self,
        path: str | Path,
    ) -> AssessmentProcedureDataset:
        path = Path(
            path
        )

        if not path.exists():
            raise AssessmentProcedureLoaderError(
                "Assessment procedure workbook "
                f"does not exist: {path}"
            )

        if not path.is_file():
            raise AssessmentProcedureLoaderError(
                "Assessment procedure path "
                f"is not a file: {path}"
            )

        if path.suffix.lower() not in {
            ".xlsx",
            ".xlsm",
        }:
            raise AssessmentProcedureLoaderError(
                "Unsupported assessment procedure "
                f"workbook type: {path.suffix}"
            )

        try:
            workbook = load_workbook(
                filename=path,
                read_only=True,
                data_only=True,
            )

        except Exception as error:
            raise AssessmentProcedureLoaderError(
                "Unable to open assessment "
                f"procedure workbook {path}: "
                f"{error}"
            ) from error

        try:
            if (
                self.sheet_name
                not in workbook.sheetnames
            ):
                raise AssessmentProcedureLoaderError(
                    "Assessment procedure worksheet "
                    f"{self.sheet_name!r} was not found. "
                    "Available worksheets: "
                    + ", ".join(
                        workbook.sheetnames
                    )
                )

            worksheet = workbook[
                self.sheet_name
            ]

            rows = self._load_worksheet(
                worksheet
            )

        finally:
            workbook.close()

        return AssessmentProcedureDataset(
            framework_id=self.framework_id,
            framework_name=(
                self.framework_name
            ),
            framework_version=(
                self.framework_version
            ),
            source_document=(
                self.source_document
            ),
            source_revision=(
                self.source_revision
            ),
            source_file=path.name,
            source_sheet=self.sheet_name,
            rows=tuple(
                rows
            ),
        )

    def _load_worksheet(
        self,
        worksheet: Worksheet,
    ) -> List[AssessmentProcedureRow]:
        header_map = self._read_headers(
            worksheet
        )

        result: List[
            AssessmentProcedureRow
        ] = []

        current: Optional[
            _RequirementContext
        ] = None

        for excel_row in range(
            2,
            worksheet.max_row + 1,
        ):
            values = self._row_values(
                worksheet,
                excel_row,
                header_map,
            )

            identifier = self._clean_text(
                values.get(
                    "Identifier"
                )
            )

            family = self._clean_text(
                values.get(
                    "Family"
                )
            )

            requirement_text = (
                self._clean_text(
                    values.get(
                        "Security Requirement"
                    )
                )
            )

            objective_text = (
                self._clean_objective_text(
                    values.get(
                        "Assessment Objective"
                    )
                )
            )

            examine = self._clean_text(
                values.get(
                    (
                        "Potential Assessment Method "
                        "and Objects: Examine"
                    )
                )
            )

            interview = self._clean_text(
                values.get(
                    (
                        "Potential Assessment Method "
                        "and Objects: Interview"
                    )
                )
            )

            test = self._clean_text(
                values.get(
                    (
                        "Potential Assessment Method "
                        "and Objects: Test"
                    )
                )
            )

            if not any(
                (
                    identifier,
                    family,
                    requirement_text,
                    objective_text,
                    examine,
                    interview,
                    test,
                )
            ):
                continue

            if not identifier:
                continue

            parsed_identifier = (
                self._parse_identifier(
                    identifier
                )
            )

            #
            # BASE REQUIREMENT ROW
            #
            if parsed_identifier is None:
                current = (
                    self._create_requirement_context(
                        family=family,
                        source_requirement_id=(
                            identifier
                        ),
                        requirement_text=(
                            requirement_text
                        ),
                        examine=examine,
                        interview=interview,
                        test=test,
                        excel_row=excel_row,
                    )
                )

                result.append(
                    AssessmentProcedureRow(
                        framework_id=(
                            self.framework_id
                        ),
                        family=current.family,
                        requirement_id=(
                            current.requirement_id
                        ),
                        requirement_title=(
                            current.requirement_title
                        ),
                        requirement_text=(
                            current.requirement_text
                        ),
                        objective_id="",
                        objective_text="",
                        examine=current.examine,
                        interview=current.interview,
                        test=current.test,
                        sprs_weight=(
                            self._sprs_weight(
                                current.requirement_id
                            )
                        ),
                        source_document=(
                            self.source_document
                        ),
                        source_revision=(
                            self.source_revision
                        ),
                        source_location=(
                            self._source_location(
                                worksheet.title,
                                excel_row,
                                identifier,
                            )
                        ),
                    )
                )

                continue

            #
            # ASSESSMENT OBJECTIVE ROW
            #
            (
                source_requirement_id,
                objective_id,
            ) = parsed_identifier

            if current is None:
                raise AssessmentProcedureLoaderError(
                    "Assessment objective row "
                    f"{excel_row} ({identifier}) "
                    "appeared before a base "
                    "requirement row."
                )

            if (
                current.source_requirement_id
                .casefold()
                != source_requirement_id
                .casefold()
            ):
                raise AssessmentProcedureLoaderError(
                    "Assessment objective row "
                    f"{excel_row} ({identifier}) "
                    "does not belong to the current "
                    "requirement "
                    f"{current.source_requirement_id}."
                )

            if not objective_text:
                raise AssessmentProcedureLoaderError(
                    "Assessment objective row "
                    f"{excel_row} ({identifier}) "
                    "does not contain objective text."
                )

            #
            # Inherit the requirement-level authoritative
            # Examine / Interview / Test content where the
            # objective row itself is blank.
            #
            objective_examine = (
                examine
                or current.examine
            )

            objective_interview = (
                interview
                or current.interview
            )

            objective_test = (
                test
                or current.test
            )

            result.append(
        AssessmentProcedureRow(
            framework_id=(
                self.framework_id
            ),
        family=current.family, 
                    
                    requirement_id=(
                        current.requirement_id
                    ),
                    requirement_title=(
                        current.requirement_title
                    ),
                    requirement_text=(
                        current.requirement_text
                    ),
                    objective_id=objective_id,
                    objective_text=objective_text,
                    examine=objective_examine,
                    interview=objective_interview,
                    test=objective_test,
                    sprs_weight=(
                        self._sprs_weight(
                            current.requirement_id
                        )
                    ),
                    source_document=(
                        self.source_document
                    ),
                    source_revision=(
                        self.source_revision
                    ),
                    source_location=(
                        self._source_location(
                            worksheet.title,
                            excel_row,
                            identifier,
                        )
                    ),
                )
            )

        if not result:
            raise AssessmentProcedureLoaderError(
                "No assessment procedure rows "
                f"were found in worksheet "
                f"{worksheet.title!r}."
            )

        return result

    def _create_requirement_context(
        self,
        *,
        family: str,
        source_requirement_id: str,
        requirement_text: str,
        examine: str,
        interview: str,
        test: str,
        excel_row: int,
    ) -> _RequirementContext:
        requirement_id = (
            self._canonical_requirement_id(
                family,
                source_requirement_id,
            )
        )

        if not requirement_text:
            requirement_text = (
                self._get_requirement_text(
                    requirement_id
                )
            )

        if not requirement_text:
            raise AssessmentProcedureLoaderError(
                "Base requirement row "
                f"{excel_row} "
                f"({source_requirement_id}) "
                "does not contain Security Requirement "
                "text, and no fallback requirement text "
                f"was available for {requirement_id}."
            )

        title = self._map_requirement_title(
            requirement_id,
            family,
            requirement_text,
        )

        return _RequirementContext(
            family=self._canonical_family_code(
                family
            ),
            source_requirement_id=(
                source_requirement_id.strip()
            ),
            requirement_id=requirement_id,
            requirement_title=title,
            requirement_text=(
                requirement_text
            ),
            examine=examine,
            interview=interview,
            test=test,
            source_row=excel_row,
        )

    @classmethod
    def _canonical_requirement_id(
        cls,
        family: str,
        source_requirement_id: str,
    ) -> str:
        """
        Produce the canonical CMMC Level 2 control ID.

        The source workbook may provide either a two-letter family
        code such as ``AC`` or a full family name such as
        ``Access Control``. Both normalize to the same control ID.
        """
        family_code = cls._canonical_family_code(
            family
        )
        requirement = source_requirement_id.strip()

        if not requirement:
            raise AssessmentProcedureLoaderError(
                "Cannot construct canonical control ID "
                "from a blank requirement identifier."
            )

        expected_prefix = f"{family_code}.L2-"

        if requirement.upper().startswith(
            expected_prefix.upper()
        ):
            return requirement

        return f"{family_code}.L2-{requirement}"

    @classmethod
    def _canonical_family_code(
        cls,
        family: str,
    ) -> str:
        """Normalize a source family value to its CMMC family code."""
        raw_family = family.strip()

        if not raw_family:
            raise AssessmentProcedureLoaderError(
                "Control family cannot be blank."
            )

        if len(raw_family) == 2 and raw_family.isalpha():
            return raw_family.upper()

        family_key = raw_family.casefold()

        try:
            return cls.FAMILY_CODES[family_key]
        except KeyError as error:
            raise AssessmentProcedureLoaderError(
                "Unknown assessment control family: "
                f"{raw_family!r}"
            ) from error

    def _get_requirement_text(
        self,
        requirement_id: str,
    ) -> str:
        if self.requirement_text_provider is None:
            return ""

        try:
            value = self.requirement_text_provider(
                requirement_id
            )

        except Exception as error:
            raise AssessmentProcedureLoaderError(
                "Unable to obtain fallback "
                "requirement text for "
                f"{requirement_id!r}: "
                f"{error}"
            ) from error

        return self._clean_text(
            value
        )

    def _map_requirement_title(
        self,
        requirement_id: str,
        family: str,
        requirement_text: str,
    ) -> str:
        if (
            self.requirement_title_mapper
            is None
        ):
            return ""

        try:
            title = (
                self.requirement_title_mapper(
                    requirement_id,
                    family,
                    requirement_text,
                )
            )

        except Exception as error:
            raise AssessmentProcedureLoaderError(
                "Unable to map requirement title "
                f"for {requirement_id!r}: "
                f"{error}"
            ) from error

        return self._clean_text(
            title
        )

    def _sprs_weight(
        self,
        requirement_id: str,
    ) -> Optional[int]:
        if (
            self.sprs_weight_provider
            is None
        ):
            return None

        try:
            value = (
                self.sprs_weight_provider(
                    requirement_id
                )
            )

        except Exception as error:
            raise AssessmentProcedureLoaderError(
                "Unable to determine SPRS weight "
                f"for {requirement_id!r}: "
                f"{error}"
            ) from error

        if value is None:
            return None

        try:
            result = int(
                value
            )

        except (
            TypeError,
            ValueError,
        ) as error:
            raise AssessmentProcedureLoaderError(
                "Invalid SPRS weight returned "
                f"for {requirement_id!r}: "
                f"{value!r}"
            ) from error

        if result < 0:
            raise AssessmentProcedureLoaderError(
                "SPRS weight cannot be negative "
                f"for {requirement_id!r}."
            )

        return result

    def _read_headers(
        self,
        worksheet: Worksheet,
    ) -> Dict[str, int]:
        header_map: Dict[
            str,
            int,
        ] = {}

        for column in range(
            1,
            worksheet.max_column + 1,
        ):
            value = self._clean_text(
                worksheet.cell(
                    row=1,
                    column=column,
                ).value
            )

            if not value:
                continue

            header_map[
                value
            ] = column

        missing = [
            header
            for header
            in self.REQUIRED_HEADERS
            if header not in header_map
        ]

        if missing:
            raise AssessmentProcedureLoaderError(
                "Assessment procedure worksheet "
                "is missing required header(s): "
                + ", ".join(
                    missing
                )
            )

        return header_map

    @staticmethod
    def _row_values(
        worksheet: Worksheet,
        row: int,
        header_map: Dict[str, int],
    ) -> Dict[str, object]:
        return {
            header: worksheet.cell(
                row=row,
                column=column,
            ).value
            for (
                header,
                column,
            )
            in header_map.items()
        }

    @classmethod
    def _parse_identifier(
        cls,
        identifier: str,
    ) -> Optional[
        Tuple[str, str]
    ]:
        """
        Parse an objective identifier.

        Example:

            3.1.1[a]

        becomes:

            ("3.1.1", "a")

        A base control identifier such as:

            3.1.1

        returns None.
        """

        value = cls._clean_text(
            identifier
        )

        match = cls._OBJECTIVE_PATTERN.match(
            value
        )

        if match is None:
            return None

        
        requirement = cls._clean_text(
        match.group(
            "requirement"
            )
        )

        #
        # Some authoritative source rows use identifiers
        # such as:
        #
        #     3.12.4.[h]
        #
        # while others use:
        #
        #     3.12.4[h]
        #
        # Both represent the same parent requirement.
        #
        requirement = requirement.rstrip(
            "."
        ).strip()

        objective = cls._clean_text(
            match.group(
                "objective"
            )
        )

        if not requirement:
            return None

        if not objective:
            return None

        return (
            requirement,
            objective,
        )

    @staticmethod
    def _source_location(
        worksheet_name: str,
        row: int,
        identifier: str,
    ) -> str:
        return (
            f"{worksheet_name}!Row {row} "
            f"({identifier.strip()})"
        )

    @staticmethod
    def _clean_text(
        value: object,
    ) -> str:
        if value is None:
            return ""

        text = str(
            value
        )

        text = text.replace(
            "\r\n",
            "\n",
        ).replace(
            "\r",
            "\n",
        )

        text = re.sub(
            r"\s+",
            " ",
            text,
        )

        return text.strip()

    @classmethod
    def _clean_objective_text(
        cls,
        value: object,
    ) -> str:
        text = cls._clean_text(
            value
        )

        if (
            text
            .rstrip(":")
            .strip()
            .casefold()
            == "determine if"
        ):
            return ""

        return text