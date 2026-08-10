from __future__ import annotations

import csv
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook, workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from src.workbook.scoring_data import (
    WorkbookScoringData,
    WorkbookScoringRule,
)
from src.workbook.workbook_styles import WorkbookStyles
from src.workbook.worksheet_factory import WorksheetFactory

from src.workbook.evidence_sheet import EvidenceSheetBuilder

from src.workbook.evidence_sheet import (
    EvidenceSheetBuilder,
)

from src.workbook.workbook_pipeline import WorkbookPipeline

from src.workbook.poam_sheet import POAMSheetBuilder

class WorkbookBuilder:
    """Build the scoring-integrated CMMC Level 2 assessment workbook."""

    SHEET_ORDER = [
        "Cover",
        "Dashboard",
        "Assessment",
        "Domain Summary",
        "Evidence",
        "POA&M",
        "SSP Crosswalk",
        "Assessment History",
        "Executive Report",
        "Settings",
        "_Lists",
    ]

    EXPECTED_CONTROL_COUNT = 110
    FIRST_ASSESSMENT_ROW = 6
    MAXIMUM_SCORE = 110

    def __init__(
        self,
        project_root: Path,
        output_path: Path | None = None,
    ) -> None:
        self.project_root = project_root.resolve()

        self.controls_path = (
            self.project_root
            / "data"
            / "controls"
            / "cmmc_level2_controls.csv"
        )

        self.scoring_weights_path = (
            self.project_root
            / "data"
            / "scoring"
            / "scoring_weights.csv"
        )

        self.output_path = output_path or (
            self.project_root
            / "output"
            / "Omni_CMMC_Assessment_v0.2.xlsx"
        )

        self.styles = WorkbookStyles()
        self.factory = WorksheetFactory(
            self.styles
        )

        self.evidence_builder = EvidenceSheetBuilder(
        styles=self.styles,
        factory=self.factory,
        )

        self.poam_builder = POAMSheetBuilder(
        styles=self.styles,
        factory=self.factory,
        )

        self.styles = WorkbookStyles()

        self.factory = WorksheetFactory(
        self.styles
        )

        self.evidence_builder = EvidenceSheetBuilder(
        styles=self.styles,
        factory=self.factory,
        )

        self.poam_builder = POAMSheetBuilder(
        styles=self.styles,
        factory=self.factory,
        )

        self.pipeline = WorkbookPipeline()

    def build(self) -> Path:
        controls = self._load_controls()

        scoring_rules = WorkbookScoringData(
            self.scoring_weights_path
        ).load_map()

        self._validate_control_scoring_alignment(
            controls,
            scoring_rules,
        )

        workbook = Workbook()
        self.styles.configure_workbook(workbook)

        default_sheet = workbook.active
        workbook.remove(default_sheet)

        worksheets = {
            sheet_name: workbook.create_sheet(
                title=sheet_name
            )
            for sheet_name in self.SHEET_ORDER
        }

        self._build_cover(
            worksheets["Cover"]
        )

        self._build_dashboard(
            worksheets["Dashboard"]
        )

        self._build_assessment(
            worksheets["Assessment"],
            controls,
            scoring_rules,
        )

        self._build_domain_summary(
            worksheets["Domain Summary"],
            controls,
        )

        self.evidence_builder.build(
            worksheets["Evidence"]
        )

        self.poam_builder.build(
            worksheets["POA&M"]
        )
        
        self._build_placeholder_sheet(
            worksheets["SSP Crosswalk"],
            title="SSP Crosswalk",
            subtitle=(
                "Map CMMC requirements to System Security "
                "Plan sections, policies, procedures, and owners."
            ),
        )

        self._build_placeholder_sheet(
            worksheets["Assessment History"],
            title="Assessment History",
            subtitle=(
                "Store historical assessment snapshots and "
                "track score and readiness trends."
            ),
        )

        self._build_placeholder_sheet(
            worksheets["Executive Report"],
            title="Executive Report",
            subtitle=(
                "Printable management summary of assessment "
                "results, risks, and remediation priorities."
            ),
        )

        self._build_settings(
            worksheets["Settings"],
            scoring_rules,
        )

        self._build_lists(
            worksheets["_Lists"]
        )

        worksheets["_Lists"].sheet_state = (
            "veryHidden"
        )

        workbook.active = (
        workbook.sheetnames.index("Cover")
        )

        self.output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    
        )

        self.pipeline.run(workbook)
        workbook.save(self.output_path)

        return self.output_path

    def _load_controls(
        self,
    ) -> List[Dict[str, str]]:
        if not self.controls_path.exists():
            raise FileNotFoundError(
                "Compiled controls CSV was not found: "
                f"{self.controls_path}. "
                "Run python compile_guide.py first."
            )

        with self.controls_path.open(
            "r",
            encoding="utf-8-sig",
            newline="",
        ) as file:
            reader = csv.DictReader(file)

            if reader.fieldnames is None:
                raise ValueError(
                    "Compiled controls CSV has no header."
                )

            required_columns = {
                "domain_code",
                "requirement_id",
                "title",
                "statement",
                "source_page_start",
                "source_page_end",
            }

            missing_columns = (
                required_columns
                - set(reader.fieldnames)
            )

            if missing_columns:
                raise ValueError(
                    "Compiled controls CSV is missing columns: "
                    + ", ".join(
                        sorted(missing_columns)
                    )
                )

            controls = list(reader)

        if len(controls) != self.EXPECTED_CONTROL_COUNT:
            raise ValueError(
                f"Expected {self.EXPECTED_CONTROL_COUNT} "
                "compiled CMMC requirements, but found "
                f"{len(controls)}."
            )

        return controls

    @staticmethod
    def _validate_control_scoring_alignment(
        controls: List[Dict[str, str]],
        scoring_rules: Dict[
            str,
            WorkbookScoringRule,
        ],
    ) -> None:
        control_ids = {
            control["requirement_id"]
            .strip()
            .upper()
            for control in controls
        }

        scoring_ids = set(scoring_rules)

        missing_scoring = sorted(
            control_ids - scoring_ids
        )

        unexpected_scoring = sorted(
            scoring_ids - control_ids
        )

        if missing_scoring:
            raise ValueError(
                "Controls are missing scoring rules: "
                + ", ".join(missing_scoring)
            )

        if unexpected_scoring:
            raise ValueError(
                "Scoring data contains unknown requirements: "
                + ", ".join(unexpected_scoring)
            )

    def _build_cover(
        self,
        worksheet: Worksheet,
    ) -> None:
        self.factory.configure_standard_sheet(
            worksheet,
            zoom_scale=90,
        )

        self.factory.create_title_band(
            worksheet,
            title="Omni by R!SC",
            subtitle=(
                "CMMC Level 2 assessment readiness, evidence, "
                "POA&M, SSP mapping, and weighted scoring"
            ),
            end_column=10,
        )

        fields = [
            ("Organization Name", ""),
            ("Assessment Name", ""),
            ("Assessment Scope", ""),
            ("CAGE Code", ""),
            (
                "Assessment Type",
                "Level 2 Self-Assessment",
            ),
            ("Assessment Start Date", ""),
            ("Assessment End Date", ""),
            ("Lead Assessor", ""),
            ("Workbook Version", "0.2"),
            (
                "Assessment Guide",
                "Level 2 Version 2.13",
            ),
            (
                "Scoring Source",
                "32 CFR 170.24",
            ),
        ]

        start_row = 6

        for offset, (label, value) in enumerate(
            fields
        ):
            row = start_row + offset

            label_cell = worksheet.cell(
                row=row,
                column=2,
                value=label,
            )

            value_cell = worksheet.cell(
                row=row,
                column=3,
                value=value,
            )

            label_cell.font = (
                self.styles.header_font()
            )
            label_cell.fill = (
                self.styles.section_fill()
            )
            label_cell.border = (
                self.styles.thin_border()
            )
            label_cell.alignment = (
                self.styles.left_alignment()
            )

            value_cell.fill = (
                self.styles.input_fill()
            )
            value_cell.font = (
                self.styles.input_font()
            )
            value_cell.protection = (
                self.styles.unlocked_protection()
            )
            value_cell.border = (
                self.styles.thin_border()
            )
            value_cell.alignment = (
                self.styles.left_alignment()
            )

        worksheet["B19"] = "Instructions"
        worksheet["B19"].font = (
            self.styles.section_font()
        )
        worksheet["B19"].fill = (
            self.styles.section_fill()
        )

        worksheet.merge_cells(
            start_row=20,
            start_column=2,
            end_row=24,
            end_column=8,
        )

        instruction_cell = worksheet["B20"]
        instruction_cell.value = (
            "Complete the organization metadata above, then "
            "evaluate all 110 requirements on the Assessment "
            "worksheet. Yellow cells are user-input fields. "
            "Gray cells contain formulas or official scoring "
            "metadata. Requirements IA.L2-3.5.3 and "
            "SC.L2-3.13.11 include limited partial-credit "
            "implementation-state options."
        )
        instruction_cell.font = (
            self.styles.body_font()
        )
        instruction_cell.fill = (
            self.styles.subheader_fill()
        )
        instruction_cell.border = (
            self.styles.thin_border()
        )
        instruction_cell.alignment = (
            self.styles.left_alignment()
        )

        widths = {
            "A": 3,
            "B": 28,
            "C": 48,
            "D": 4,
            "E": 18,
            "F": 18,
            "G": 18,
            "H": 18,
            "I": 18,
            "J": 18,
        }

        for column, width in widths.items():
            worksheet.column_dimensions[
                column
            ].width = width

    def _build_dashboard(
        self,
        worksheet: Worksheet,
    ) -> None:
        self.factory.configure_standard_sheet(
            worksheet,
            zoom_scale=90,
        )

        self.factory.create_title_band(
            worksheet,
            title="Executive Dashboard",
            subtitle=(
                "Weighted CMMC Level 2 assessment and "
                "readiness summary"
            ),
            end_column=13,
        )

        first_row = self.FIRST_ASSESSMENT_ROW
        last_row = (
            first_row
            + self.EXPECTED_CONTROL_COUNT
            - 1
        )

        cards = [
            (
                "Current Score",
                (
                    f"={self.MAXIMUM_SCORE}-"
                    f"SUM(Assessment!N{first_row}:"
                    f"N{last_row})"
                ),
            ),
            (
                "Requirements Met",
                (
                    '=COUNTIF('
                    f'Assessment!I{first_row}:'
                    f'I{last_row},"MET")'
                ),
            ),
            (
                "Requirements Not Met",
                (
                    '=COUNTIF('
                    f'Assessment!I{first_row}:'
                    f'I{last_row},"NOT MET")'
                ),
            ),
            (
                "Not Assessed",
                (
                    '=COUNTIF('
                    f'Assessment!I{first_row}:'
                    f'I{last_row},"NOT ASSESSED")'
                ),
            ),
        ]

        start_columns = [
            2,
            5,
            8,
            11,
        ]

        for (
            label,
            formula,
        ), start_column in zip(
            cards,
            start_columns,
        ):
            worksheet.merge_cells(
                start_row=6,
                start_column=start_column,
                end_row=6,
                end_column=start_column + 1,
            )

            worksheet.merge_cells(
                start_row=7,
                start_column=start_column,
                end_row=9,
                end_column=start_column + 1,
            )

            label_cell = worksheet.cell(
                row=6,
                column=start_column,
                value=label,
            )

            value_cell = worksheet.cell(
                row=7,
                column=start_column,
                value=formula,
            )

            label_cell.font = (
                self.styles.header_font()
            )
            label_cell.fill = (
                self.styles.header_fill()
            )
            label_cell.alignment = (
                self.styles.center_alignment()
            )
            label_cell.border = (
                self.styles.thin_border()
            )

            value_cell.font = (
                self.styles.title_font()
            )
            value_cell.fill = (
                self.styles.section_fill()
            )
            value_cell.alignment = (
                self.styles.center_alignment()
            )
            value_cell.border = (
                self.styles.thin_border()
            )

        dashboard_metrics = [
            (
                "Assessment Completion",
                (
                    '=IFERROR(('
                    f'COUNTIF(Assessment!I{first_row}:'
                    f'I{last_row},"MET")+'
                    f'COUNTIF(Assessment!I{first_row}:'
                    f'I{last_row},"NOT MET")+'
                    f'COUNTIF(Assessment!I{first_row}:'
                    f'I{last_row},"NOT APPLICABLE"))/'
                    f'{self.EXPECTED_CONTROL_COUNT},0)'
                ),
                "0%",
            ),
            (
                "Evidence Complete",
                (
                    '=COUNTIF('
                    f'Assessment!O{first_row}:'
                    f'O{last_row},"Complete")'
                ),
                "0",
            ),
            (
                "Partial Credit Applied",
                (
                    '=COUNTIF('
                    f'Assessment!M{first_row}:'
                    f'M{last_row},"Yes")'
                ),
                "0",
            ),
            (
                "POA&M Items",
                (
                    '=COUNTIF('
                    f'Assessment!S{first_row}:'
                    f'S{last_row},"Yes")'
                ),
                "0",
            ),
        ]

        metric_columns = [
            2,
            5,
            8,
            11,
        ]

        for (
            label,
            formula,
            number_format,
        ), column in zip(
            dashboard_metrics,
            metric_columns,
        ):
            label_cell = worksheet.cell(
                row=12,
                column=column,
                value=label,
            )

            value_cell = worksheet.cell(
                row=12,
                column=column + 1,
                value=formula,
            )

            label_cell.font = (
                self.styles.header_font()
            )
            label_cell.fill = (
                self.styles.header_fill()
            )
            label_cell.border = (
                self.styles.thin_border()
            )
            label_cell.alignment = (
                self.styles.center_alignment()
            )

            value_cell.fill = (
                self.styles.formula_fill()
            )
            value_cell.border = (
                self.styles.thin_border()
            )
            value_cell.alignment = (
                self.styles.center_alignment()
            )
            value_cell.number_format = (
                number_format
            )

        worksheet["B15"] = "Scoring Status"
        worksheet["B15"].font = (
            self.styles.header_font()
        )
        worksheet["B15"].fill = (
            self.styles.header_fill()
        )
        worksheet["B15"].border = (
            self.styles.thin_border()
        )

        worksheet["C15"] = (
            '=IF(COUNTIF('
            f'Assessment!I{first_row}:'
            f'I{last_row},"NOT ASSESSED")=0,'
            '"COMPLETE","PROVISIONAL")'
        )
        worksheet["C15"].fill = (
            self.styles.formula_fill()
        )
        worksheet["C15"].border = (
            self.styles.thin_border()
        )
        worksheet["C15"].alignment = (
            self.styles.center_alignment()
        )

        worksheet["E15"] = (
            "Maximum Deduction"
        )
        worksheet["E15"].font = (
            self.styles.header_font()
        )
        worksheet["E15"].fill = (
            self.styles.header_fill()
        )
        worksheet["E15"].border = (
            self.styles.thin_border()
        )

        worksheet["F15"] = (
            f"=SUM(Assessment!L{first_row}:"
            f"L{last_row})"
        )
        worksheet["F15"].fill = (
            self.styles.formula_fill()
        )
        worksheet["F15"].border = (
            self.styles.thin_border()
        )
        worksheet["F15"].alignment = (
            self.styles.center_alignment()
        )

        worksheet["H15"] = (
            "Mathematical Minimum"
        )
        worksheet["H15"].font = (
            self.styles.header_font()
        )
        worksheet["H15"].fill = (
            self.styles.header_fill()
        )
        worksheet["H15"].border = (
            self.styles.thin_border()
        )

        worksheet["I15"] = (
            f"={self.MAXIMUM_SCORE}-F15"
        )
        worksheet["I15"].fill = (
            self.styles.formula_fill()
        )
        worksheet["I15"].border = (
            self.styles.thin_border()
        )
        worksheet["I15"].alignment = (
            self.styles.center_alignment()
        )

        for column in range(1, 14):
            column_letter = get_column_letter(
                column
            )
            worksheet.column_dimensions[
                column_letter
            ].width = 13

    def _build_assessment(
        self,
        worksheet: Worksheet,
        controls: List[Dict[str, str]],
        scoring_rules: Dict[
            str,
            WorkbookScoringRule,
        ],
    ) -> None:
        self.factory.configure_standard_sheet(
            worksheet,
            freeze_cell="A6",
            show_gridlines=False,
            zoom_scale=75,
        )

        self.factory.create_title_band(
            worksheet,
            title=(
                "CMMC Level 2 Requirement Assessment"
            ),
            subtitle=(
                "Evaluate implementation, evidence, ownership, "
                "SSP mapping, weighted deductions, and POA&M "
                "requirements for all 110 controls."
            ),
            end_column=22,
        )

        headers = [
            "Domain",                    # A
            "Requirement ID",            # B
            "Title",                     # C
            "Requirement Statement",     # D
            "Source Start",              # E
            "Source End",                # F
            "Applicable",                # G
            "Scoring Category",          # H
            "Status",                    # I
            "Implementation State",      # J
            "Partial Credit Allowed",    # K
            "Full Deduction",            # L
            "Partial Credit Applied",    # M
            "Calculated Deduction",      # N
            "Evidence Status",           # O
            "Control Owner",             # P
            "SSP Reference",             # Q
            "Assessor Notes",            # R
            "POA&M Required",            # S
            "Partial Credit Condition",  # T
            "Scoring Source",            # U
            "Score Explanation",         # V
        ]

        header_row = 5

        for column, header in enumerate(
            headers,
            start=1,
        ):
            worksheet.cell(
                row=header_row,
                column=column,
                value=header,
            )

        self.factory.style_table_header(
            worksheet,
            row_number=header_row,
            start_column=1,
            end_column=len(headers),
        )

        first_data_row = (
            self.FIRST_ASSESSMENT_ROW
        )

        for row_offset, control in enumerate(
            controls
        ):
            row = first_data_row + row_offset

            requirement_id = (
                control["requirement_id"]
                .strip()
                .upper()
            )

            rule = scoring_rules[
                requirement_id
            ]

            worksheet.cell(
                row=row,
                column=1,
                value=control["domain_code"],
            )

            worksheet.cell(
                row=row,
                column=2,
                value=requirement_id,
            )

            worksheet.cell(
                row=row,
                column=3,
                value=control["title"],
            )

            worksheet.cell(
                row=row,
                column=4,
                value=control["statement"],
            )

            worksheet.cell(
                row=row,
                column=5,
                value=int(
                    control[
                        "source_page_start"
                    ]
                ),
            )

            worksheet.cell(
                row=row,
                column=6,
                value=int(
                    control[
                        "source_page_end"
                    ]
                ),
            )

            worksheet.cell(
                row=row,
                column=7,
                value="Yes",
            )

            worksheet.cell(
                row=row,
                column=8,
                value=rule.scoring_category,
            )

            worksheet.cell(
                row=row,
                column=9,
                value="NOT ASSESSED",
            )

            worksheet.cell(
                row=row,
                column=10,
                value=(
                    "NOT ASSESSED"
                    if rule.partial_credit_allowed
                    else ""
                ),
            )

            worksheet.cell(
                row=row,
                column=11,
                value=(
                    "Yes"
                    if rule.partial_credit_allowed
                    else "No"
                ),
            )

            worksheet.cell(
                row=row,
                column=12,
                value=rule.full_deduction_points,
            )

            worksheet.cell(
                row=row,
                column=13,
                value=(
                    f'=IF(AND('
                    f'K{row}="Yes",'
                    f'I{row}="NOT MET",'
                    f'J{row}="PARTIALLY IMPLEMENTED"),'
                    f'"Yes","No")'
                ),
            )

            worksheet.cell(
                row=row,
                column=14,
                value=self._deduction_formula(row),
            )

            worksheet.cell(
                row=row,
                column=15,
                value="Not Started",
            )

            worksheet.cell(
                row=row,
                column=16,
                value="",
            )

            worksheet.cell(
                row=row,
                column=17,
                value="",
            )

            worksheet.cell(
                row=row,
                column=18,
                value="",
            )

            worksheet.cell(
                row=row,
                column=19,
                value=(
                    f'=IF(I{row}="NOT MET",'
                    f'"Yes","No")'
                ),
            )

            worksheet.cell(
                row=row,
                column=20,
                value=rule.partial_credit_condition,
            )

            worksheet.cell(
                row=row,
                column=21,
                value=(
                    f"{rule.scoring_source} | "
                    f"{rule.scoring_source_version}"
                ),
            )

            worksheet.cell(
                row=row,
                column=22,
                value=self._explanation_formula(row),
            )

            for column in range(
                1,
                len(headers) + 1,
            ):
                cell = worksheet.cell(
                    row=row,
                    column=column,
                )

                cell.border = (
                    self.styles.thin_border()
                )
                cell.alignment = (
                    self.styles.left_alignment()
                )
                cell.font = (
                    self.styles.body_font()
                )

            input_columns = [
                7,
                9,
                10,
                15,
                16,
                17,
                18,
            ]

            for input_column in input_columns:
                cell = worksheet.cell(
                    row=row,
                    column=input_column,
                )

                if (
                    input_column == 10
                    and not rule.partial_credit_allowed
                ):
                    cell.fill = (
                        self.styles.formula_fill()
                    )
                    cell.protection = (
                        self.styles.locked_protection()
                    )
                else:
                    cell.fill = (
                        self.styles.input_fill()
                    )
                    cell.protection = (
                        self.styles.unlocked_protection()
                    )

            formula_columns = [
                8,
                11,
                12,
                13,
                14,
                19,
                20,
                21,
                22,
            ]

            for formula_column in formula_columns:
                worksheet.cell(
                    row=row,
                    column=formula_column,
                ).fill = (
                    self.styles.formula_fill()
                )

        last_data_row = (
            first_data_row
            + len(controls)
            - 1
        )

        self._add_assessment_validations(
            worksheet,
            first_data_row,
            last_data_row,
        )

        self._add_assessment_conditional_formatting(
            worksheet,
            first_data_row,
            last_data_row,
        )

        widths = {
            "A": 10,
            "B": 18,
            "C": 34,
            "D": 65,
            "E": 12,
            "F": 12,
            "G": 12,
            "H": 20,
            "I": 18,
            "J": 24,
            "K": 16,
            "L": 14,
            "M": 18,
            "N": 18,
            "O": 18,
            "P": 22,
            "Q": 22,
            "R": 45,
            "S": 16,
            "T": 60,
            "U": 28,
            "V": 60,
        }

        for column, width in widths.items():
            worksheet.column_dimensions[
                column
            ].width = width

        worksheet.auto_filter.ref = (
            f"A{header_row}:V{last_data_row}"
        )

        worksheet.print_title_rows = (
            f"1:{header_row}"
        )

        worksheet.print_area = (
            f"A1:V{last_data_row}"
        )

    @staticmethod
    def _deduction_formula(
        row: int,
    ) -> str:
        return (
            f'=IF(OR('
            f'I{row}="MET",'
            f'I{row}="NOT APPLICABLE",'
            f'I{row}="NOT ASSESSED"),'
            f'0,'
            f'IF(AND('
            f'K{row}="Yes",'
            f'J{row}="PARTIALLY IMPLEMENTED"),'
            f'3,'
            f'L{row}))'
        )

    @staticmethod
    def _explanation_formula(
        row: int,
    ) -> str:
        return (
            f'=IF(I{row}="MET",'
            f'"Requirement assessed MET; no deduction.",'
            f'IF(I{row}="NOT APPLICABLE",'
            f'"Requirement assessed NOT APPLICABLE; '
            f'no deduction.",'
            f'IF(I{row}="NOT ASSESSED",'
            f'"Requirement not assessed; score remains '
            f'provisional.",'
            f'IF(M{row}="Yes",'
            f'"Partial-credit condition applied; '
            f'3 points deducted.",'
            f'"Requirement assessed NOT MET; "&'
            f'L{row}&" points deducted."))))'
        )

    def _add_assessment_validations(
        self,
        worksheet: Worksheet,
        first_row: int,
        last_row: int,
    ) -> None:
        applicable_validation = DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False,
        )

        worksheet.add_data_validation(
            applicable_validation
        )

        applicable_validation.add(
            f"G{first_row}:G{last_row}"
        )

        status_validation = DataValidation(
            type="list",
            formula1=(
                '"MET,NOT MET,'
                'NOT APPLICABLE,'
                'NOT ASSESSED"'
            ),
            allow_blank=False,
        )

        worksheet.add_data_validation(
            status_validation
        )

        status_validation.add(
            f"I{first_row}:I{last_row}"
        )

        implementation_validation = (
            DataValidation(
                type="list",
                formula1=(
                    '"FULLY IMPLEMENTED,'
                    'PARTIALLY IMPLEMENTED,'
                    'NOT IMPLEMENTED,'
                    'NOT APPLICABLE,'
                    'NOT ASSESSED"'
                ),
                allow_blank=True,
            )
        )

        worksheet.add_data_validation(
            implementation_validation
        )

        implementation_validation.add(
            f"J{first_row}:J{last_row}"
        )

        evidence_validation = DataValidation(
            type="list",
            formula1=(
                '"Not Started,In Progress,'
                'Complete,Not Applicable"'
            ),
            allow_blank=False,
        )

        worksheet.add_data_validation(
            evidence_validation
        )

        evidence_validation.add(
            f"O{first_row}:O{last_row}"
        )

    def _add_assessment_conditional_formatting(
        self,
        worksheet: Worksheet,
        first_row: int,
        last_row: int,
    ) -> None:
        status_range = (
            f"I{first_row}:I{last_row}"
        )

        worksheet.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=[
                    f'I{first_row}="MET"'
                ],
                fill=self.styles.good_fill(),
            ),
        )

        worksheet.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=[
                    f'I{first_row}="NOT MET"'
                ],
                fill=self.styles.bad_fill(),
            ),
        )

        worksheet.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=[
                    (
                        f'I{first_row}'
                        '="NOT ASSESSED"'
                    )
                ],
                fill=self.styles.warning_fill(),
            ),
        )

        partial_range = (
            f"M{first_row}:M{last_row}"
        )

        worksheet.conditional_formatting.add(
            partial_range,
            FormulaRule(
                formula=[
                    f'M{first_row}="Yes"'
                ],
                fill=self.styles.warning_fill(),
            ),
        )

        poam_range = (
            f"S{first_row}:S{last_row}"
        )

        worksheet.conditional_formatting.add(
            poam_range,
            FormulaRule(
                formula=[
                    f'S{first_row}="Yes"'
                ],
                fill=self.styles.bad_fill(),
            ),
        )

    def _build_domain_summary(
        self,
        worksheet: Worksheet,
        controls: List[Dict[str, str]],
    ) -> None:
        self.factory.configure_standard_sheet(
            worksheet,
            freeze_cell="A6",
        )

        self.factory.create_title_band(
            worksheet,
            title="Domain Summary",
            subtitle=(
                "Weighted scoring, completion, and findings "
                "by CMMC domain"
            ),
            end_column=10,
        )

        headers = [
            "Domain",
            "Requirement Count",
            "Met",
            "Not Met",
            "Not Assessed",
            "Full Weight",
            "Current Deduction",
            "Completion %",
            "Domain Score",
            "Readiness",
        ]

        for column, header in enumerate(
            headers,
            start=1,
        ):
            worksheet.cell(
                row=5,
                column=column,
                value=header,
            )

        self.factory.style_table_header(
            worksheet,
            row_number=5,
            start_column=1,
            end_column=len(headers),
        )

        domain_codes: List[str] = []

        for control in controls:
            domain_code = (
                control["domain_code"]
                .strip()
                .upper()
            )

            if domain_code not in domain_codes:
                domain_codes.append(
                    domain_code
                )

        first_row = self.FIRST_ASSESSMENT_ROW
        last_row = (
            first_row
            + self.EXPECTED_CONTROL_COUNT
            - 1
        )

        for offset, domain_code in enumerate(
            domain_codes
        ):
            row = 6 + offset

            worksheet.cell(
                row=row,
                column=1,
                value=domain_code,
            )

            worksheet.cell(
                row=row,
                column=2,
                value=(
                    '=COUNTIF('
                    f'Assessment!$A${first_row}:'
                    f'$A${last_row},A{row})'
                ),
            )

            worksheet.cell(
                row=row,
                column=3,
                value=(
                    '=COUNTIFS('
                    f'Assessment!$A${first_row}:'
                    f'$A${last_row},A{row},'
                    f'Assessment!$I${first_row}:'
                    f'$I${last_row},"MET")'
                ),
            )

            worksheet.cell(
                row=row,
                column=4,
                value=(
                    '=COUNTIFS('
                    f'Assessment!$A${first_row}:'
                    f'$A${last_row},A{row},'
                    f'Assessment!$I${first_row}:'
                    f'$I${last_row},"NOT MET")'
                ),
            )

            worksheet.cell(
                row=row,
                column=5,
                value=(
                    '=COUNTIFS('
                    f'Assessment!$A${first_row}:'
                    f'$A${last_row},A{row},'
                    f'Assessment!$I${first_row}:'
                    f'$I${last_row},"NOT ASSESSED")'
                ),
            )

            worksheet.cell(
                row=row,
                column=6,
                value=(
                    '=SUMIF('
                    f'Assessment!$A${first_row}:'
                    f'$A${last_row},A{row},'
                    f'Assessment!$L${first_row}:'
                    f'$L${last_row})'
                ),
            )

            worksheet.cell(
                row=row,
                column=7,
                value=(
                    '=SUMIF('
                    f'Assessment!$A${first_row}:'
                    f'$A${last_row},A{row},'
                    f'Assessment!$N${first_row}:'
                    f'$N${last_row})'
                ),
            )

            worksheet.cell(
                row=row,
                column=8,
                value=(
                    f'=IF(B{row}=0,0,'
                    f'(B{row}-E{row})/B{row})'
                ),
            )

            worksheet.cell(
                row=row,
                column=9,
                value=(
                    f'=F{row}-G{row}'
                ),
            )

            worksheet.cell(
                row=row,
                column=10,
                value=(
                    f'=IF(E{row}>0,'
                    f'"INCOMPLETE",'
                    f'IF(D{row}=0,'
                    f'"READY","NEEDS WORK"))'
                ),
            )

            for column in range(1, 11):
                cell = worksheet.cell(
                    row=row,
                    column=column,
                )

                cell.border = (
                    self.styles.thin_border()
                )
                cell.alignment = (
                    self.styles.center_alignment()
                )
                cell.font = (
                    self.styles.body_font()
                )

            worksheet.cell(
                row=row,
                column=8,
            ).number_format = "0%"

        for column in range(1, 11):
            column_letter = (
                get_column_letter(column)
            )
            worksheet.column_dimensions[
                column_letter
            ].width = 18

        last_domain_row = (
            5 + len(domain_codes)
        )

        worksheet.auto_filter.ref = (
            f"A5:J{last_domain_row}"
        )

    def _build_settings(
        self,
        worksheet: Worksheet,
        scoring_rules: Dict[
            str,
            WorkbookScoringRule,
        ],
    ) -> None:
        self.factory.configure_standard_sheet(
            worksheet
        )

        self.factory.create_title_band(
            worksheet,
            title="Workbook Settings",
            subtitle=(
                "Workbook configuration and official "
                "scoring metadata"
            ),
            end_column=8,
        )

        maximum_deduction = sum(
            rule.full_deduction_points
            for rule in scoring_rules.values()
        )

        minimum_score = (
            self.MAXIMUM_SCORE
            - maximum_deduction
        )

        partial_rule_count = sum(
            rule.partial_credit_allowed
            for rule in scoring_rules.values()
        )

        settings = [
            ("Workbook Version", "0.2"),
            (
                "Assessment Guide Version",
                "2.13",
            ),
            (
                "Maximum Score",
                self.MAXIMUM_SCORE,
            ),
            (
                "Maximum Deduction",
                maximum_deduction,
            ),
            (
                "Mathematical Minimum Score",
                minimum_score,
            ),
            (
                "Requirement Count",
                len(scoring_rules),
            ),
            (
                "Partial Credit Requirements",
                partial_rule_count,
            ),
            (
                "Assessment Level",
                "CMMC Level 2",
            ),
            (
                "Scoring Source",
                "32 CFR 170.24",
            ),
            (
                "Workbook Generated By",
                "Omni by R!SC",
            ),
        ]

        for offset, (name, value) in enumerate(
            settings
        ):
            row = 6 + offset

            name_cell = worksheet.cell(
                row=row,
                column=2,
                value=name,
            )

            value_cell = worksheet.cell(
                row=row,
                column=3,
                value=value,
            )

            name_cell.fill = (
                self.styles.section_fill()
            )
            name_cell.font = (
                self.styles.header_font()
            )
            name_cell.border = (
                self.styles.thin_border()
            )

            value_cell.fill = (
                self.styles.formula_fill()
            )
            value_cell.font = (
                self.styles.body_font()
            )
            value_cell.border = (
                self.styles.thin_border()
            )

        worksheet.column_dimensions[
            "B"
        ].width = 34

        worksheet.column_dimensions[
            "C"
        ].width = 42

    def _build_lists(
        self,
        worksheet: Worksheet,
    ) -> None:
        lists = {
            "A": (
                "Assessment Status",
                [
                    "MET",
                    "NOT MET",
                    "NOT APPLICABLE",
                    "NOT ASSESSED",
                ],
            ),
            "B": (
                "Evidence Status",
                [
                    "Not Started",
                    "In Progress",
                    "Complete",
                    "Not Applicable",
                ],
            ),
            "C": (
                "Yes / No",
                [
                    "Yes",
                    "No",
                ],
            ),
            "D": (
                "Implementation State",
                [
                    "FULLY IMPLEMENTED",
                    "PARTIALLY IMPLEMENTED",
                    "NOT IMPLEMENTED",
                    "NOT APPLICABLE",
                    "NOT ASSESSED",
                ],
            ),
        }

        for column, (
            heading,
            values,
        ) in lists.items():
            worksheet[f"{column}1"] = heading

            for row, value in enumerate(
                values,
                start=2,
            ):
                worksheet[
                    f"{column}{row}"
                ] = value

    def _build_placeholder_sheet(
        self,
        worksheet: Worksheet,
        title: str,
        subtitle: str,
    ) -> None:
        self.factory.configure_standard_sheet(
            worksheet
        )

        self.factory.create_title_band(
            worksheet,
            title=title,
            subtitle=subtitle,
            end_column=10,
        )

        worksheet.merge_cells(
            start_row=6,
            start_column=2,
            end_row=8,
            end_column=8,
        )

        placeholder_cell = worksheet["B6"]

        placeholder_cell.value = (
            "This module will be implemented in "
            "the next workbook sprint."
        )

        placeholder_cell.font = (
            self.styles.body_font()
        )
        placeholder_cell.fill = (
            self.styles.subheader_fill()
        )
        placeholder_cell.border = (
            self.styles.thin_border()
        )
        placeholder_cell.alignment = (
            self.styles.center_alignment()
        )

        for column in range(1, 11):
            column_letter = (
                get_column_letter(column)
            )

            worksheet.column_dimensions[
                column_letter
            ].width = 14
