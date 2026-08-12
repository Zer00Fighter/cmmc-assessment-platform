from __future__ import annotations

from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from src.workbook.workbook_styles import WorkbookStyles
from src.workbook.worksheet_factory import WorksheetFactory


class POAMSheetBuilder:
    """Build the Remediation Action Plan (POA&M) worksheet."""

    HEADER_ROW = 5
    FIRST_DATA_ROW = 6
    MAX_POAM_ROWS = 300

    HEADERS = [
        "Remediation ID (POA&M ID)",
        "Requirement ID",
        "Requirement Title",
        "Domain",
        "Weakness Description",
        "Root Cause",
        "Corrective Action",
        "Current Milestone",
        "Milestone Owner",
        "Status",
        "Priority",
        "Severity",
        "Likelihood",
        "Risk Score",
        "Date Identified",
        "Planned Completion",
        "Actual Completion",
        "Days Open",
        "Aging Bucket",
        "Residual Risk",
        "Validation Status",
        "Evidence IDs",
        "Security Plan Reference (SSP)",
        "Assessor Notes",
    ]

    def __init__(
        self,
        styles: WorkbookStyles,
        factory: WorksheetFactory,
    ) -> None:
        self.styles = styles
        self.factory = factory

    @property
    def last_data_row(self) -> int:
        return self.FIRST_DATA_ROW + self.MAX_POAM_ROWS - 1

    def build(
        self,
        worksheet: Worksheet,
    ) -> None:
        self.factory.configure_standard_sheet(
            worksheet,
            freeze_cell="A6",
            show_gridlines=False,
            zoom_scale=75,
        )

        self.factory.create_title_band(
            worksheet,
            title="Remediation Action Plan (POA&M)",
            subtitle=(
                "Track remediation actions, ownership, milestones, "
                "risk, aging, validation, and closure for CMMC findings."
            ),
            end_column=len(self.HEADERS),
        )

        self._write_headers(worksheet)
        self._create_rows(worksheet)
        self._add_validations(worksheet)
        self._add_conditional_formatting(worksheet)
        self._configure_columns(worksheet)
        self._configure_printing(worksheet)

    def _write_headers(
        self,
        worksheet: Worksheet,
    ) -> None:
        for column, header in enumerate(
            self.HEADERS,
            start=1,
        ):
            worksheet.cell(
                row=self.HEADER_ROW,
                column=column,
                value=header,
            )

        self.factory.style_table_header(
            worksheet,
            row_number=self.HEADER_ROW,
            start_column=1,
            end_column=len(self.HEADERS),
        )

    def _create_rows(
        self,
        worksheet: Worksheet,
    ) -> None:
        for row in range(
            self.FIRST_DATA_ROW,
            self.last_data_row + 1,
        ):
            worksheet.cell(
                row=row,
                column=1,
                value="",
            )

            worksheet.cell(
                row=row,
                column=2,
                value="",
            )

            worksheet.cell(
                row=row,
                column=3,
                value="",
            )

            worksheet.cell(
                row=row,
                column=4,
                value="",
            )

            worksheet.cell(
                row=row,
                column=5,
                value="",
            )

            worksheet.cell(
                row=row,
                column=6,
                value="",
            )

            worksheet.cell(
                row=row,
                column=7,
                value="",
            )

            worksheet.cell(
                row=row,
                column=8,
                value="",
            )

            worksheet.cell(
                row=row,
                column=9,
                value="",
            )

            worksheet.cell(
                row=row,
                column=10,
                value="Open",
            )

            worksheet.cell(
                row=row,
                column=11,
                value="Medium",
            )

            worksheet.cell(
                row=row,
                column=12,
                value="Medium",
            )

            worksheet.cell(
                row=row,
                column=13,
                value="Possible",
            )

            worksheet.cell(
                row=row,
                column=14,
                value=(
                    f'=IF(OR(L{row}="",M{row}=""),"",'
                    f'IF(L{row}="Critical",4,'
                    f'IF(L{row}="High",3,'
                    f'IF(L{row}="Medium",2,1)))'
                    f"*"
                    f'IF(M{row}="Almost Certain",5,'
                    f'IF(M{row}="Likely",4,'
                    f'IF(M{row}="Possible",3,'
                    f'IF(M{row}="Unlikely",2,1)))))'
                ),
            )

            worksheet.cell(
                row=row,
                column=15,
                value="",
            )

            worksheet.cell(
                row=row,
                column=16,
                value="",
            )

            worksheet.cell(
                row=row,
                column=17,
                value="",
            )

            worksheet.cell(
                row=row,
                column=18,
                value=(
                    f'=IF(O{row}="","",' f'IF(Q{row}<>"",Q{row}-O{row},TODAY()-O{row}))'
                ),
            )

            worksheet.cell(
                row=row,
                column=19,
                value=(
                    f'=IF(R{row}="","",'
                    f'IF(R{row}<=30,"0-30 Days",'
                    f'IF(R{row}<=60,"31-60 Days",'
                    f'IF(R{row}<=90,"61-90 Days",'
                    f'IF(R{row}<=180,"91-180 Days","181+ Days")))))'
                ),
            )

            worksheet.cell(
                row=row,
                column=20,
                value="Medium",
            )

            worksheet.cell(
                row=row,
                column=21,
                value="Pending",
            )

            worksheet.cell(
                row=row,
                column=22,
                value="",
            )

            worksheet.cell(
                row=row,
                column=23,
                value="",
            )

            worksheet.cell(
                row=row,
                column=24,
                value="",
            )

            for column in range(
                1,
                len(self.HEADERS) + 1,
            ):
                cell = worksheet.cell(
                    row=row,
                    column=column,
                )

                cell.font = self.styles.body_font()
                cell.border = self.styles.thin_border()
                cell.alignment = self.styles.left_alignment()

                if column in {
                    3,
                    4,
                    14,
                    18,
                    19,
                }:
                    cell.fill = self.styles.formula_fill()
                    cell.protection = self.styles.locked_protection()
                else:
                    cell.fill = self.styles.input_fill()
                    cell.protection = self.styles.unlocked_protection()

            for column in {
                15,
                16,
                17,
            }:
                worksheet.cell(
                    row=row,
                    column=column,
                ).number_format = "mm/dd/yyyy"

            worksheet.cell(
                row=row,
                column=14,
            ).number_format = "0"

            worksheet.cell(
                row=row,
                column=18,
            ).number_format = "0"

    def _add_validations(
        self,
        worksheet: Worksheet,
    ) -> None:
        status_validation = DataValidation(
            type="list",
            formula1=(
                '"Open,In Progress,On Hold,Deferred,' 'Completed,Closed,Risk Accepted"'
            ),
            allow_blank=False,
        )

        worksheet.add_data_validation(status_validation)

        status_validation.add(f"J{self.FIRST_DATA_ROW}:" f"J{self.last_data_row}")

        priority_validation = DataValidation(
            type="list",
            formula1=('"Critical,High,Medium,Low"'),
            allow_blank=False,
        )

        worksheet.add_data_validation(priority_validation)

        priority_validation.add(f"K{self.FIRST_DATA_ROW}:" f"K{self.last_data_row}")

        severity_validation = DataValidation(
            type="list",
            formula1=('"Critical,High,Medium,Low"'),
            allow_blank=False,
        )

        worksheet.add_data_validation(severity_validation)

        severity_validation.add(f"L{self.FIRST_DATA_ROW}:" f"L{self.last_data_row}")

        likelihood_validation = DataValidation(
            type="list",
            formula1=('"Almost Certain,Likely,Possible,' 'Unlikely,Rare"'),
            allow_blank=False,
        )

        worksheet.add_data_validation(likelihood_validation)

        likelihood_validation.add(f"M{self.FIRST_DATA_ROW}:" f"M{self.last_data_row}")

        residual_risk_validation = DataValidation(
            type="list",
            formula1=('"Critical,High,Medium,Low,Accepted"'),
            allow_blank=False,
        )

        worksheet.add_data_validation(residual_risk_validation)

        residual_risk_validation.add(
            f"T{self.FIRST_DATA_ROW}:" f"T{self.last_data_row}"
        )

        validation_status_validation = DataValidation(
            type="list",
            formula1=('"Pending,In Review,Verified,Rejected,' 'Accepted,Not Required"'),
            allow_blank=False,
        )

        worksheet.add_data_validation(validation_status_validation)

        validation_status_validation.add(
            f"U{self.FIRST_DATA_ROW}:" f"U{self.last_data_row}"
        )

        date_validation = DataValidation(
            type="date",
            operator="between",
            formula1="DATE(2000,1,1)",
            formula2="DATE(2100,12,31)",
            allow_blank=True,
        )

        worksheet.add_data_validation(date_validation)

        date_validation.add(f"O{self.FIRST_DATA_ROW}:" f"Q{self.last_data_row}")

    def _add_conditional_formatting(
        self,
        worksheet: Worksheet,
    ) -> None:
        status_range = f"J{self.FIRST_DATA_ROW}:" f"J{self.last_data_row}"

        worksheet.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=[f'J{self.FIRST_DATA_ROW}="Closed"'],
                fill=self.styles.good_fill(),
            ),
        )

        worksheet.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=[f'J{self.FIRST_DATA_ROW}="Completed"'],
                fill=self.styles.good_fill(),
            ),
        )

        worksheet.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=[f'J{self.FIRST_DATA_ROW}="In Progress"'],
                fill=self.styles.warning_fill(),
            ),
        )

        worksheet.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=[f'J{self.FIRST_DATA_ROW}="Open"'],
                fill=self.styles.bad_fill(),
            ),
        )

        risk_range = f"N{self.FIRST_DATA_ROW}:" f"N{self.last_data_row}"

        worksheet.conditional_formatting.add(
            risk_range,
            FormulaRule(
                formula=[f"N{self.FIRST_DATA_ROW}>=15"],
                fill=self.styles.bad_fill(),
            ),
        )

        worksheet.conditional_formatting.add(
            risk_range,
            FormulaRule(
                formula=[
                    f"AND(N{self.FIRST_DATA_ROW}>=8," f"N{self.FIRST_DATA_ROW}<15)"
                ],
                fill=self.styles.warning_fill(),
            ),
        )

        overdue_range = f"P{self.FIRST_DATA_ROW}:" f"P{self.last_data_row}"

        worksheet.conditional_formatting.add(
            overdue_range,
            FormulaRule(
                formula=[
                    (
                        f"AND(P{self.FIRST_DATA_ROW}<TODAY(),"
                        f'P{self.FIRST_DATA_ROW}<>"",'
                        f'J{self.FIRST_DATA_ROW}<>"Closed",'
                        f'J{self.FIRST_DATA_ROW}<>"Completed")'
                    )
                ],
                fill=self.styles.bad_fill(),
            ),
        )

        aging_range = f"S{self.FIRST_DATA_ROW}:" f"S{self.last_data_row}"

        worksheet.conditional_formatting.add(
            aging_range,
            FormulaRule(
                formula=[f'S{self.FIRST_DATA_ROW}="181+ Days"'],
                fill=self.styles.bad_fill(),
            ),
        )

        worksheet.conditional_formatting.add(
            aging_range,
            FormulaRule(
                formula=[
                    (
                        f'OR(S{self.FIRST_DATA_ROW}="61-90 Days",'
                        f'S{self.FIRST_DATA_ROW}="91-180 Days")'
                    )
                ],
                fill=self.styles.warning_fill(),
            ),
        )

    def _configure_columns(
        self,
        worksheet: Worksheet,
    ) -> None:
        widths = {
            "A": 16,
            "B": 18,
            "C": 36,
            "D": 10,
            "E": 48,
            "F": 42,
            "G": 48,
            "H": 36,
            "I": 22,
            "J": 18,
            "K": 14,
            "L": 14,
            "M": 18,
            "N": 12,
            "O": 15,
            "P": 18,
            "Q": 18,
            "R": 12,
            "S": 16,
            "T": 16,
            "U": 18,
            "V": 28,
            "W": 24,
            "X": 45,
        }

        for column, width in widths.items():
            worksheet.column_dimensions[column].width = width

        worksheet.auto_filter.ref = f"A{self.HEADER_ROW}:" f"X{self.last_data_row}"

    def _configure_printing(
        self,
        worksheet: Worksheet,
    ) -> None:
        worksheet.print_title_rows = f"1:{self.HEADER_ROW}"

        worksheet.print_area = f"A1:X{self.last_data_row}"

        worksheet.page_setup.orientation = "landscape"

        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
