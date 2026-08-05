from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from src.workbook.workbook_styles import WorkbookStyles


class WorksheetFactory:
    """Common worksheet configuration and formatting utilities."""

    def __init__(self, styles: WorkbookStyles) -> None:
        self.styles = styles

    def configure_standard_sheet(
        self,
        worksheet: Worksheet,
        *,
        freeze_cell: str | None = None,
        show_gridlines: bool = False,
        zoom_scale: int = 90,
    ) -> None:
        worksheet.sheet_view.showGridLines = show_gridlines
        worksheet.sheet_view.zoomScale = zoom_scale

        if freeze_cell:
            worksheet.freeze_panes = freeze_cell

        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.page_setup.orientation = "landscape"
        worksheet.page_margins.left = 0.25
        worksheet.page_margins.right = 0.25
        worksheet.page_margins.top = 0.5
        worksheet.page_margins.bottom = 0.5

    def create_title_band(
        self,
        worksheet: Worksheet,
        title: str,
        subtitle: str = "",
        *,
        start_column: int = 1,
        end_column: int = 10,
    ) -> None:
        worksheet.merge_cells(
            start_row=1,
            start_column=start_column,
            end_row=2,
            end_column=end_column,
        )

        title_cell = worksheet.cell(
            row=1,
            column=start_column,
        )

        title_cell.value = title
        title_cell.font = self.styles.title_font()
        title_cell.fill = self.styles.title_fill()
        title_cell.alignment = self.styles.left_alignment()

        for row in range(1, 3):
            for column in range(
                start_column,
                end_column + 1,
            ):
                worksheet.cell(
                    row=row,
                    column=column,
                ).fill = self.styles.title_fill()

        if subtitle:
            worksheet.merge_cells(
                start_row=3,
                start_column=start_column,
                end_row=3,
                end_column=end_column,
            )

            subtitle_cell = worksheet.cell(
                row=3,
                column=start_column,
            )

            subtitle_cell.value = subtitle
            subtitle_cell.font = self.styles.subtitle_font()
            subtitle_cell.fill = self.styles.title_fill()
            subtitle_cell.alignment = self.styles.left_alignment()

            for column in range(
                start_column,
                end_column + 1,
            ):
                worksheet.cell(
                    row=3,
                    column=column,
                ).fill = self.styles.title_fill()

        worksheet.row_dimensions[1].height = 26
        worksheet.row_dimensions[2].height = 14
        worksheet.row_dimensions[3].height = 22

    def style_table_header(
        self,
        worksheet: Worksheet,
        row_number: int,
        start_column: int,
        end_column: int,
    ) -> None:
        for column in range(
            start_column,
            end_column + 1,
        ):
            cell = worksheet.cell(
                row=row_number,
                column=column,
            )

            cell.font = self.styles.header_font()
            cell.fill = self.styles.header_fill()
            cell.border = self.styles.thin_border()
            cell.alignment = self.styles.center_alignment()

        worksheet.row_dimensions[row_number].height = 32