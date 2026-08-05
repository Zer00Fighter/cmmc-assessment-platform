from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, Iterable, List, Mapping

from openpyxl.worksheet.worksheet import Worksheet

from src.workbook.assessment_sync import (
    AssessmentRecord,
)
from src.workbook.poam_tracker import (
    POAMEntry,
    POAMTracker,
)


class POAMSyncError(Exception):
    """Raised when POA&M worksheet synchronization fails."""


@dataclass(frozen=True)
class POAMSyncResult:
    assessment_record_count: int
    not_met_count: int
    existing_poam_count: int
    created_count: int
    updated_count: int
    cleared_count: int
    final_poam_count: int


class POAMWorksheetSynchronizer:
    """
    Synchronize Assessment records into the POA&M worksheet.

    The synchronizer:

    - identifies NOT MET requirements;
    - creates or preserves POA&M entries;
    - writes requirement metadata into the POA&M worksheet;
    - preserves user-entered remediation fields;
    - clears obsolete generated entries;
    - prevents duplicate requirement mappings.
    """

    FIRST_DATA_ROW = 6
    MAX_POAM_ROWS = 300

    POAM_ID_COLUMN = 1
    REQUIREMENT_ID_COLUMN = 2
    TITLE_COLUMN = 3
    DOMAIN_COLUMN = 4
    WEAKNESS_COLUMN = 5
    ROOT_CAUSE_COLUMN = 6
    CORRECTIVE_ACTION_COLUMN = 7
    MILESTONE_COLUMN = 8
    OWNER_COLUMN = 9
    STATUS_COLUMN = 10
    PRIORITY_COLUMN = 11
    SEVERITY_COLUMN = 12
    LIKELIHOOD_COLUMN = 13
    RISK_SCORE_COLUMN = 14
    DATE_IDENTIFIED_COLUMN = 15
    PLANNED_COMPLETION_COLUMN = 16
    ACTUAL_COMPLETION_COLUMN = 17
    DAYS_OPEN_COLUMN = 18
    AGING_COLUMN = 19
    RESIDUAL_RISK_COLUMN = 20
    VALIDATION_STATUS_COLUMN = 21
    EVIDENCE_IDS_COLUMN = 22
    SSP_REFERENCE_COLUMN = 23
    NOTES_COLUMN = 24

    PRESERVED_COLUMNS = {
        WEAKNESS_COLUMN,
        ROOT_CAUSE_COLUMN,
        CORRECTIVE_ACTION_COLUMN,
        MILESTONE_COLUMN,
        OWNER_COLUMN,
        STATUS_COLUMN,
        PRIORITY_COLUMN,
        SEVERITY_COLUMN,
        LIKELIHOOD_COLUMN,
        DATE_IDENTIFIED_COLUMN,
        PLANNED_COMPLETION_COLUMN,
        ACTUAL_COMPLETION_COLUMN,
        RESIDUAL_RISK_COLUMN,
        VALIDATION_STATUS_COLUMN,
        EVIDENCE_IDS_COLUMN,
        SSP_REFERENCE_COLUMN,
        NOTES_COLUMN,
    }

    def __init__(
        self,
        tracker: POAMTracker | None = None,
    ) -> None:
        self.tracker = tracker or POAMTracker()

    @property
    def last_data_row(self) -> int:
        return (
            self.FIRST_DATA_ROW
            + self.MAX_POAM_ROWS
            - 1
        )

    def synchronize(
        self,
        worksheet: Worksheet,
        assessment_records: Iterable[AssessmentRecord],
    ) -> POAMSyncResult:
        records = list(assessment_records)

        assessment_rows = [
            self._to_tracker_row(record)
            for record in records
        ]

        not_met_records = [
            record
            for record in records
            if self._normalize_status(
                record.status
            )
            == "NOT MET"
        ]

        existing_rows = self._read_existing_rows(
            worksheet
        )

        existing_poam_count = len(existing_rows)

        tracker_entries = self.tracker.synchronize(
            assessment_rows
        )

        if len(tracker_entries) > self.MAX_POAM_ROWS:
            raise POAMSyncError(
                "POA&M synchronization produced "
                f"{len(tracker_entries)} entries, but the "
                f"worksheet supports only {self.MAX_POAM_ROWS}."
            )

        created_count = 0
        updated_count = 0

        active_requirement_ids = {
            entry.requirement_id
            for entry in tracker_entries
        }

        row_assignments = self._assign_rows(
            tracker_entries=tracker_entries,
            existing_rows=existing_rows,
        )

        for entry in tracker_entries:
            row = row_assignments[
                entry.requirement_id
            ]

            existing = existing_rows.get(
                entry.requirement_id
            )

            if existing is None:
                created_count += 1
            else:
                updated_count += 1

            record = next(
                item
                for item in not_met_records
                if item.requirement_id
                == entry.requirement_id
            )

            self._write_entry(
                worksheet=worksheet,
                row=row,
                entry=entry,
                record=record,
                existing_values=(
                    existing["values"]
                    if existing is not None
                    else {}
                ),
            )

        cleared_count = 0

        for requirement_id, existing in (
            existing_rows.items()
        ):
            if requirement_id in active_requirement_ids:
                continue

            self._clear_generated_row(
                worksheet=worksheet,
                row=existing["row"],
            )

            cleared_count += 1

        return POAMSyncResult(
            assessment_record_count=len(records),
            not_met_count=len(not_met_records),
            existing_poam_count=existing_poam_count,
            created_count=created_count,
            updated_count=updated_count,
            cleared_count=cleared_count,
            final_poam_count=len(tracker_entries),
        )

    def _read_existing_rows(
        self,
        worksheet: Worksheet,
    ) -> Dict[str, Dict[str, object]]:
        existing: Dict[
            str,
            Dict[str, object],
        ] = {}

        for row in range(
            self.FIRST_DATA_ROW,
            self.last_data_row + 1,
        ):
            requirement_value = worksheet.cell(
                row=row,
                column=self.REQUIREMENT_ID_COLUMN,
            ).value

            if requirement_value in {
                None,
                "",
            }:
                continue

            requirement_id = (
                self._normalize_requirement_id(
                    str(requirement_value)
                )
            )

            if requirement_id in existing:
                raise POAMSyncError(
                    "Duplicate POA&M worksheet entries exist "
                    f"for requirement {requirement_id}."
                )

            values = {
                column: worksheet.cell(
                    row=row,
                    column=column,
                ).value
                for column in self.PRESERVED_COLUMNS
            }

            existing[requirement_id] = {
                "row": row,
                "values": values,
                "poam_id": worksheet.cell(
                    row=row,
                    column=self.POAM_ID_COLUMN,
                ).value,
            }

        return existing

    def _assign_rows(
        self,
        tracker_entries: List[POAMEntry],
        existing_rows: Mapping[
            str,
            Dict[str, object],
        ],
    ) -> Dict[str, int]:
        assignments: Dict[str, int] = {}

        occupied_rows = {
            int(existing["row"])
            for existing in existing_rows.values()
            if existing["row"] is not None
        }

        for entry in tracker_entries:
            existing = existing_rows.get(
                entry.requirement_id
            )

            if existing is not None:
                assignments[
                    entry.requirement_id
                ] = int(existing["row"])

        available_rows = [
            row
            for row in range(
                self.FIRST_DATA_ROW,
                self.last_data_row + 1,
            )
            if row not in occupied_rows
        ]

        for entry in tracker_entries:
            if entry.requirement_id in assignments:
                continue

            if not available_rows:
                raise POAMSyncError(
                    "No available POA&M worksheet rows remain."
                )

            assignments[
                entry.requirement_id
            ] = available_rows.pop(0)

        return assignments

    def _write_entry(
        self,
        worksheet: Worksheet,
        row: int,
        entry: POAMEntry,
        record: AssessmentRecord,
        existing_values: Mapping[int, object],
    ) -> None:
        worksheet.cell(
            row=row,
            column=self.POAM_ID_COLUMN,
            value=entry.poam_id,
        )

        worksheet.cell(
            row=row,
            column=self.REQUIREMENT_ID_COLUMN,
            value=entry.requirement_id,
        )

        worksheet.cell(
            row=row,
            column=self.TITLE_COLUMN,
            value=record.title,
        )

        worksheet.cell(
            row=row,
            column=self.DOMAIN_COLUMN,
            value=record.domain,
        )

        worksheet.cell(
            row=row,
            column=self.WEAKNESS_COLUMN,
            value=self._preserved_or_default(
                existing_values,
                self.WEAKNESS_COLUMN,
                (
                    f"{entry.requirement_id} is assessed "
                    "NOT MET."
                ),
            ),
        )

        worksheet.cell(
            row=row,
            column=self.ROOT_CAUSE_COLUMN,
            value=self._preserved_or_default(
                existing_values,
                self.ROOT_CAUSE_COLUMN,
                "",
            ),
        )

        worksheet.cell(
            row=row,
            column=self.CORRECTIVE_ACTION_COLUMN,
            value=self._preserved_or_default(
                existing_values,
                self.CORRECTIVE_ACTION_COLUMN,
                entry.corrective_action,
            ),
        )

        worksheet.cell(
            row=row,
            column=self.MILESTONE_COLUMN,
            value=self._preserved_or_default(
                existing_values,
                self.MILESTONE_COLUMN,
                entry.milestone,
            ),
        )

        worksheet.cell(
            row=row,
            column=self.OWNER_COLUMN,
            value=self._preserved_or_default(
                existing_values,
                self.OWNER_COLUMN,
                record.owner or entry.owner,
            ),
        )

        worksheet.cell(
            row=row,
            column=self.STATUS_COLUMN,
            value=self._preserved_or_default(
                existing_values,
                self.STATUS_COLUMN,
                entry.status,
            ),
        )

        worksheet.cell(
            row=row,
            column=self.PRIORITY_COLUMN,
            value=self._preserved_or_default(
                existing_values,
                self.PRIORITY_COLUMN,
                "Medium",
            ),
        )

        worksheet.cell(
            row=row,
            column=self.SEVERITY_COLUMN,
            value=self._preserved_or_default(
                existing_values,
                self.SEVERITY_COLUMN,
                "Medium",
            ),
        )

        worksheet.cell(
            row=row,
            column=self.LIKELIHOOD_COLUMN,
            value=self._preserved_or_default(
                existing_values,
                self.LIKELIHOOD_COLUMN,
                "Possible",
            ),
        )

        worksheet.cell(
            row=row,
            column=self.RISK_SCORE_COLUMN,
            value=self._risk_formula(row),
        )

        worksheet.cell(
            row=row,
            column=self.DATE_IDENTIFIED_COLUMN,
            value=self._preserved_or_default(
                existing_values,
                self.DATE_IDENTIFIED_COLUMN,
                "",
            ),
        )

        worksheet.cell(
            row=row,
            column=self.PLANNED_COMPLETION_COLUMN,
            value=self._preserved_or_default(
                existing_values,
                self.PLANNED_COMPLETION_COLUMN,
                "",
            ),
        )

        worksheet.cell(
            row=row,
            column=self.ACTUAL_COMPLETION_COLUMN,
            value=self._preserved_or_default(
                existing_values,
                self.ACTUAL_COMPLETION_COLUMN,
                "",
            ),
        )

        worksheet.cell(
            row=row,
            column=self.DAYS_OPEN_COLUMN,
            value=self._days_open_formula(row),
        )

        worksheet.cell(
            row=row,
            column=self.AGING_COLUMN,
            value=self._aging_formula(row),
        )

        worksheet.cell(
            row=row,
            column=self.RESIDUAL_RISK_COLUMN,
            value=self._preserved_or_default(
                existing_values,
                self.RESIDUAL_RISK_COLUMN,
                "Medium",
            ),
        )

        worksheet.cell(
            row=row,
            column=self.VALIDATION_STATUS_COLUMN,
            value=self._preserved_or_default(
                existing_values,
                self.VALIDATION_STATUS_COLUMN,
                "Pending",
            ),
        )

        worksheet.cell(
            row=row,
            column=self.EVIDENCE_IDS_COLUMN,
            value=self._preserved_or_default(
                existing_values,
                self.EVIDENCE_IDS_COLUMN,
                "",
            ),
        )

        worksheet.cell(
            row=row,
            column=self.SSP_REFERENCE_COLUMN,
            value=self._preserved_or_default(
                existing_values,
                self.SSP_REFERENCE_COLUMN,
                "",
            ),
        )

        worksheet.cell(
            row=row,
            column=self.NOTES_COLUMN,
            value=self._preserved_or_default(
                existing_values,
                self.NOTES_COLUMN,
                "",
            ),
        )

    def _clear_generated_row(
        self,
        worksheet: Worksheet,
        row: int,
    ) -> None:
        for column in range(
            1,
            self.NOTES_COLUMN + 1,
        ):
            worksheet.cell(
                row=row,
                column=column,
                value="",
            )

        worksheet.cell(
            row=row,
            column=self.STATUS_COLUMN,
            value="Open",
        )

        worksheet.cell(
            row=row,
            column=self.PRIORITY_COLUMN,
            value="Medium",
        )

        worksheet.cell(
            row=row,
            column=self.SEVERITY_COLUMN,
            value="Medium",
        )

        worksheet.cell(
            row=row,
            column=self.LIKELIHOOD_COLUMN,
            value="Possible",
        )

        worksheet.cell(
            row=row,
            column=self.RISK_SCORE_COLUMN,
            value=self._risk_formula(row),
        )

        worksheet.cell(
            row=row,
            column=self.DAYS_OPEN_COLUMN,
            value=self._days_open_formula(row),
        )

        worksheet.cell(
            row=row,
            column=self.AGING_COLUMN,
            value=self._aging_formula(row),
        )

        worksheet.cell(
            row=row,
            column=self.RESIDUAL_RISK_COLUMN,
            value="Medium",
        )

        worksheet.cell(
            row=row,
            column=self.VALIDATION_STATUS_COLUMN,
            value="Pending",
        )

    @staticmethod
    def _preserved_or_default(
        existing_values: Mapping[int, object],
        column: int,
        default: object,
    ) -> object:
        existing_value = existing_values.get(
            column
        )

        if existing_value not in {
            None,
            "",
        }:
            return existing_value

        return default

    @staticmethod
    def _to_tracker_row(
        record: AssessmentRecord,
    ) -> Dict[str, str]:
        return {
            "requirement_id": (
                record.requirement_id
            ),
            "title": record.title,
            "domain": record.domain,
            "status": (
                POAMWorksheetSynchronizer
                ._normalize_status(record.status)
            ),
        }

    @staticmethod
    def _normalize_status(
        status: str,
    ) -> str:
        return (
            str(status)
            .strip()
            .upper()
            .replace("_", " ")
        )

    @staticmethod
    def _normalize_requirement_id(
        requirement_id: str,
    ) -> str:
        normalized = (
            requirement_id
            .strip()
            .upper()
        )

        if (
            len(normalized) >= 6
            and normalized[2:6] == "-L2-"
        ):
            normalized = (
                normalized[:2]
                + ".L2-"
                + normalized[6:]
            )

        return normalized

    @staticmethod
    def _risk_formula(
        row: int,
    ) -> str:
        return (
            f'=IF(OR(L{row}="",M{row}=""),"",'
            f'IF(L{row}="Critical",4,'
            f'IF(L{row}="High",3,'
            f'IF(L{row}="Medium",2,1)))'
            f'*'
            f'IF(M{row}="Almost Certain",5,'
            f'IF(M{row}="Likely",4,'
            f'IF(M{row}="Possible",3,'
            f'IF(M{row}="Unlikely",2,1)))))'
        )

    @staticmethod
    def _days_open_formula(
        row: int,
    ) -> str:
        return (
            f'=IF(O{row}="","",'
            f'IF(Q{row}<>"",'
            f'Q{row}-O{row},'
            f'TODAY()-O{row}))'
        )

    @staticmethod
    def _aging_formula(
        row: int,
    ) -> str:
        return (
            f'=IF(R{row}="","",'
            f'IF(R{row}<=30,"0-30 Days",'
            f'IF(R{row}<=60,"31-60 Days",'
            f'IF(R{row}<=90,"61-90 Days",'
            f'IF(R{row}<=180,'
            f'"91-180 Days","181+ Days")))))'
        )