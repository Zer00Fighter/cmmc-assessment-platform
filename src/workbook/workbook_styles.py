from __future__ import annotations

from dataclasses import dataclass

from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Protection,
    Side,
)
from openpyxl.workbook import Workbook


@dataclass(frozen=True)
class ThemeColors:
    navy: str = "17365D"
    dark_navy: str = "0B1F33"
    blue: str = "2F75B5"
    light_blue: str = "D9EAF7"
    teal: str = "00A6A6"
    green: str = "70AD47"
    light_green: str = "E2F0D9"
    amber: str = "FFC000"
    light_amber: str = "FFF2CC"
    red: str = "C00000"
    light_red: str = "F4CCCC"
    gray: str = "7F8C8D"
    light_gray: str = "E7E6E6"
    very_light_gray: str = "F4F6F7"
    white: str = "FFFFFF"
    black: str = "000000"


class WorkbookStyles:
    """Central style library for the generated CMMC workbook."""

    def __init__(self) -> None:
        self.colors = ThemeColors()

        self.thin_gray_side = Side(
            style="thin",
            color="B7B7B7",
        )

        self.medium_navy_side = Side(
            style="medium",
            color=self.colors.navy,
        )

    def configure_workbook(
        self,
        workbook: Workbook,
    ) -> None:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"

    def title_font(self) -> Font:
        return Font(
            name="Aptos Display",
            size=24,
            bold=True,
            color=self.colors.white,
        )

    def subtitle_font(self) -> Font:
        return Font(
            name="Aptos",
            size=12,
            color=self.colors.white,
        )

    def section_font(self) -> Font:
        return Font(
            name="Aptos Display",
            size=14,
            bold=True,
            color=self.colors.white,
        )

    def header_font(self) -> Font:
        return Font(
            name="Aptos",
            size=10,
            bold=True,
            color=self.colors.white,
        )

    def body_font(self) -> Font:
        return Font(
            name="Aptos",
            size=10,
            color=self.colors.black,
        )

    def small_font(self) -> Font:
        return Font(
            name="Aptos",
            size=9,
            color=self.colors.gray,
        )

    def input_font(self) -> Font:
        return Font(
            name="Aptos",
            size=10,
            color=self.colors.blue,
        )

    def title_fill(self) -> PatternFill:
        return PatternFill(
            fill_type="solid",
            fgColor=self.colors.dark_navy,
        )

    def section_fill(self) -> PatternFill:
        return PatternFill(
            fill_type="solid",
            fgColor=self.colors.navy,
        )

    def header_fill(self) -> PatternFill:
        return PatternFill(
            fill_type="solid",
            fgColor=self.colors.blue,
        )

    def subheader_fill(self) -> PatternFill:
        return PatternFill(
            fill_type="solid",
            fgColor=self.colors.light_blue,
        )

    def input_fill(self) -> PatternFill:
        return PatternFill(
            fill_type="solid",
            fgColor="FFFDEB",
        )

    def formula_fill(self) -> PatternFill:
        return PatternFill(
            fill_type="solid",
            fgColor=self.colors.very_light_gray,
        )

    def good_fill(self) -> PatternFill:
        return PatternFill(
            fill_type="solid",
            fgColor=self.colors.light_green,
        )

    def warning_fill(self) -> PatternFill:
        return PatternFill(
            fill_type="solid",
            fgColor=self.colors.light_amber,
        )

    def bad_fill(self) -> PatternFill:
        return PatternFill(
            fill_type="solid",
            fgColor=self.colors.light_red,
        )

    def thin_border(self) -> Border:
        return Border(
            left=self.thin_gray_side,
            right=self.thin_gray_side,
            top=self.thin_gray_side,
            bottom=self.thin_gray_side,
        )

    def section_border(self) -> Border:
        return Border(
            bottom=self.medium_navy_side,
        )

    def center_alignment(self) -> Alignment:
        return Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    def left_alignment(self) -> Alignment:
        return Alignment(
            horizontal="left",
            vertical="top",
            wrap_text=True,
        )

    def unlocked_protection(self) -> Protection:
        return Protection(
            locked=False,
            hidden=False,
        )

    def locked_protection(self) -> Protection:
        return Protection(
            locked=True,
            hidden=False,
        )