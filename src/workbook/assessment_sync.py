from __future__ import annotations

from dataclasses import dataclass
from typing import Iterable, List

from openpyxl.worksheet.worksheet import Worksheet


class AssessmentSyncError(Exception):
    """Raised when Assessment worksheet synchronization fails."""


@dataclass(slots=True)
class AssessmentRecord:
    requirement_id: str
    title: str
    domain: str
    status: str
    implementation_state: str
    applicable: bool
    owner: str
    evidence_status: str
    row_number: int


class AssessmentSynchronizer:
    """
    Reads the Assessment worksheet and produces
    normalized AssessmentRecord objects for the
    synchronization pipeline.
    """

    FIRST_DATA_ROW = 6

    #
    # Assessment worksheet columns
    #
    DOMAIN_COLUMN = 1
    REQUIREMENT_COLUMN = 2
    TITLE_COLUMN = 3
    APPLICABLE_COLUMN = 7
    STATUS_COLUMN = 9
    IMPLEMENTATION_COLUMN = 10
    EVIDENCE_COLUMN = 15
    OWNER_COLUMN = 16

    def synchronize(
        self,
        worksheet: Worksheet,
    ) -> List[AssessmentRecord]:

        records: List[AssessmentRecord] = []

        row = self.FIRST_DATA_ROW

        while True:

            requirement = worksheet.cell(
                row=row,
                column=self.REQUIREMENT_COLUMN,
            ).value

            if requirement in (None, ""):
                break

            records.append(
                self._record_from_row(
                    worksheet,
                    row,
                )
            )

            row += 1

        return records

    def _record_from_row(
        self,
        worksheet: Worksheet,
        row: int,
    ) -> AssessmentRecord:

        applicable_value = worksheet.cell(
            row=row,
            column=self.APPLICABLE_COLUMN,
        ).value

        applicable = str(
            applicable_value
        ).strip().upper() != "NO"

        return AssessmentRecord(
            requirement_id=str(
                worksheet.cell(
                    row=row,
                    column=self.REQUIREMENT_COLUMN,
                ).value
            ),
            title=str(
                worksheet.cell(
                    row=row,
                    column=self.TITLE_COLUMN,
                ).value
            ),
            domain=str(
                worksheet.cell(
                    row=row,
                    column=self.DOMAIN_COLUMN,
                ).value
            ),
            status=str(
                worksheet.cell(
                    row=row,
                    column=self.STATUS_COLUMN,
                ).value
            ),
            implementation_state=str(
                worksheet.cell(
                    row=row,
                    column=self.IMPLEMENTATION_COLUMN,
                ).value
                or ""
            ),
            applicable=applicable,
            owner=str(
                worksheet.cell(
                    row=row,
                    column=self.OWNER_COLUMN,
                ).value
                or ""
            ),
            evidence_status=str(
                worksheet.cell(
                    row=row,
                    column=self.EVIDENCE_COLUMN,
                ).value
                or ""
            ),
            row_number=row,
        )

    def not_met_records(
        self,
        records: Iterable[AssessmentRecord],
    ) -> List[AssessmentRecord]:
        return [
            record
            for record in records
            if record.status == "NOT MET"
        ]

    def applicable_records(
        self,
        records: Iterable[AssessmentRecord],
    ) -> List[AssessmentRecord]:
        return [
            record
            for record in records
            if record.applicable
        ]