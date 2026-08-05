from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, Iterable, List, Set

from openpyxl.worksheet.worksheet import Worksheet


class EvidenceSyncError(Exception):
    """Raised when Evidence worksheet synchronization fails."""


@dataclass(frozen=True)
class EvidenceRecord:
    evidence_id: str
    title: str
    evidence_type: str
    status: str
    requirement_ids: List[str]
    objective_ids: List[str]
    owner: str
    row_number: int


@dataclass(frozen=True)
class EvidenceCoverage:
    total_evidence: int
    covered_requirements: int
    uncovered_requirements: int
    orphaned_evidence: int
    duplicate_evidence: int


class EvidenceSynchronizer:

    FIRST_DATA_ROW = 6

    COL_EVIDENCE_ID = 1
    COL_TITLE = 2
    COL_TYPE = 3
    COL_OWNER = 6
    COL_STATUS = 7
    COL_REQUIREMENTS = 12
    COL_OBJECTIVES = 13

    def synchronize(
        self,
        worksheet: Worksheet,
    ) -> List[EvidenceRecord]:

        records: List[EvidenceRecord] = []

        row = self.FIRST_DATA_ROW

        while True:

            evidence_id = worksheet.cell(
                row=row,
                column=self.COL_EVIDENCE_ID,
            ).value

            if evidence_id in (None, ""):
                break

            records.append(
                self._read_record(
                    worksheet,
                    row,
                )
            )

            row += 1

        return records

    def calculate_coverage(
        self,
        records: Iterable[EvidenceRecord],
        all_requirements: Iterable[str],
    ) -> EvidenceCoverage:

        records = list(records)

        referenced_requirements: Set[str] = set()

        evidence_ids: Set[str] = set()

        duplicates = 0

        orphaned = 0

        for record in records:

            if record.evidence_id in evidence_ids:
                duplicates += 1
            else:
                evidence_ids.add(
                    record.evidence_id
                )

            if not record.requirement_ids:
                orphaned += 1

            referenced_requirements.update(
                record.requirement_ids
            )

        all_requirements = {
            requirement.upper()
            for requirement in all_requirements
        }

        covered = len(
            referenced_requirements
            & all_requirements
        )

        uncovered = (
            len(all_requirements)
            - covered
        )

        return EvidenceCoverage(
            total_evidence=len(records),
            covered_requirements=covered,
            uncovered_requirements=uncovered,
            orphaned_evidence=orphaned,
            duplicate_evidence=duplicates,
        )

    def _read_record(
        self,
        worksheet: Worksheet,
        row: int,
    ) -> EvidenceRecord:

        requirements = self._split_list(
            worksheet.cell(
                row=row,
                column=self.COL_REQUIREMENTS,
            ).value
        )

        objectives = self._split_list(
            worksheet.cell(
                row=row,
                column=self.COL_OBJECTIVES,
            ).value
        )

        return EvidenceRecord(
            evidence_id=str(
                worksheet.cell(
                    row=row,
                    column=self.COL_EVIDENCE_ID,
                ).value
            ),
            title=str(
                worksheet.cell(
                    row=row,
                    column=self.COL_TITLE,
                ).value
                or ""
            ),
            evidence_type=str(
                worksheet.cell(
                    row=row,
                    column=self.COL_TYPE,
                ).value
                or ""
            ),
            status=str(
                worksheet.cell(
                    row=row,
                    column=self.COL_STATUS,
                ).value
                or ""
            ),
            requirement_ids=requirements,
            objective_ids=objectives,
            owner=str(
                worksheet.cell(
                    row=row,
                    column=self.COL_OWNER,
                ).value
                or ""
            ),
            row_number=row,
        )

    @staticmethod
    def _split_list(
        value: object,
    ) -> List[str]:

        if value in (None, ""):
            return []

        return [
            item.strip().upper()
            for item in str(value).split(",")
            if item.strip()
        ]