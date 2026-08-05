from __future__ import annotations

from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from src.workbook.workbook_styles import WorkbookStyles
from src.workbook.worksheet_factory import WorksheetFactory


class EvidenceSheetBuilder:
    """Build the workbook Evidence Register."""

    HEADER_ROW = 5
    FIRST_DATA_ROW = 6
    MAX_EVIDENCE_ROWS = 500

    HEADERS = [
        "Evidence ID",
        "Evidence Title",
        "Evidence Type",
        "Description",
        "Storage Location / URL",
        "Document Owner",
        "Evidence Status",
        "Review Status",
        "Reviewer",
        "Review Date",
        "Expiration Date",
        "Requirement IDs",
        "Objective IDs",
        "Confidentiality",
        "Version",
        "Notes",
    ]

    def __init__(
        self,
        styles: WorkbookStyles,
        factory: WorksheetFactory,
    ) -> None:
        self.styles = styles
        self.factory = factory

    @property
    def last_data_row(self) -> int:
        return (
            self.FIRST_DATA_ROW
            + self.MAX_EVIDENCE_ROWS
            - 1
        )

    def build(
        self,
        worksheet: Worksheet,
    ) -> None:
        self.factory.configure_standard_sheet(
            worksheet,
            freeze_cell="A6",
            show_gridlines=False,
            zoom_scale=80,
        )

        self.factory.create_title_band(
            worksheet,
            title="Evidence Register",
            subtitle=(
                "Track policies, procedures, screenshots, "
                "configurations, interviews, tests, records, "
                "and other evidence supporting the assessment."
            ),
            end_column=len(self.HEADERS),
        )

        self._write_headers(worksheet)
        self._create_rows(worksheet)
        self._add_validations(worksheet)
        self._add_conditional_formatting(worksheet)
        self._configure_columns(worksheet)
        self._configure_printing(worksheet)

    def _write_headers(
        self,
        worksheet: Worksheet,
    ) -> None:
        for column, header in enumerate(
            self.HEADERS,
            start=1,
        ):
            worksheet.cell(
                row=self.HEADER_ROW,
                column=column,
                value=header,
            )

        self.factory.style_table_header(
            worksheet,
            row_number=self.HEADER_ROW,
            start_column=1,
            end_column=len(self.HEADERS),
        )

    def _create_rows(
        self,
        worksheet: Worksheet,
    ) -> None:
        for row in range(
            self.FIRST_DATA_ROW,
            self.last_data_row + 1,
        ):
            worksheet.cell(
                row=row,
                column=1,
                value="",
            )

            worksheet.cell(
                row=row,
                column=2,
                value="",
            )

            worksheet.cell(
                row=row,
                column=3,
                value="Document",
            )

            worksheet.cell(
                row=row,
                column=4,
                value="",
            )

            worksheet.cell(
                row=row,
                column=5,
                value="",
            )

            worksheet.cell(
                row=row,
                column=6,
                value="",
            )

            worksheet.cell(
                row=row,
                column=7,
                value="Not Started",
            )

            worksheet.cell(
                row=row,
                column=8,
                value="Not Reviewed",
            )

            worksheet.cell(
                row=row,
                column=9,
                value="",
            )

            worksheet.cell(
                row=row,
                column=10,
                value="",
            )

            worksheet.cell(
                row=row,
                column=11,
                value="",
            )

            worksheet.cell(
                row=row,
                column=12,
                value="",
            )

            worksheet.cell(
                row=row,
                column=13,
                value="",
            )

            worksheet.cell(
                row=row,
                column=14,
                value="CUI",
            )

            worksheet.cell(
                row=row,
                column=15,
                value="",
            )

            worksheet.cell(
                row=row,
                column=16,
                value="",
            )

            for column in range(
                1,
                len(self.HEADERS) + 1,
            ):
                cell = worksheet.cell(
                    row=row,
                    column=column,
                )

                cell.font = self.styles.body_font()
                cell.border = self.styles.thin_border()
                cell.alignment = self.styles.left_alignment()
                cell.fill = self.styles.input_fill()
                cell.protection = (
                    self.styles.unlocked_protection()
                )

            worksheet.cell(
                row=row,
                column=10,
            ).number_format = "mm/dd/yyyy"

            worksheet.cell(
                row=row,
                column=11,
            ).number_format = "mm/dd/yyyy"

    def _add_validations(
        self,
        worksheet: Worksheet,
    ) -> None:
        evidence_type_validation = DataValidation(
            type="list",
            formula1=(
                '"Policy,Procedure,Plan,Configuration,'
                'Screenshot,Record,Report,Interview,'
                'Test Result,Log,Diagram,Contract,Other"'
            ),
            allow_blank=False,
        )

        worksheet.add_data_validation(
            evidence_type_validation
        )

        evidence_type_validation.add(
            f"C{self.FIRST_DATA_ROW}:"
            f"C{self.last_data_row}"
        )

        evidence_status_validation = DataValidation(
            type="list",
            formula1=(
                '"Not Started,In Progress,Complete,'
                'Expired,Not Applicable"'
            ),
            allow_blank=False,
        )

        worksheet.add_data_validation(
            evidence_status_validation
        )

        evidence_status_validation.add(
            f"G{self.FIRST_DATA_ROW}:"
            f"G{self.last_data_row}"
        )

        review_status_validation = DataValidation(
            type="list",
            formula1=(
                '"Not Reviewed,In Review,Approved,'
                'Rejected,Needs Update"'
            ),
            allow_blank=False,
        )

        worksheet.add_data_validation(
            review_status_validation
        )

        review_status_validation.add(
            f"H{self.FIRST_DATA_ROW}:"
            f"H{self.last_data_row}"
        )

        confidentiality_validation = DataValidation(
            type="list",
            formula1=(
                '"Public,Internal,Confidential,CUI,'
                'Controlled,Restricted"'
            ),
            allow_blank=False,
        )

        worksheet.add_data_validation(
            confidentiality_validation
        )

        confidentiality_validation.add(
            f"N{self.FIRST_DATA_ROW}:"
            f"N{self.last_data_row}"
        )

        date_validation = DataValidation(
            type="date",
            operator="between",
            formula1="DATE(2000,1,1)",
            formula2="DATE(2100,12,31)",
            allow_blank=True,
        )

        worksheet.add_data_validation(
            date_validation
        )

        date_validation.add(
            f"J{self.FIRST_DATA_ROW}:"
            f"K{self.last_data_row}"
        )

    def _add_conditional_formatting(
        self,
        worksheet: Worksheet,
    ) -> None:
        status_range = (
            f"G{self.FIRST_DATA_ROW}:"
            f"G{self.last_data_row}"
        )

        worksheet.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=[
                    f'G{self.FIRST_DATA_ROW}="Complete"'
                ],
                fill=self.styles.good_fill(),
            ),
        )

        worksheet.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=[
                    f'G{self.FIRST_DATA_ROW}="In Progress"'
                ],
                fill=self.styles.warning_fill(),
            ),
        )

        worksheet.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=[
                    f'G{self.FIRST_DATA_ROW}="Expired"'
                ],
                fill=self.styles.bad_fill(),
            ),
        )

        review_range = (
            f"H{self.FIRST_DATA_ROW}:"
            f"H{self.last_data_row}"
        )

        worksheet.conditional_formatting.add(
            review_range,
            FormulaRule(
                formula=[
                    f'H{self.FIRST_DATA_ROW}="Approved"'
                ],
                fill=self.styles.good_fill(),
            ),
        )

        worksheet.conditional_formatting.add(
            review_range,
            FormulaRule(
                formula=[
                    f'H{self.FIRST_DATA_ROW}="Rejected"'
                ],
                fill=self.styles.bad_fill(),
            ),
        )

        expiration_range = (
            f"K{self.FIRST_DATA_ROW}:"
            f"K{self.last_data_row}"
        )

        worksheet.conditional_formatting.add(
            expiration_range,
            FormulaRule(
                formula=[
                    (
                        f'AND(K{self.FIRST_DATA_ROW}<>"",'
                        f'K{self.FIRST_DATA_ROW}<TODAY())'
                    )
                ],
                fill=self.styles.bad_fill(),
            ),
        )

    def _configure_columns(
        self,
        worksheet: Worksheet,
    ) -> None:
        widths = {
            "A": 16,
            "B": 34,
            "C": 18,
            "D": 45,
            "E": 48,
            "F": 22,
            "G": 18,
            "H": 18,
            "I": 22,
            "J": 14,
            "K": 14,
            "L": 35,
            "M": 28,
            "N": 18,
            "O": 14,
            "P": 45,
        }

        for column, width in widths.items():
            worksheet.column_dimensions[
                column
            ].width = width

        worksheet.auto_filter.ref = (
            f"A{self.HEADER_ROW}:"
            f"P{self.last_data_row}"
        )

    def _configure_printing(
        self,
        worksheet: Worksheet,
    ) -> None:
        worksheet.print_title_rows = (
            f"1:{self.HEADER_ROW}"
        )

        worksheet.print_area = (
            f"A1:P{self.last_data_row}"
        )

        worksheet.page_setup.orientation = (
            "landscape"
        )