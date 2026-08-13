"""Readiness validation for generating a completed Word Security Plan."""

from __future__ import annotations

from dataclasses import dataclass
from enum import Enum
from pathlib import Path

from openpyxl import load_workbook


class ReadinessSeverity(str, Enum):
    BLOCKER = "BLOCKER"
    WARNING = "WARNING"


@dataclass(frozen=True)
class ReadinessIssue:
    severity: ReadinessSeverity
    field: str
    message: str
    requirement_id: str = ""


@dataclass(frozen=True)
class SSPReadinessReport:
    workbook_path: Path
    total_controls: int
    assessed_controls: int
    issues: tuple[ReadinessIssue, ...]

    @property
    def blockers(self) -> tuple[ReadinessIssue, ...]:
        return tuple(
            issue
            for issue in self.issues
            if issue.severity == ReadinessSeverity.BLOCKER
        )

    @property
    def warnings(self) -> tuple[ReadinessIssue, ...]:
        return tuple(
            issue
            for issue in self.issues
            if issue.severity == ReadinessSeverity.WARNING
        )

    @property
    def ready(self) -> bool:
        return not self.blockers

    @property
    def completion_percent(self) -> float:
        if self.total_controls == 0:
            return 0.0
        return round(self.assessed_controls / self.total_controls * 100, 2)

    def to_text(self) -> str:
        status = "READY" if self.ready else "NOT READY"
        lines = [
            "Omni Word SSP Readiness Report",
            "=" * 31,
            f"Workbook: {self.workbook_path}",
            f"Status: {status}",
            f"Assessment completion: {self.assessed_controls}/{self.total_controls} "
            f"({self.completion_percent:.2f}%)",
            f"Blockers: {len(self.blockers)}",
            f"Warnings: {len(self.warnings)}",
            "",
        ]
        for heading, issues in (
            ("BLOCKERS", self.blockers),
            ("WARNINGS", self.warnings),
        ):
            lines.append(heading)
            lines.append("-" * len(heading))
            if not issues:
                lines.append("None")
            else:
                for issue in issues:
                    control = (
                        f" [{issue.requirement_id}]" if issue.requirement_id else ""
                    )
                    lines.append(f"- {issue.field}{control}: {issue.message}")
            lines.append("")
        return "\n".join(lines)


def _text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.startswith("=") else text


def validate_ssp_readiness(workbook_path: str | Path) -> SSPReadinessReport:
    """Validate whether an Omni workbook can produce a completed SSP."""

    path = Path(workbook_path).resolve()
    if not path.is_file():
        raise FileNotFoundError(f"Omni workbook not found: {path}")

    workbook = load_workbook(path, data_only=False, read_only=True)
    try:
        required_sheets = {"Cover", "Assessment", "SSP Crosswalk"}
        missing_sheets = sorted(required_sheets - set(workbook.sheetnames))
        if missing_sheets:
            issues = tuple(
                ReadinessIssue(
                    ReadinessSeverity.BLOCKER,
                    "Workbook structure",
                    f"Missing required worksheet: {sheet}",
                )
                for sheet in missing_sheets
            )
            return SSPReadinessReport(path, 0, 0, issues)

        issues: list[ReadinessIssue] = []
        cover = workbook["Cover"]
        cover_values = {
            _text(cover.cell(row, 2).value): _text(cover.cell(row, 3).value)
            for row in range(1, cover.max_row + 1)
            if _text(cover.cell(row, 2).value)
        }
        required_demographics = (
            "Organization Name",
            "Assessment Name",
            "Assessment Scope",
            "CAGE Code",
            "Assessment Start Date",
            "Assessment End Date",
            "Lead Assessor",
        )
        for field in required_demographics:
            if not cover_values.get(field):
                issues.append(
                    ReadinessIssue(
                        ReadinessSeverity.BLOCKER,
                        field,
                        "Required SSP demographic value is missing.",
                    )
                )

        assessment = workbook["Assessment"]
        crosswalk = workbook["SSP Crosswalk"]
        total_controls = 0
        assessed_controls = 0
        for row in range(6, assessment.max_row + 1):
            requirement_id = _text(assessment.cell(row, 2).value)
            if not requirement_id:
                continue
            total_controls += 1
            status = _text(assessment.cell(row, 9).value).upper()
            notes = _text(assessment.cell(row, 18).value)
            ssp_reference = _text(assessment.cell(row, 17).value)
            governance = _text(crosswalk.cell(row, 4).value)
            evidence = _text(crosswalk.cell(row, 5).value)
            mapping_status = _text(crosswalk.cell(row, 7).value)

            if status in {"MET", "NOT MET", "NOT APPLICABLE"}:
                assessed_controls += 1
            else:
                issues.append(
                    ReadinessIssue(
                        ReadinessSeverity.BLOCKER,
                        "Assessment Status",
                        "Control has not been assessed.",
                        requirement_id,
                    )
                )
            if status in {"MET", "NOT MET", "NOT APPLICABLE"} and not notes:
                issues.append(
                    ReadinessIssue(
                        ReadinessSeverity.BLOCKER,
                        "Assessor Notes / Findings",
                        "A conformity statement, finding, or N/A justification is required.",
                        requirement_id,
                    )
                )
            if not ssp_reference:
                issues.append(
                    ReadinessIssue(
                        ReadinessSeverity.WARNING,
                        "Security Plan Reference (SSP)",
                        "No SSP reference is recorded.",
                        requirement_id,
                    )
                )
            if not governance:
                issues.append(
                    ReadinessIssue(
                        ReadinessSeverity.WARNING,
                        "System Design Documentation",
                        "No policy, plan, standard, or procedure is mapped.",
                        requirement_id,
                    )
                )
            if not evidence:
                issues.append(
                    ReadinessIssue(
                        ReadinessSeverity.WARNING,
                        "Supporting Artifacts",
                        "No evidence reference is mapped.",
                        requirement_id,
                    )
                )
            if mapping_status not in {"Mapped", "Partially Mapped"}:
                issues.append(
                    ReadinessIssue(
                        ReadinessSeverity.WARNING,
                        "SSP Mapping Status",
                        "Control is not marked Mapped or Partially Mapped.",
                        requirement_id,
                    )
                )

        return SSPReadinessReport(
            workbook_path=path,
            total_controls=total_controls,
            assessed_controls=assessed_controls,
            issues=tuple(issues),
        )
    finally:
        workbook.close()


def write_readiness_report(report: SSPReadinessReport, output_path: str | Path) -> Path:
    path = Path(output_path).resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(report.to_text(), encoding="utf-8")
    return path
