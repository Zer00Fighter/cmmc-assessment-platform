from __future__ import annotations

from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import RemediationPlan


HEADERS = (
    "Remediation ID (POA&M ID)", "Requirement ID", "Requirement Title", "Domain",
    "Weakness Description", "Root Cause", "Corrective Action", "Current Milestone",
    "Milestone Owner", "Status", "Priority", "Severity", "Likelihood", "Risk Score",
    "Date Identified", "Planned Completion", "Actual Completion", "Days Open",
    "Aging Bucket", "Residual Risk", "Validation Status", "Evidence IDs",
    "Security Plan Reference (SSP)", "Assessor Notes",
)


def _aging(days: int) -> str:
    if days <= 30:
        return "0-30 Days"
    if days <= 60:
        return "31-60 Days"
    if days <= 90:
        return "61-90 Days"
    if days <= 180:
        return "91-180 Days"
    return "181+ Days"


def build_remediation_workbook(assessment) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Remediation Action Plan"
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    sheet.cell(1, 1, "Omni Remediation Action Plan (POA&M)")
    sheet.cell(1, 1).font = Font(size=18, bold=True, color="FFFFFF")
    sheet.cell(1, 1).fill = PatternFill("solid", fgColor="102A43")
    sheet.cell(2, 1, f"Assessment: {assessment.name}")
    sheet.cell(3, 1, f"System: {assessment.system.name}")
    for column, header in enumerate(HEADERS, 1):
        cell = sheet.cell(5, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1677A6")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    row = 6
    today = timezone.localdate()
    plans = assessment.remediation_plans.prefetch_related(
        "controls__requirement", "milestones__owner__user", "closure_evidence"
    ).select_related("owner__user", "validated_by")
    for plan in plans:
        controls = list(plan.controls.select_related("requirement"))
        requirement_ids = ", ".join(item.requirement.requirement_id for item in controls)
        titles = "; ".join(item.requirement.title for item in controls)
        domains = ", ".join(sorted({item.requirement.domain for item in controls}))
        milestone = plan.milestones.exclude(status="COMPLETE").first() or plan.milestones.first()
        end = plan.actual_completion or today
        days_open = max((end - plan.date_identified).days, 0)
        evidence_ids = ", ".join(str(item.id) for item in plan.closure_evidence.all())
        ssp_refs = ", ".join(sorted({item.ssp_reference for item in controls if item.ssp_reference}))
        values = (
            plan.remediation_id, requirement_ids, titles, domains,
            plan.weakness_description, plan.root_cause, plan.corrective_action,
            milestone.title if milestone else "",
            str(milestone.owner) if milestone and milestone.owner else "",
            plan.get_status_display(), plan.get_priority_display(), plan.get_severity_display(),
            plan.get_likelihood_display(), plan.risk_score, plan.date_identified,
            plan.planned_completion, plan.actual_completion, days_open, _aging(days_open),
            plan.get_residual_risk_display(), plan.get_validation_status_display(),
            evidence_ids, ssp_refs, plan.validation_notes,
        )
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row, column, value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if column in (15, 16, 17) and value:
                cell.number_format = "yyyy-mm-dd"
        row += 1
    sheet.freeze_panes = "A6"
    sheet.auto_filter.ref = f"A5:X{max(row - 1, 5)}"
    for column in range(1, len(HEADERS) + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 18
    for column in (3, 5, 6, 7, 24):
        sheet.column_dimensions[get_column_letter(column)].width = 32
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
