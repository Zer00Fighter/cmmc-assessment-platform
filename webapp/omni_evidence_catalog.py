from __future__ import annotations
import hashlib
from pathlib import Path
from openpyxl import load_workbook
from django.db import transaction
from src.evidence.evidence_resolver import EvidenceResolver
from .models import OmniEvidenceSourceRequest


def _split(value):
    return [item.strip() for item in str(value or "").replace(";", "\n").replace(",", "\n").splitlines() if item.strip()]


def normalize_cmmc(identifier):
    # Level 1 practices included in a Level 2 assessment use the same NIST practice number.
    return identifier.replace(".L1-", ".L2-") if ".L1-" in identifier else identifier


def read_catalog(path):
    source = Path(path); raw = source.read_bytes(); digest = hashlib.sha256(raw).hexdigest()
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook.active
    resolver = EvidenceResolver(); records = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        resolution = resolver.resolve(str(row[3] or "").strip())
        records.append({"source_identifier": str(row[1] or "").strip(), "area_of_focus": str(row[2] or "").strip(),
            "source_title": str(row[3] or "").strip(), "source_description": str(row[4] or "").strip(),
            "omni_control_ids": _split(row[5]), "source_cmmc_ids": _split(row[6]),
            "normalized_cmmc_ids": [normalize_cmmc(x) for x in _split(row[6])],
            "canonical_evidence_code": resolution.evidence_id or "",
            "resolution": OmniEvidenceSourceRequest.Resolution.EXACT if resolution.resolved else OmniEvidenceSourceRequest.Resolution.REVIEW,
            "source_row": row_number, "source_filename": source.name, "source_sha256": digest})
    workbook.close()
    return records


@transaction.atomic
def import_catalog(path, apply=False):
    records = read_catalog(path)
    if apply:
        for item in records:
            existing = OmniEvidenceSourceRequest.objects.filter(source_identifier=item["source_identifier"]).first()
            # Preserve governed human decisions when refreshing the same source row.
            if existing and existing.resolution != OmniEvidenceSourceRequest.Resolution.EXACT:
                for key in ("canonical_evidence_code", "resolution"):
                    item[key] = getattr(existing, key)
            OmniEvidenceSourceRequest.objects.update_or_create(source_identifier=item["source_identifier"], defaults=item)
    return {"requests": len(records), "exact": sum(x["resolution"] == "EXACT" for x in records),
            "review": sum(x["resolution"] == "REVIEW" for x in records),
            "controls": len({c for x in records for c in x["omni_control_ids"]}),
            "applied": apply, "sha256": records[0]["source_sha256"] if records else ""}
