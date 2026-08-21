"""Private ingestion for user-authorized AICPA Points of Focus content."""
from __future__ import annotations

import csv
import hashlib
import re
from pathlib import Path

from django.db import transaction
from openpyxl import load_workbook

from .models import Requirement, Soc2PointOfFocus
from .soc2_activity_import import TSC_FRAMEWORK_CODE

ALIASES = {
    "criterion_id": {"criterion id", "control id", "requirement id", "tsc id"},
    "point_id": {"point of focus id", "point id", "pof id"},
    "text": {"point of focus", "point of focus text", "pof text", "text"},
    "source_reference": {"source reference", "source", "reference"},
    "source_page": {"source page", "page"},
}


def _key(value) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().casefold()).strip()


def _header_map(row) -> dict[str, int]:
    result = {}
    for index, value in enumerate(row):
        normalized = _key(value)
        for field, aliases in ALIASES.items():
            if normalized in aliases and field not in result:
                result[field] = index
    return result


def _structured_rows(path: Path):
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as stream:
            yield "CSV", list(csv.reader(stream))
        return
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                yield sheet.title, list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()
        return
    raise ValueError("Points of Focus sources must be CSV, XLSX, or XLSM.")


def normalize_points_of_focus(path: str | Path) -> tuple[dict, dict]:
    path = Path(path)
    content = path.read_bytes()
    digest = hashlib.sha256(content).hexdigest()
    rows, errors = [], []
    for source_sheet, values in _structured_rows(path):
        header_index = header = None
        for index, row in enumerate(values[:20]):
            candidate = _header_map(row)
            if {"criterion_id", "point_id", "text"}.issubset(candidate):
                header_index, header = index, candidate
                break
        if header is None:
            errors.append(f"{source_sheet}: Points of Focus header not found.")
            continue
        for source_row, raw in enumerate(values[header_index + 1:], header_index + 2):
            def value(field):
                position = header.get(field)
                return str(raw[position] or "").strip() if position is not None and position < len(raw) else ""
            criterion_id, point_id, text = value("criterion_id"), value("point_id"), value("text")
            if not any((criterion_id, point_id, text)):
                continue
            if not all((criterion_id, point_id, text)):
                errors.append(f"{source_sheet} row {source_row}: criterion, point ID, and text are required.")
                continue
            page = value("source_page")
            rows.append({
                "criterion_id": criterion_id, "point_id": point_id,
                "licensed_text": text, "source_sheet": source_sheet,
                "source_row": source_row, "source_reference": value("source_reference"),
                "source_page": int(page) if page.isdigit() else None,
            })
    official = set(Requirement.objects.filter(
        framework__code=TSC_FRAMEWORK_CODE
    ).values_list("requirement_id", flat=True))
    unknown = sorted({row["criterion_id"] for row in rows} - official)
    if unknown:
        errors.append(f"Unknown AICPA TSC criteria: {', '.join(unknown)}")
    keys = [(row["criterion_id"].casefold(), row["point_id"].casefold()) for row in rows]
    duplicates = sorted({f"{criterion}:{point}" for criterion, point in keys if keys.count((criterion, point)) > 1})
    if duplicates:
        errors.append(f"Duplicate Points of Focus: {', '.join(duplicates[:25])}")
    normalized = {"source_filename": path.name, "source_sha256": digest, "points": rows}
    return normalized, {"valid": bool(rows) and not errors, "errors": errors, "point_count": len(rows)}


@transaction.atomic
def import_points_of_focus(path: str | Path) -> tuple[dict, dict]:
    normalized, report = normalize_points_of_focus(path)
    if not report["valid"]:
        raise ValueError("Invalid Points of Focus source: " + "; ".join(report["errors"]))
    digest = normalized["source_sha256"]
    existing = Soc2PointOfFocus.objects.filter(source_sha256=digest)
    if existing.exists():
        return {"created": 0, "existing": existing.count()}, report
    requirements = {item.requirement_id: item for item in Requirement.objects.filter(
        framework__code=TSC_FRAMEWORK_CODE
    )}
    created = Soc2PointOfFocus.objects.bulk_create([
        Soc2PointOfFocus(
            requirement=requirements[item["criterion_id"]], point_id=item["point_id"],
            licensed_text=item["licensed_text"], source_filename=normalized["source_filename"],
            source_sha256=digest, source_reference=item["source_reference"],
            source_row=item["source_row"], source_page=item["source_page"],
        ) for item in normalized["points"]
    ])
    return {"created": len(created), "existing": 0}, report
