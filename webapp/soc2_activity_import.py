"""Normalize practical SOC 2 checklist activities without treating them as TSC criteria."""
from __future__ import annotations

import hashlib
import re
from datetime import date, datetime
from pathlib import Path

from django.db import transaction
from openpyxl import load_workbook

from .models import (
    Framework,
    ImplementationActivity,
    ImplementationActivityMapping,
)
from .soc2_tsc import EXPECTED_BY_DOMAIN

TSC_FRAMEWORK_CODE = "AICPA-TSC-2017-RPOF-2022"
MASTER_SHEET = "All Controls"
REQUIRED_COLUMNS = (
    "Control ID", "TSC / Control Area", "CC Category", "Control Activity",
    "Control Type", "Owner", "Frequency", "System / Scope", "Evidence Artifact",
    "Evidence Source", "Status", "Last Reviewed", "Notes", "ISO 27001 Ref",
)
MIRROR_SHEETS = {
    "CC-Series (Required)": "CC",
    "Security Controls": "SEC-",
    "Privacy Controls": "PRIV-",
    "Confidentiality Controls": "CONF-",
    "Availability Controls": "AVAIL-",
    "Processing Integrity": "PI-",
}

# These are implementation-support relationships, not equivalence assertions.
ACTIVITY_TARGETS = {
    "CC1.1": ("CC1.1",), "CC1.2": ("CC1.3", "CC1.5"),
    "CC1.3": ("CC1.4",), "CC1.4": ("CC1.4",), "CC1.5": ("CC1.2",),
    "CC2.1": ("CC2.2",), "CC2.2": ("CC2.3",), "CC2.3": ("CC2.1", "CC5.3"),
    "CC3.1": ("CC3.2",), "CC3.2": ("CC3.2",), "CC3.3": ("CC3.3", "CC3.4"),
    "CC4.1": ("CC4.1",), "CC4.2": ("CC4.2",),
    "CC5.1": ("CC5.3",), "CC5.2": ("CC4.1", "CC5.3"),
    "CC6.1": ("CC6.1", "CC6.2"), "CC6.2": ("CC6.1", "CC6.3"),
    "CC6.3": ("CC6.2", "CC6.3"), "CC6.4": ("CC6.2", "CC6.3"),
    "CC6.5": ("CC6.3",), "CC6.6": ("CC6.4",), "CC6.7": ("CC6.6",),
    "CC7.1": ("CC7.2",), "CC7.2": ("CC7.2",), "CC7.3": ("CC7.1",),
    "CC7.4": ("CC7.4", "CC7.5"),
    "CC8.1": ("CC8.1",), "CC8.2": ("CC8.1",), "CC8.3": ("CC8.1",),
    "CC9.1": ("CC9.2",), "CC9.2": ("CC9.2",), "CC9.3": ("A1.3",),
    "SEC-01": ("CC6.1", "CC6.2"), "SEC-02": ("CC6.6",),
    "SEC-03": ("CC7.2",), "SEC-04": ("CC6.4",),
    "SEC-05": ("CC1.4",), "SEC-06": ("CC1.4",),
    "PRIV-01": ("P2.1",), "PRIV-02": ("P3.1",), "PRIV-03": ("P3.2",),
    "PRIV-04": ("P4.1",), "PRIV-05": ("P4.2", "P4.3"), "PRIV-06": ("CC2.1",),
    "CONF-01": ("C1.1",), "CONF-02": ("C1.1", "CC6.1"),
    "CONF-03": ("C1.1", "CC6.7"), "CONF-04": ("C1.2",), "CONF-05": ("C1.1",),
    "AVAIL-01": ("A1.2", "A1.3"), "AVAIL-02": ("A1.3",),
    "AVAIL-03": ("A1.3",), "AVAIL-04": ("A1.2",), "AVAIL-05": ("A1.1",),
    "PI-01": ("A1.1", "PI1.3"), "PI-02": ("PI1.2",),
    "PI-03": ("PI1.2", "PI1.3", "PI1.4"), "PI-04": ("PI1.3", "PI1.4"),
    "PI-05": ("CC8.1", "PI1.3"), "PI-06": ("PI1.3",),
}


def _text(value) -> str:
    return str(value or "").strip()


def _json_value(value):
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value if value is not None else ""


def _rows(sheet) -> list[dict]:
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        return []
    headers = tuple(_text(value) for value in values[0])
    if headers[: len(REQUIRED_COLUMNS)] != REQUIRED_COLUMNS:
        raise ValueError(f"{sheet.title}: expected the 14-column SOC 2 checklist header.")
    rows = []
    for source_row, raw in enumerate(values[1:], 2):
        identifier = _text(raw[0] if raw else "")
        if not identifier:
            continue
        record = {headers[index]: _json_value(raw[index] if index < len(raw) else "")
                  for index in range(len(REQUIRED_COLUMNS))}
        record["source_row"] = source_row
        rows.append(record)
    return rows


def normalize_workbook(path: str | Path) -> tuple[dict, dict]:
    path = Path(path)
    content = path.read_bytes()
    digest = hashlib.sha256(content).hexdigest()
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        errors, warnings = [], []
        missing_sheets = [name for name in (MASTER_SHEET, *MIRROR_SHEETS) if name not in workbook.sheetnames]
        if missing_sheets:
            return {}, {"valid": False, "errors": [f"Missing sheets: {', '.join(missing_sheets)}"]}
        try:
            master_rows = _rows(workbook[MASTER_SHEET])
        except ValueError as exc:
            return {}, {"valid": False, "errors": [str(exc)]}
        identifiers = [_text(row["Control ID"]) for row in master_rows]
        duplicates = sorted({item for item in identifiers if identifiers.count(item) > 1})
        if duplicates:
            errors.append(f"Duplicate master identifiers: {', '.join(duplicates)}")
        missing_mappings = sorted(set(identifiers) - set(ACTIVITY_TARGETS))
        unexpected_mappings = sorted(set(ACTIVITY_TARGETS) - set(identifiers))
        if missing_mappings:
            errors.append(f"Activities without proposed mappings: {', '.join(missing_mappings)}")
        if unexpected_mappings:
            errors.append(f"Mapping rules without source activities: {', '.join(unexpected_mappings)}")
        official_ids = {item for group in EXPECTED_BY_DOMAIN.values() for item in group}
        invalid_targets = sorted({target for targets in ACTIVITY_TARGETS.values()
                                  for target in targets if target not in official_ids})
        if invalid_targets:
            errors.append(f"Unknown TSC mapping targets: {', '.join(invalid_targets)}")
        master_by_id = {identifier: row for identifier, row in zip(identifiers, master_rows)}
        mirror_counts = {}
        for sheet_name, prefix in MIRROR_SHEETS.items():
            try:
                mirror = _rows(workbook[sheet_name])
            except ValueError as exc:
                errors.append(str(exc))
                continue
            mirror_counts[sheet_name] = len(mirror)
            expected = [row for identifier, row in master_by_id.items()
                        if identifier.startswith(prefix)]
            mirror_by_id = {_text(row["Control ID"]): row for row in mirror}
            if set(mirror_by_id) != {_text(row["Control ID"]) for row in expected}:
                errors.append(f"{sheet_name}: identifiers do not mirror the master list.")
                continue
            for identifier, row in mirror_by_id.items():
                source = master_by_id[identifier]
                for column in REQUIRED_COLUMNS[:5]:
                    if _text(row[column]) != _text(source[column]):
                        errors.append(f"{sheet_name}: {identifier} differs from master column {column}.")
        activities = []
        for row in master_rows:
            identifier = _text(row["Control ID"])
            activities.append({
                "source_identifier": identifier,
                "source_area": _text(row["TSC / Control Area"]),
                "category": _text(row["CC Category"]),
                "activity": _text(row["Control Activity"]),
                "control_type": _text(row["Control Type"]),
                "source_sheet": MASTER_SHEET,
                "source_row": row["source_row"],
                "metadata": {column: row[column] for column in REQUIRED_COLUMNS[5:]},
                "mappings": [{
                    "target_framework_code": TSC_FRAMEWORK_CODE,
                    "target_requirement_id": target,
                    "relationship": "SUPPORTS",
                    "review_status": "PROPOSED",
                    "confidence": "0.900" if len(ACTIVITY_TARGETS[identifier]) == 1 else "0.800",
                    "rationale": "Semantic implementation-support mapping; assessor approval required.",
                } for target in ACTIVITY_TARGETS.get(identifier, ())],
            })
        if len(master_rows) != 60:
            warnings.append(f"Expected 60 master activities from the supplied workbook; found {len(master_rows)}.")
        normalized = {
            "source_filename": path.name, "source_sha256": digest,
            "source_sheet": MASTER_SHEET, "activities": activities,
        }
        report = {
            "valid": not errors, "errors": errors, "warnings": warnings,
            "activity_count": len(activities),
            "mapping_count": sum(len(item["mappings"]) for item in activities),
            "mirror_counts": mirror_counts,
        }
        return normalized, report
    finally:
        workbook.close()


@transaction.atomic
def import_activities(path: str | Path) -> tuple[dict, dict]:
    normalized, report = normalize_workbook(path)
    if not report.get("valid"):
        raise ValueError("Invalid SOC 2 activity workbook: " + "; ".join(report.get("errors", [])))
    framework = Framework.objects.filter(code=TSC_FRAMEWORK_CODE).first()
    if not framework:
        raise ValueError("Install Sprint 19.7 with seed_soc2_tsc before importing activities.")
    digest = normalized["source_sha256"]
    existing = ImplementationActivity.objects.filter(source_sha256=digest)
    if existing.exists():
        if existing.count() != report["activity_count"]:
            raise ValueError("Existing activity import is incomplete for this source digest.")
        return {"created": 0, "existing": existing.count(), "mappings_created": 0}, report
    requirements = {item.requirement_id: item for item in framework.requirements.all()}
    activities = ImplementationActivity.objects.bulk_create([
        ImplementationActivity(
            source_identifier=item["source_identifier"], source_area=item["source_area"],
            category=item["category"], activity=item["activity"],
            control_type=item["control_type"], source_filename=normalized["source_filename"],
            source_sha256=digest, source_sheet=item["source_sheet"],
            source_row=item["source_row"], source_metadata=item["metadata"],
        ) for item in normalized["activities"]
    ])
    mappings = []
    for activity, item in zip(activities, normalized["activities"]):
        for mapping in item["mappings"]:
            target_id = mapping["target_requirement_id"]
            mappings.append(ImplementationActivityMapping(
                activity=activity, target_framework_code=TSC_FRAMEWORK_CODE,
                target_requirement_id_text=target_id,
                target_requirement=requirements[target_id], relationship="SUPPORTS",
                review_status="PROPOSED", confidence=mapping["confidence"],
                rationale=mapping["rationale"],
            ))
    ImplementationActivityMapping.objects.bulk_create(mappings)
    return {"created": len(activities), "existing": 0, "mappings_created": len(mappings)}, report
