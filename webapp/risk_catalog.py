from __future__ import annotations

import hashlib
import re
from pathlib import Path

from django.db import transaction
from openpyxl import load_workbook

from .models import RiskCatalogEntry


RISK_ID = re.compile(r"^R-[A-Z]{2}-\d+$")


def read_risk_catalog(path):
    source = Path(path)
    digest = hashlib.sha256(source.read_bytes()).hexdigest()
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook.active
    records, issues, grouping = [], [], ""
    try:
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1):
            values = list(row[:4]) + [None] * max(0, 4 - len(row))
            raw_group, raw_id, raw_title, raw_description = values[:4]
            risk_id = str(raw_id or "").strip()
            if not risk_id:
                continue
            if risk_id.casefold() in {"risk #", "risk id"}:
                continue
            if str(raw_group or "").strip():
                grouping = str(raw_group).strip()
            title = str(raw_title or "").strip()
            description = str(raw_description or "").strip()
            if not RISK_ID.fullmatch(risk_id):
                issues.append({"code": "INVALID_RISK_ID", "row": row_number, "value": risk_id})
                continue
            if not grouping or not title or not description:
                issues.append({"code": "MISSING_REQUIRED_VALUE", "row": row_number, "risk_id": risk_id})
                continue
            records.append({"risk_id": risk_id, "grouping": grouping, "title": title,
                            "description": description, "source_row": row_number,
                            "source_filename": source.name, "source_sha256": digest})
    finally:
        workbook.close()
    duplicates = sorted({item["risk_id"] for item in records if sum(
        candidate["risk_id"] == item["risk_id"] for candidate in records
    ) > 1})
    if duplicates:
        issues.append({"code": "DUPLICATE_RISK_IDS", "values": duplicates})
    return records, issues, digest


@transaction.atomic
def import_risk_catalog(path, apply=False):
    records, issues, digest = read_risk_catalog(path)
    if apply and not issues:
        incoming = set()
        for item in records:
            incoming.add(item["risk_id"])
            RiskCatalogEntry.objects.update_or_create(
                risk_id=item["risk_id"], defaults={**item, "active": True}
            )
        RiskCatalogEntry.objects.exclude(risk_id__in=incoming).update(active=False)
    return {"records": len(records), "groups": len({x["grouping"] for x in records}),
            "issues": issues, "valid": not issues, "applied": apply and not issues,
            "sha256": digest}
