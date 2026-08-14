"""Governed, format-independent framework ingestion for Omni."""
from __future__ import annotations

import csv
import hashlib
import re
from io import BytesIO, StringIO

from django.db import transaction
from django.utils import timezone
from django.utils.text import slugify
from openpyxl import load_workbook
from pypdf import PdfReader

from .models import ExternalAuthority, Framework, FrameworkImport, MappingReference, Requirement, RequirementMapping

ALIASES = {
    "requirement_id": {"requirement id", "control id", "control", "id", "practice id", "reference", "ccf"},
    "domain": {"domain", "family", "category", "function", "ccf domain"},
    "title": {"title", "control title", "requirement title", "name", "ccf control"},
    "statement": {"statement", "requirement", "control statement", "description", "text", "comprehensive controls framework ccf control description"},
    "target_framework": {"target framework", "mapped framework", "framework mapping"},
    "target_requirement": {"target requirement", "mapped requirement", "mapping", "mapped control"},
    "relationship": {"relationship", "mapping type", "type"},
}


def _key(value) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().casefold()).strip()


def _header_map(row) -> dict[str, int]:
    found = {}
    for index, value in enumerate(row):
        normalized = _key(value)
        for field, aliases in ALIASES.items():
            if normalized in aliases and field not in found:
                found[field] = index
    return found


def _rows_to_requirements(rows, source_name: str, page: int | None = None):
    rows = list(rows)
    header_index = mapping = None
    for index, row in enumerate(rows[:30]):
        candidate = _header_map(row)
        if "requirement_id" in candidate and ("statement" in candidate or "title" in candidate):
            header_index, mapping = index, candidate
            break
    if mapping is None:
        return [], [{"code": "HEADER_NOT_FOUND", "source": source_name}]
    results = []
    header = rows[header_index]
    is_ccf_matrix = any(_key(cell) == "ccf domain" for cell in header) and len(header) > 5
    mapped_columns = []
    if is_ccf_matrix:
        used = set(mapping.values())
        mapped_columns = [(index, str(value or "").strip()) for index, value in enumerate(header)
                          if index not in used and str(value or "").strip()
                          and _key(value) != "ccf control question"]
    for row_number, row in enumerate(rows[header_index + 1 :], header_index + 2):
        def value(field):
            pos = mapping.get(field)
            return str(row[pos] or "").strip() if pos is not None and pos < len(row) else ""
        requirement_id = value("requirement_id")
        if not requirement_id:
            continue
        mapping_refs = []
        for column, target_framework in mapped_columns:
            raw = str(row[column] or "").strip() if column < len(row) else ""
            if raw:
                mapping_refs.append({"target_framework": target_framework, "target_requirement": raw,
                                     "source_column": column + 1})
        results.append({
            "requirement_id": requirement_id,
            "domain": value("domain"),
            "title": value("title") or requirement_id,
            "statement": value("statement") or value("title"),
            "target_framework": value("target_framework"),
            "target_requirement": value("target_requirement"),
            "relationship": value("relationship") or "RELATED",
            "source_reference": source_name,
            "source_page": page,
            "source_row": row_number,
            "mapping_refs": mapping_refs,
        })
    if results and is_ccf_matrix:
        results[0]["authority_columns"] = [
            {"name": name, "source_column": column + 1}
            for column, name in mapped_columns
        ]
    return results, []


def parse_upload(upload, metadata: dict) -> tuple[dict, dict, str, str]:
    content = upload.read()
    upload.seek(0)
    digest = hashlib.sha256(content).hexdigest()
    suffix = upload.name.rsplit(".", 1)[-1].lower()
    requirements, issues = [], []
    if suffix == "csv":
        text = content.decode("utf-8-sig", errors="replace")
        requirements, issues = _rows_to_requirements(csv.reader(StringIO(text)), upload.name)
        source_format = FrameworkImport.SourceFormat.CSV
    elif suffix in {"xlsx", "xlsm"}:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            parsed, sheet_issues = _rows_to_requirements(
                sheet.iter_rows(values_only=True), f"{upload.name}#{sheet.title}"
            )
            requirements.extend(parsed)
            if not parsed:
                issues.extend(sheet_issues)
        source_format = FrameworkImport.SourceFormat.XLSX
    elif suffix == "pdf":
        reader = PdfReader(BytesIO(content))
        extracted = 0
        for page_number, page in enumerate(reader.pages, 1):
            text = page.extract_text() or ""
            if text.strip():
                extracted += 1
                parsed, page_issues = _rows_to_requirements(
                    csv.reader(StringIO(text), delimiter="|"), upload.name, page_number
                )
                requirements.extend(parsed)
                if not parsed:
                    issues.extend(page_issues)
        if not extracted:
            issues.append({"code": "OCR_REQUIRED", "message": "The PDF has no extractable text."})
        source_format = FrameworkImport.SourceFormat.PDF
    else:
        raise ValueError("Supported framework files are CSV, XLSX, XLSM, and PDF.")

    duplicates, seen = [], set()
    for item in requirements:
        key = item["requirement_id"].casefold()
        if key in seen:
            duplicates.append(item["requirement_id"])
        seen.add(key)
        if not item["statement"]:
            issues.append({"code": "MISSING_STATEMENT", "requirement_id": item["requirement_id"]})
    if duplicates:
        issues.append({"code": "DUPLICATE_REQUIREMENTS", "values": sorted(set(duplicates))[:25]})
    if not requirements:
        issues.append({"code": "NO_REQUIREMENTS", "message": "No normalized requirements were found."})
    normalized = {"framework": metadata, "requirements": requirements}
    mapping_count = sum(len(item.get("mapping_refs", [])) for item in requirements)
    mapped_authority_names = {ref["target_framework"] for item in requirements
                              for ref in item.get("mapping_refs", [])}
    authority_names = sorted(mapped_authority_names | {
        authority["name"] for item in requirements
        for authority in item.get("authority_columns", [])
    })
    report = {
        "valid": not any(i["code"] in {"OCR_REQUIRED", "NO_REQUIREMENTS", "DUPLICATE_REQUIREMENTS"} for i in issues),
        "requirement_count": len(requirements), "mapping_reference_count": mapping_count,
        "authority_count": len(authority_names),
        "mapped_authority_count": len(mapped_authority_names),
        "empty_authority_count": len(authority_names) - len(mapped_authority_names),
        "authorities": authority_names,
        "issue_count": len(issues), "issues": issues,
    }
    return normalized, report, source_format, digest


@transaction.atomic
def approve_import(job: FrameworkImport, user) -> Framework:
    if job.status != FrameworkImport.Status.PREVIEW:
        raise ValueError("Only a preview awaiting approval can be imported.")
    if not job.validation_report.get("valid"):
        raise ValueError("Resolve blocking validation issues before approval.")
    meta = job.normalized_data["framework"]
    if Framework.objects.filter(code__iexact=meta["code"]).exists():
        raise ValueError("That framework code already exists; Omni never overwrites a catalog version.")
    if meta.get("is_omni_control_framework") and Framework.objects.filter(
        is_omni_control_framework=True
    ).exists():
        raise ValueError("An Omni Control Framework mapping hub is already designated.")
    framework = Framework.objects.create(
        code=meta["code"], name=meta["name"], version=meta["version"],
        authority=meta.get("authority", ""), description=meta.get("description", ""),
        source_filename=job.source_filename, source_sha256=job.source_sha256,
        is_omni_control_framework=bool(meta.get("is_omni_control_framework")),
    )
    pending_mappings = []
    for item in job.normalized_data["requirements"]:
        requirement = Requirement.objects.create(
            framework=framework, requirement_id=item["requirement_id"],
            domain=item.get("domain", "")[:100], title=item.get("title", "")[:300],
            statement=item.get("statement", ""), source_reference=item.get("source_reference", ""),
            source_page=item.get("source_page"), source_row=item.get("source_row"),
        )
        if item.get("target_requirement"):
            pending_mappings.append((requirement, item))
        for mapping in item.get("mapping_refs", []):
            pending_mappings.append((requirement, {
                **item, "target_framework": mapping["target_framework"],
                "target_requirement": mapping["target_requirement"],
            }))
    catalog_frameworks = list(Framework.objects.exclude(pk=framework.pk))
    frameworks_by_label = {}
    for catalog_framework in catalog_frameworks:
        frameworks_by_label[_key(catalog_framework.code)] = catalog_framework
        frameworks_by_label[_key(catalog_framework.name)] = catalog_framework
    target_indexes = {}
    for source, item in pending_mappings:
        target_framework = frameworks_by_label.get(_key(item.get("target_framework", "")))
        if not target_framework:
            continue
        if target_framework.pk not in target_indexes:
            target_indexes[target_framework.pk] = {
                req.requirement_id.casefold(): req
                for req in target_framework.requirements.all()
            }
        target = target_indexes[target_framework.pk].get(item["target_requirement"].casefold())
        if target:
            relationship = item.get("relationship", "RELATED").upper()
            valid = {choice for choice, _ in RequirementMapping.Relationship.choices}
            RequirementMapping.objects.create(
                source=source, target=target, relationship=relationship if relationship in valid else "RELATED",
                source_reference=item.get("source_reference", ""), approved_by=user,
                approved_at=timezone.now(),
            )
    job.status = FrameworkImport.Status.IMPORTED
    job.approved_by = user
    job.approved_at = timezone.now()
    job.imported_framework = framework
    job.save(update_fields=("status", "approved_by", "approved_at", "imported_framework"))
    materialize_mapping_references(job)
    resolve_catalog_mappings(user)
    return framework


def _mapping_tokens(raw: str) -> list[str]:
    return [token.strip() for token in re.split(r"[,;\n]+", raw or "") if token.strip()]


def materialize_mapping_references(job: FrameworkImport) -> dict:
    """Create the governed authority registry and raw cell-level mapping ledger."""
    source_index = {item.requirement_id.casefold(): item for item in job.imported_framework.requirements.all()}
    authorities, rows = {}, []
    for item in job.normalized_data.get("requirements", []):
        for column in item.get("authority_columns", []):
            name = column["name"].strip()
            base = slugify(name)[:90] or "authority"
            code, suffix = base, 2
            while ExternalAuthority.objects.exclude(canonical_name__iexact=name).filter(code=code).exists():
                code, suffix = f"{base[:85]}-{suffix}", suffix + 1
            authority, _ = ExternalAuthority.objects.get_or_create(
                canonical_name__iexact=name,
                defaults={"code": code, "canonical_name": name, "aliases": [name],
                          "source_column": column.get("source_column")},
            )
            authorities[name.casefold()] = authority
    for item in job.normalized_data.get("requirements", []):
        source = source_index.get(item.get("requirement_id", "").casefold())
        for reference in item.get("mapping_refs", []):
            name = reference["target_framework"].strip()
            key = name.casefold()
            if key not in authorities:
                base = slugify(name)[:90] or "authority"
                code, suffix = base, 2
                while ExternalAuthority.objects.exclude(canonical_name__iexact=name).filter(code=code).exists():
                    code, suffix = f"{base[:85]}-{suffix}", suffix + 1
                authority, _ = ExternalAuthority.objects.get_or_create(
                    canonical_name__iexact=name,
                    defaults={"code": code, "canonical_name": name,
                              "aliases": [name], "source_column": reference.get("source_column")},
                )
                authorities[key] = authority
            rows.append(MappingReference(
                import_job=job, source_requirement=source,
                source_requirement_id_text=item.get("requirement_id", ""),
                authority=authorities[key], raw_reference=reference["target_requirement"],
                source_row=item.get("source_row"), source_column=reference.get("source_column"),
            ))
    MappingReference.objects.bulk_create(rows, ignore_conflicts=True, batch_size=1000)
    return {"authorities": len(authorities), "references": len(rows)}


def resolve_catalog_mappings(user) -> dict:
    """Resolve retained import references after any catalog framework arrives."""
    frameworks = list(Framework.objects.prefetch_related("requirements"))
    by_label = {}
    requirement_indexes = {}
    for framework in frameworks:
        by_label[_key(framework.code)] = framework
        by_label[_key(framework.name)] = framework
        requirement_indexes[framework.pk] = {
            requirement.requirement_id.casefold(): requirement
            for requirement in framework.requirements.all()
        }
    pending, unresolved = [], 0
    jobs = FrameworkImport.objects.filter(
        status=FrameworkImport.Status.IMPORTED, imported_framework__isnull=False
    ).select_related("imported_framework")
    for job in jobs:
        source_index = requirement_indexes.get(job.imported_framework_id, {})
        for item in job.normalized_data.get("requirements", []):
            source = source_index.get(item.get("requirement_id", "").casefold())
            if not source:
                continue
            references = list(item.get("mapping_refs", []))
            if item.get("target_requirement"):
                references.append({
                    "target_framework": item.get("target_framework", ""),
                    "target_requirement": item["target_requirement"],
                })
            for reference in references:
                target_framework = by_label.get(_key(reference.get("target_framework", "")))
                if not target_framework or target_framework.pk == source.framework_id:
                    unresolved += 1
                    continue
                target_index = requirement_indexes[target_framework.pk]
                resolved_here = False
                for token in _mapping_tokens(reference.get("target_requirement", "")):
                    target = target_index.get(token.casefold())
                    if target:
                        pending.append(RequirementMapping(
                            source=source, target=target,
                            relationship=RequirementMapping.Relationship.RELATED,
                            source_reference=item.get("source_reference", job.source_filename),
                            approved_by=user, approved_at=timezone.now(),
                        ))
                        resolved_here = True
                unresolved += int(not resolved_here)
    before = RequirementMapping.objects.count()
    RequirementMapping.objects.bulk_create(pending, ignore_conflicts=True, batch_size=1000)
    ledger_updates = []
    for reference in MappingReference.objects.filter(
        status=MappingReference.Status.UNRESOLVED
    ).select_related("authority"):
        target_framework = by_label.get(_key(reference.authority.canonical_name))
        if not target_framework:
            continue
        matches = [requirement_indexes[target_framework.pk].get(token.casefold())
                   for token in _mapping_tokens(reference.raw_reference)]
        matches = [item for item in matches if item]
        if len(matches) == 1:
            reference.target_requirement = matches[0]
            reference.parsed_reference = matches[0].requirement_id
            reference.status = MappingReference.Status.RESOLVED
            ledger_updates.append(reference)
    MappingReference.objects.bulk_update(
        ledger_updates, ("target_requirement", "parsed_reference", "status"), batch_size=1000
    )
    return {"created": RequirementMapping.objects.count() - before,
            "ledger_resolved": len(ledger_updates), "unresolved": unresolved}


def version_impact(normalized: dict) -> dict:
    meta = normalized.get("framework", {})
    previous = Framework.objects.filter(name__iexact=meta.get("name", "")).order_by("-id").first()
    if not previous:
        return {"baseline": None, "added": len(normalized.get("requirements", [])),
                "changed": 0, "removed": 0}
    old = {item.requirement_id: item.statement for item in previous.requirements.all()}
    new = {item["requirement_id"]: item.get("statement", "") for item in normalized.get("requirements", [])}
    return {"baseline": f"{previous.code} {previous.version}",
            "added": len(set(new) - set(old)), "removed": len(set(old) - set(new)),
            "changed": sum(old[key] != new[key] for key in set(old) & set(new))}


def mapping_coverage(framework: Framework) -> dict:
    total = framework.requirements.count()
    mapped = framework.requirements.filter(outbound_mappings__isnull=False).distinct().count()
    return {"total": total, "mapped": mapped, "unmapped": total - mapped,
            "percentage": round(mapped * 100 / total, 1) if total else 0}
