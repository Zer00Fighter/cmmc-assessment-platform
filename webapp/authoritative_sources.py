from __future__ import annotations
import hashlib
from collections import Counter
from pathlib import Path
from urllib.parse import urlparse
from django.db import transaction
from django.utils.text import slugify
from openpyxl import load_workbook
from .models import AuthoritativeDocument, ExternalAuthority


def read_authoritative_sources(path):
    source = Path(path); raw = source.read_bytes(); digest = hashlib.sha256(raw).hexdigest()
    workbook = load_workbook(source, read_only=True, data_only=True); sheet = workbook.active
    records, excluded = [], 0
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        values = [str(value or "").strip() for value in row[:6]]
        locale, header, adi, publisher, name, url = values
        if not header or not adi or not name:
            excluded += 1; continue
        parsed = urlparse(url)
        quality = (AuthoritativeDocument.Quality.MISSING_URL if not url else
                   AuthoritativeDocument.Quality.INVALID_URL if parsed.scheme not in ("http", "https") or not parsed.netloc else
                   AuthoritativeDocument.Quality.VALID)
        records.append({"locale": locale, "ccf_column_header": header,
            "authoritative_document_id": adi, "publisher": publisher, "formal_name": name,
            "official_url": url if quality != AuthoritativeDocument.Quality.INVALID_URL else "",
            "source_url_text": url,
            "quality": quality, "source_row": row_number, "source_filename": source.name,
            "source_sha256": digest})
    workbook.close()
    duplicates = {key for key, count in Counter(x["authoritative_document_id"] for x in records).items() if count > 1}
    for item in records:
        if item["authoritative_document_id"] in duplicates:
            item["quality"] = AuthoritativeDocument.Quality.DUPLICATE_ADI
    return records, excluded


@transaction.atomic
def import_authoritative_sources(path, apply=False):
    records, excluded = read_authoritative_sources(path)
    if apply:
        for item in records:
            header = item["ccf_column_header"]
            authority = ExternalAuthority.objects.filter(canonical_name__iexact=header).first()
            if not authority:
                base = slugify(header)[:90] or "authority"; code, suffix = base, 2
                while ExternalAuthority.objects.filter(code=code).exists():
                    code, suffix = f"{base[:85]}-{suffix}", suffix + 1
                authority = ExternalAuthority.objects.create(
                    code=code, canonical_name=header, aliases=[header]
                )
            AuthoritativeDocument.objects.update_or_create(
                source_sha256=item["source_sha256"], source_row=item["source_row"],
                defaults={**item, "authority": authority},
            )
    counts = Counter(item["quality"] for item in records)
    return {"records": len(records), "authorities": len({x["ccf_column_header"] for x in records}),
            "valid": counts["VALID"], "missing_url": counts["MISSING_URL"],
            "invalid_url": counts["INVALID_URL"], "duplicate_adi": counts["DUPLICATE_ADI"],
            "excluded_rows": excluded, "applied": apply,
            "sha256": records[0]["source_sha256"] if records else ""}
