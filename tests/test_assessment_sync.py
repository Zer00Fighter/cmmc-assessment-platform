from __future__ import annotations

from openpyxl import Workbook

from src.workbook.assessment_sync import (
    AssessmentRecord,
    AssessmentSynchronizer,
)


def build_assessment_sheet():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Assessment"

    headers = [
        "Domain",
        "Requirement ID",
        "Title",
        "Requirement Statement",
        "Source Start",
        "Source End",
        "Applicable",
        "Scoring Category",
        "Status",
        "Implementation State",
        "Partial Credit Allowed",
        "Full Deduction",
        "Partial Credit Applied",
        "Calculated Deduction",
        "Evidence Status",
        "Control Owner",
    ]

    for column, header in enumerate(
        headers,
        start=1,
    ):
        worksheet.cell(
            row=5,
            column=column,
            value=header,
        )

    rows = [
        {
            "domain": "AC",
            "requirement_id": "AC.L2-3.1.1",
            "title": "Authorized Access Control",
            "applicable": "Yes",
            "status": "NOT MET",
            "implementation_state": "",
            "evidence_status": "In Progress",
            "owner": "Security Manager",
        },
        {
            "domain": "AU",
            "requirement_id": "AU.L2-3.3.1",
            "title": "System Auditing",
            "applicable": "Yes",
            "status": "MET",
            "implementation_state": "",
            "evidence_status": "Complete",
            "owner": "Audit Manager",
        },
        {
            "domain": "IA",
            "requirement_id": "IA.L2-3.5.3",
            "title": "Multifactor Authentication",
            "applicable": "Yes",
            "status": "NOT MET",
            "implementation_state": "PARTIALLY IMPLEMENTED",
            "evidence_status": "In Progress",
            "owner": "Identity Manager",
        },
        {
            "domain": "PE",
            "requirement_id": "PE.L2-3.10.1",
            "title": "Physical Access",
            "applicable": "No",
            "status": "NOT APPLICABLE",
            "implementation_state": "",
            "evidence_status": "Not Applicable",
            "owner": "Facilities Manager",
        },
    ]

    for row_number, row_data in enumerate(
        rows,
        start=6,
    ):
        worksheet.cell(
            row=row_number,
            column=1,
            value=row_data["domain"],
        )

        worksheet.cell(
            row=row_number,
            column=2,
            value=row_data["requirement_id"],
        )

        worksheet.cell(
            row=row_number,
            column=3,
            value=row_data["title"],
        )

        worksheet.cell(
            row=row_number,
            column=7,
            value=row_data["applicable"],
        )

        worksheet.cell(
            row=row_number,
            column=9,
            value=row_data["status"],
        )

        worksheet.cell(
            row=row_number,
            column=10,
            value=row_data["implementation_state"],
        )

        worksheet.cell(
            row=row_number,
            column=15,
            value=row_data["evidence_status"],
        )

        worksheet.cell(
            row=row_number,
            column=16,
            value=row_data["owner"],
        )

    return worksheet


def test_synchronize_returns_records() -> None:
    worksheet = build_assessment_sheet()
    synchronizer = AssessmentSynchronizer()

    records = synchronizer.synchronize(
        worksheet
    )

    assert len(records) == 4

    assert all(
        isinstance(record, AssessmentRecord)
        for record in records
    )


def test_first_record_values() -> None:
    worksheet = build_assessment_sheet()

    records = AssessmentSynchronizer().synchronize(
        worksheet
    )

    first = records[0]

    assert first.requirement_id == (
        "AC.L2-3.1.1"
    )
    assert first.title == (
        "Authorized Access Control"
    )
    assert first.domain == "AC"
    assert first.status == "NOT MET"
    assert first.implementation_state == ""
    assert first.applicable is True
    assert first.owner == "Security Manager"
    assert first.evidence_status == "In Progress"
    assert first.row_number == 6


def test_partial_implementation_state_is_read() -> None:
    worksheet = build_assessment_sheet()

    records = AssessmentSynchronizer().synchronize(
        worksheet
    )

    record = records[2]

    assert record.requirement_id == (
        "IA.L2-3.5.3"
    )

    assert record.implementation_state == (
        "PARTIALLY IMPLEMENTED"
    )


def test_no_value_is_not_applicable() -> None:
    worksheet = build_assessment_sheet()

    records = AssessmentSynchronizer().synchronize(
        worksheet
    )

    record = records[3]

    assert record.requirement_id == (
        "PE.L2-3.10.1"
    )
    assert record.applicable is False
    assert record.status == "NOT APPLICABLE"


def test_not_met_records_filter() -> None:
    worksheet = build_assessment_sheet()
    synchronizer = AssessmentSynchronizer()

    records = synchronizer.synchronize(
        worksheet
    )

    not_met = synchronizer.not_met_records(
        records
    )

    assert len(not_met) == 2

    assert {
        record.requirement_id
        for record in not_met
    } == {
        "AC.L2-3.1.1",
        "IA.L2-3.5.3",
    }


def test_applicable_records_filter() -> None:
    worksheet = build_assessment_sheet()
    synchronizer = AssessmentSynchronizer()

    records = synchronizer.synchronize(
        worksheet
    )

    applicable = (
        synchronizer.applicable_records(
            records
        )
    )

    assert len(applicable) == 3

    assert {
        record.requirement_id
        for record in applicable
    } == {
        "AC.L2-3.1.1",
        "AU.L2-3.3.1",
        "IA.L2-3.5.3",
    }


def test_stops_at_first_blank_requirement() -> None:
    worksheet = build_assessment_sheet()

    worksheet["B8"] = None

    records = AssessmentSynchronizer().synchronize(
        worksheet
    )

    assert len(records) == 2

    assert [
        record.requirement_id
        for record in records
    ] == [
        "AC.L2-3.1.1",
        "AU.L2-3.3.1",
    ]


def test_empty_worksheet_returns_empty_list() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Assessment"

    records = AssessmentSynchronizer().synchronize(
        worksheet
    )

    assert records == []


def test_blank_optional_values_become_empty_strings() -> None:
    worksheet = build_assessment_sheet()

    worksheet["J6"] = None
    worksheet["O6"] = None
    worksheet["P6"] = None

    record = AssessmentSynchronizer().synchronize(
        worksheet
    )[0]

    assert record.implementation_state == ""
    assert record.evidence_status == ""
    assert record.owner == ""


def test_applicable_defaults_to_true_for_blank_value() -> None:
    worksheet = build_assessment_sheet()

    worksheet["G6"] = None

    record = AssessmentSynchronizer().synchronize(
        worksheet
    )[0]

    assert record.applicable is True


def test_applicable_is_case_insensitive() -> None:
    worksheet = build_assessment_sheet()

    worksheet["G6"] = "no"

    record = AssessmentSynchronizer().synchronize(
        worksheet
    )[0]

    assert record.applicable is False


def test_record_row_numbers_are_preserved() -> None:
    worksheet = build_assessment_sheet()

    records = AssessmentSynchronizer().synchronize(
        worksheet
    )

    assert [
        record.row_number
        for record in records
    ] == [
        6,
        7,
        8,
        9,
    ]


def test_record_fields_are_independent() -> None:
    worksheet = build_assessment_sheet()

    records = AssessmentSynchronizer().synchronize(
        worksheet
    )

    assert records[0].owner == "Security Manager"
    assert records[1].owner == "Audit Manager"
    assert records[2].owner == "Identity Manager"

    assert records[0].evidence_status == "In Progress"
    assert records[1].evidence_status == "Complete"


def test_not_met_filter_accepts_iterable() -> None:
    worksheet = build_assessment_sheet()
    synchronizer = AssessmentSynchronizer()

    records = synchronizer.synchronize(
        worksheet
    )

    not_met = synchronizer.not_met_records(
        record
        for record in records
    )

    assert len(not_met) == 2


def test_applicable_filter_accepts_iterable() -> None:
    worksheet = build_assessment_sheet()
    synchronizer = AssessmentSynchronizer()

    records = synchronizer.synchronize(
        worksheet
    )

    applicable = (
        synchronizer.applicable_records(
            record
            for record in records
        )
    )

    assert len(applicable) == 3