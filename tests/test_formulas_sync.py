from __future__ import annotations

from openpyxl import Workbook

from src.workbook.formulas_sync import (
    FormulaSynchronizer,
    FormulaSyncResult,
)


def build_workbook_sheets():
    workbook = Workbook()

    assessment = workbook.active
    assessment.title = "Assessment"

    poam = workbook.create_sheet("POA&M")
    dashboard = workbook.create_sheet("Dashboard")

    return assessment, poam, dashboard


def populate_assessment_rows(
    worksheet,
    count: int,
) -> None:
    for offset in range(count):
        row = 6 + offset

        worksheet.cell(
            row=row,
            column=2,
            value=f"AC.L2-3.1.{offset + 1}",
        )

        worksheet.cell(
            row=row,
            column=9,
            value="NOT ASSESSED",
        )

        worksheet.cell(
            row=row,
            column=12,
            value=5,
        )


def populate_poam_rows(
    worksheet,
    count: int,
) -> None:
    for offset in range(count):
        row = 6 + offset

        worksheet.cell(
            row=row,
            column=1,
            value=f"POAM-{offset + 1:04d}",
        )

        worksheet.cell(
            row=row,
            column=12,
            value="Medium",
        )

        worksheet.cell(
            row=row,
            column=13,
            value="Possible",
        )


def test_synchronize_returns_result() -> None:
    assessment, poam, dashboard = (
        build_workbook_sheets()
    )

    populate_assessment_rows(
        assessment,
        2,
    )

    populate_poam_rows(
        poam,
        1,
    )

    result = FormulaSynchronizer().synchronize(
        assessment,
        poam,
        dashboard,
    )

    assert isinstance(
        result,
        FormulaSyncResult,
    )


def test_assessment_formulas_are_restored() -> None:
    assessment, poam, _ = (
        build_workbook_sheets()
    )

    populate_assessment_rows(
        assessment,
        3,
    )

    FormulaSynchronizer().synchronize(
        assessment,
        poam,
    )

    for row in range(6, 9):
        assert assessment[
            f"N{row}"
        ].value.startswith("=IF(")


def test_assessment_formula_uses_row_number() -> None:
    assessment, poam, _ = (
        build_workbook_sheets()
    )

    populate_assessment_rows(
        assessment,
        2,
    )

    FormulaSynchronizer().synchronize(
        assessment,
        poam,
    )

    assert "I6" in assessment["N6"].value
    assert "L6" in assessment["N6"].value

    assert "I7" in assessment["N7"].value
    assert "L7" in assessment["N7"].value


def test_assessment_sync_stops_at_first_blank_requirement() -> None:
    assessment, poam, _ = (
        build_workbook_sheets()
    )

    populate_assessment_rows(
        assessment,
        3,
    )

    assessment["B7"] = None
    assessment["N8"] = "MANUAL VALUE"

    result = FormulaSynchronizer().synchronize(
        assessment,
        poam,
    )

    assert result.assessment_rows_updated == 1
    assert assessment["N6"].value.startswith(
        "=IF("
    )
    assert assessment["N8"].value == (
        "MANUAL VALUE"
    )


def test_blank_assessment_returns_zero_updates() -> None:
    assessment, poam, _ = (
        build_workbook_sheets()
    )

    result = FormulaSynchronizer().synchronize(
        assessment,
        poam,
    )

    assert result.assessment_rows_updated == 0


def test_poam_formulas_are_restored() -> None:
    assessment, poam, _ = (
        build_workbook_sheets()
    )

    populate_poam_rows(
        poam,
        2,
    )

    FormulaSynchronizer().synchronize(
        assessment,
        poam,
    )

    for row in range(6, 8):
        assert poam[
            f"N{row}"
        ].value.startswith("=IF(")

        assert poam[
            f"R{row}"
        ].value.startswith("=IF(")

        assert poam[
            f"S{row}"
        ].value.startswith("=IF(")


def test_poam_formula_uses_row_number() -> None:
    assessment, poam, _ = (
        build_workbook_sheets()
    )

    populate_poam_rows(
        poam,
        1,
    )

    FormulaSynchronizer().synchronize(
        assessment,
        poam,
    )

    assert "L6" in poam["N6"].value
    assert "M6" in poam["N6"].value

    assert "O6" in poam["R6"].value
    assert "Q6" in poam["R6"].value

    assert "R6" in poam["S6"].value


def test_blank_poam_rows_are_skipped() -> None:
    assessment, poam, _ = (
        build_workbook_sheets()
    )

    poam["N6"] = "KEEP"
    poam["R6"] = "KEEP"
    poam["S6"] = "KEEP"

    result = FormulaSynchronizer().synchronize(
        assessment,
        poam,
    )

    assert result.poam_rows_updated == 0
    assert poam["N6"].value == "KEEP"
    assert poam["R6"].value == "KEEP"
    assert poam["S6"].value == "KEEP"


def test_noncontiguous_poam_rows_are_updated() -> None:
    assessment, poam, _ = (
        build_workbook_sheets()
    )

    poam["A6"] = "POAM-0001"
    poam["A8"] = "POAM-0002"

    result = FormulaSynchronizer().synchronize(
        assessment,
        poam,
    )

    assert result.poam_rows_updated == 2

    assert poam["N6"].value.startswith(
        "=IF("
    )

    assert poam["N8"].value.startswith(
        "=IF("
    )


def test_dashboard_is_updated_when_provided() -> None:
    assessment, poam, dashboard = (
        build_workbook_sheets()
    )

    result = FormulaSynchronizer().synchronize(
        assessment,
        poam,
        dashboard,
    )

    assert result.dashboard_updated is True
    assert dashboard["B2"].value == "=TODAY()"


def test_dashboard_is_optional() -> None:
    assessment, poam, _ = (
        build_workbook_sheets()
    )

    result = FormulaSynchronizer().synchronize(
        assessment,
        poam,
    )

    assert result.dashboard_updated is False


def test_result_counts_updated_rows() -> None:
    assessment, poam, dashboard = (
        build_workbook_sheets()
    )

    populate_assessment_rows(
        assessment,
        4,
    )

    populate_poam_rows(
        poam,
        3,
    )

    result = FormulaSynchronizer().synchronize(
        assessment,
        poam,
        dashboard,
    )

    assert result.assessment_rows_updated == 4
    assert result.poam_rows_updated == 3
    assert result.dashboard_updated is True


def test_manual_assessment_formula_is_replaced() -> None:
    assessment, poam, _ = (
        build_workbook_sheets()
    )

    populate_assessment_rows(
        assessment,
        1,
    )

    assessment["N6"] = "MANUAL VALUE"

    FormulaSynchronizer().synchronize(
        assessment,
        poam,
    )

    assert assessment["N6"].value.startswith(
        "=IF("
    )
    assert assessment["N6"].value != (
        "MANUAL VALUE"
    )


def test_manual_poam_formulas_are_replaced() -> None:
    assessment, poam, _ = (
        build_workbook_sheets()
    )

    populate_poam_rows(
        poam,
        1,
    )

    poam["N6"] = "MANUAL RISK"
    poam["R6"] = "MANUAL DAYS"
    poam["S6"] = "MANUAL AGING"

    FormulaSynchronizer().synchronize(
        assessment,
        poam,
    )

    assert poam["N6"].value.startswith(
        "=IF("
    )
    assert poam["R6"].value.startswith(
        "=IF("
    )
    assert poam["S6"].value.startswith(
        "=IF("
    )


def test_synchronization_is_idempotent() -> None:
    assessment, poam, dashboard = (
        build_workbook_sheets()
    )

    populate_assessment_rows(
        assessment,
        2,
    )

    populate_poam_rows(
        poam,
        2,
    )

    synchronizer = FormulaSynchronizer()

    first_result = synchronizer.synchronize(
        assessment,
        poam,
        dashboard,
    )

    first_assessment_formula = (
        assessment["N6"].value
    )
    first_risk_formula = poam["N6"].value
    first_days_formula = poam["R6"].value
    first_aging_formula = poam["S6"].value

    second_result = synchronizer.synchronize(
        assessment,
        poam,
        dashboard,
    )

    assert first_result == second_result

    assert (
        assessment["N6"].value
        == first_assessment_formula
    )

    assert poam["N6"].value == (
        first_risk_formula
    )

    assert poam["R6"].value == (
        first_days_formula
    )

    assert poam["S6"].value == (
        first_aging_formula
    )


def test_assessment_formula_content() -> None:
    formula = (
        FormulaSynchronizer
        ._assessment_formula(6)
    )

    assert formula == (
        '=IF(OR('
        'I6="MET",'
        'I6="NOT APPLICABLE",'
        'I6="NOT ASSESSED"),'
        '0,'
        'IF(AND('
        'K6="Yes",'
        'J6="PARTIALLY IMPLEMENTED"),'
        '3,'
        'L6))'
    )

    assert formula == (
        '=IF(OR('
        'I6="MET",'
        'I6="NOT APPLICABLE",'
        'I6="NOT ASSESSED"),'
        '0,'
        'IF(AND('
        'K6="Yes",'
        'J6="PARTIALLY IMPLEMENTED"),'
        '3,'
        'L6))'
    )
    
def test_risk_formula_content() -> None:
    formula = (
        FormulaSynchronizer
        ._risk_formula(6)
    )

    assert formula.startswith(
        '=IF(OR(L6="",M6=""),"",'
    )

    assert 'L6="Critical"' in formula
    assert 'M6="Almost Certain"' in formula


def test_days_open_formula_content() -> None:
    formula = (
        FormulaSynchronizer
        ._days_open_formula(6)
    )

    assert formula == (
        '=IF(O6="","",'
        'IF(Q6<>"",'
        'Q6-O6,'
        'TODAY()-O6))'
    )


def test_aging_formula_content() -> None:
    formula = (
        FormulaSynchronizer
        ._aging_formula(6)
    )

    assert formula.startswith(
        '=IF(R6="","",'
    )

    assert '"0-30 Days"' in formula
    assert '"31-60 Days"' in formula
    assert '"61-90 Days"' in formula
    assert '"91-180 Days"' in formula
    assert '"181+ Days"' in formula