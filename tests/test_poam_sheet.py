from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from src.workbook import WorkbookBuilder

ROOT = Path(__file__).resolve().parents[1]


@pytest.fixture(scope="module")
def generated_workbook(
    tmp_path_factory: pytest.TempPathFactory,
) -> Workbook:
    output_dir = tmp_path_factory.mktemp("poam_workbook")

    output_path = output_dir / "CMMC_Assessment.xlsx"

    WorkbookBuilder(
        project_root=ROOT,
        output_path=output_path,
    ).build()

    return load_workbook(
        output_path,
        data_only=False,
    )


def test_poam_sheet_exists(
    generated_workbook: Workbook,
) -> None:
    assert "POA&M" in generated_workbook.sheetnames


def test_poam_headers(
    generated_workbook: Workbook,
) -> None:
    worksheet = generated_workbook["POA&M"]

    headers = [
        worksheet.cell(
            row=5,
            column=column,
        ).value
        for column in range(1, 25)
    ]

    assert headers == [
        "Remediation ID (POA&M ID)",
        "Requirement ID",
        "Requirement Title",
        "Domain",
        "Weakness Description",
        "Root Cause",
        "Corrective Action",
        "Current Milestone",
        "Milestone Owner",
        "Status",
        "Priority",
        "Severity",
        "Likelihood",
        "Risk Score",
        "Date Identified",
        "Planned Completion",
        "Actual Completion",
        "Days Open",
        "Aging Bucket",
        "Residual Risk",
        "Validation Status",
        "Evidence IDs",
        "Security Plan Reference (SSP)",
        "Assessor Notes",
    ]


def test_poam_has_300_rows(
    generated_workbook: Workbook,
) -> None:
    worksheet = generated_workbook["POA&M"]

    assert worksheet.max_row >= 305


def test_default_values(
    generated_workbook: Workbook,
) -> None:
    worksheet = generated_workbook["POA&M"]

    assert worksheet["J6"].value == "Open"
    assert worksheet["K6"].value == "Medium"
    assert worksheet["L6"].value == "Medium"
    assert worksheet["M6"].value == "Possible"
    assert worksheet["T6"].value == "Medium"
    assert worksheet["U6"].value == "Pending"


def test_formula_columns(
    generated_workbook: Workbook,
) -> None:
    worksheet = generated_workbook["POA&M"]

    assert worksheet["N6"].value.startswith("=IF(")
    assert worksheet["R6"].value.startswith("=IF(")
    assert worksheet["S6"].value.startswith("=IF(")


def test_date_formats(
    generated_workbook: Workbook,
) -> None:
    worksheet = generated_workbook["POA&M"]

    assert worksheet["O6"].number_format == "mm/dd/yyyy"
    assert worksheet["P6"].number_format == "mm/dd/yyyy"
    assert worksheet["Q6"].number_format == "mm/dd/yyyy"


def test_validations_exist(
    generated_workbook: Workbook,
) -> None:
    worksheet = generated_workbook["POA&M"]

    validation_ranges = {
        str(v.sqref) for v in worksheet.data_validations.dataValidation
    }

    assert "J6:J305" in validation_ranges
    assert "K6:K305" in validation_ranges
    assert "L6:L305" in validation_ranges
    assert "M6:M305" in validation_ranges
    assert "T6:T305" in validation_ranges
    assert "U6:U305" in validation_ranges
    assert "O6:Q305" in validation_ranges


def test_filter_and_freeze(
    generated_workbook: Workbook,
) -> None:
    worksheet = generated_workbook["POA&M"]

    assert worksheet.freeze_panes == "A6"
    assert worksheet.auto_filter.ref == "A5:X305"


def test_print_area(
    generated_workbook: Workbook,
) -> None:
    worksheet = generated_workbook["POA&M"]

    assert "$A$1:$X$305" in str(worksheet.print_area)


def test_conditional_formatting_exists(
    generated_workbook: Workbook,
) -> None:
    worksheet = generated_workbook["POA&M"]

    formatting = {str(item) for item in worksheet.conditional_formatting}

    assert any("J6:J305" in value for value in formatting)

    assert any("N6:N305" in value for value in formatting)

    assert any("P6:P305" in value for value in formatting)

    assert any("S6:S305" in value for value in formatting)


def test_locked_and_unlocked_cells(
    generated_workbook: Workbook,
) -> None:
    worksheet = generated_workbook["POA&M"]

    assert worksheet["E6"].protection.locked is False
    assert worksheet["F6"].protection.locked is False
    assert worksheet["G6"].protection.locked is False

    assert worksheet["N6"].protection.locked is True
    assert worksheet["R6"].protection.locked is True
    assert worksheet["S6"].protection.locked is True
