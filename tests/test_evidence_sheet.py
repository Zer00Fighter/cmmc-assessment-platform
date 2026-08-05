from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from src.workbook import WorkbookBuilder


ROOT = Path(__file__).resolve().parents[1]


@pytest.fixture(scope="module")
def generated_workbook(
    tmp_path_factory: pytest.TempPathFactory,
) -> Workbook:
    output_directory = tmp_path_factory.mktemp(
        "evidence_workbook"
    )

    output_path = (
        output_directory
        / "CMMC_Assessment.xlsx"
    )

    builder = WorkbookBuilder(
        project_root=ROOT,
        output_path=output_path,
    )

    builder.build()

    return load_workbook(
        output_path,
        data_only=False,
    )


def test_evidence_sheet_exists(
    generated_workbook: Workbook,
) -> None:
    assert "Evidence" in (
        generated_workbook.sheetnames
    )


def test_evidence_sheet_has_expected_headers(
    generated_workbook: Workbook,
) -> None:
    worksheet = generated_workbook["Evidence"]

    headers = [
        worksheet.cell(
            row=5,
            column=column,
        ).value
        for column in range(1, 17)
    ]

    assert headers == [
        "Evidence ID",
        "Evidence Title",
        "Evidence Type",
        "Description",
        "Storage Location / URL",
        "Document Owner",
        "Evidence Status",
        "Review Status",
        "Reviewer",
        "Review Date",
        "Expiration Date",
        "Requirement IDs",
        "Objective IDs",
        "Confidentiality",
        "Version",
        "Notes",
    ]


def test_evidence_sheet_has_500_entry_rows(
    generated_workbook: Workbook,
) -> None:
    worksheet = generated_workbook["Evidence"]

    first_data_row = 6
    last_data_row = 505

    assert (
        last_data_row
        - first_data_row
        + 1
    ) == 500

    assert worksheet.max_row >= last_data_row


def test_evidence_sheet_default_values(
    generated_workbook: Workbook,
) -> None:
    worksheet = generated_workbook["Evidence"]

    assert worksheet["A6"].value in {
        None,
        "",
    }

    assert worksheet["B6"].value in {
        None,
        "",
    }

    assert worksheet["C6"].value == "Document"
    assert worksheet["G6"].value == "Not Started"
    assert worksheet["H6"].value == "Not Reviewed"
    assert worksheet["N6"].value == "CUI"


def test_evidence_sheet_date_formats(
    generated_workbook: Workbook,
) -> None:
    worksheet = generated_workbook["Evidence"]

    assert worksheet["J6"].number_format == (
        "mm/dd/yyyy"
    )

    assert worksheet["K6"].number_format == (
        "mm/dd/yyyy"
    )


def test_evidence_sheet_has_expected_validations(
    generated_workbook: Workbook,
) -> None:
    worksheet = generated_workbook["Evidence"]

    validations = list(
        worksheet
        .data_validations
        .dataValidation
    )

    validation_ranges = {
        str(validation.sqref)
        for validation in validations
    }

    assert "C6:C505" in validation_ranges
    assert "G6:G505" in validation_ranges
    assert "H6:H505" in validation_ranges
    assert "J6:K505" in validation_ranges
    assert "N6:N505" in validation_ranges


def test_evidence_type_validation_values(
    generated_workbook: Workbook,
) -> None:
    worksheet = generated_workbook["Evidence"]

    validation = next(
        item
        for item in (
            worksheet
            .data_validations
            .dataValidation
        )
        if str(item.sqref) == "C6:C505"
    )

    assert validation.type == "list"
    assert "Policy" in validation.formula1
    assert "Procedure" in validation.formula1
    assert "Screenshot" in validation.formula1
    assert "Test Result" in validation.formula1


def test_evidence_status_validation_values(
    generated_workbook: Workbook,
) -> None:
    worksheet = generated_workbook["Evidence"]

    validation = next(
        item
        for item in (
            worksheet
            .data_validations
            .dataValidation
        )
        if str(item.sqref) == "G6:G505"
    )

    assert validation.type == "list"
    assert "Not Started" in validation.formula1
    assert "In Progress" in validation.formula1
    assert "Complete" in validation.formula1
    assert "Expired" in validation.formula1


def test_review_status_validation_values(
    generated_workbook: Workbook,
) -> None:
    worksheet = generated_workbook["Evidence"]

    validation = next(
        item
        for item in (
            worksheet
            .data_validations
            .dataValidation
        )
        if str(item.sqref) == "H6:H505"
    )

    assert validation.type == "list"
    assert "Not Reviewed" in validation.formula1
    assert "Approved" in validation.formula1
    assert "Rejected" in validation.formula1
    assert "Needs Update" in validation.formula1


def test_confidentiality_validation_values(
    generated_workbook: Workbook,
) -> None:
    worksheet = generated_workbook["Evidence"]

    validation = next(
        item
        for item in (
            worksheet
            .data_validations
            .dataValidation
        )
        if str(item.sqref) == "N6:N505"
    )

    assert validation.type == "list"
    assert "Public" in validation.formula1
    assert "Confidential" in validation.formula1
    assert "CUI" in validation.formula1
    assert "Restricted" in validation.formula1


def test_date_validation_configuration(
    generated_workbook: Workbook,
) -> None:
    worksheet = generated_workbook["Evidence"]

    validation = next(
        item
        for item in (
            worksheet
            .data_validations
            .dataValidation
        )
        if str(item.sqref) == "J6:K505"
    )

    assert validation.type == "date"
    assert validation.operator == "between"
    assert validation.formula1 == (
        "DATE(2000,1,1)"
    )
    assert validation.formula2 == (
        "DATE(2100,12,31)"
    )


def test_evidence_sheet_has_conditional_formatting(
    generated_workbook: Workbook,
) -> None:
    worksheet = generated_workbook["Evidence"]

    conditional_ranges = {
        str(item)
        for item in worksheet.conditional_formatting
    }

    assert any(
        "G6:G505" in item
        for item in conditional_ranges
    )

    assert any(
        "H6:H505" in item
        for item in conditional_ranges
    )

    assert any(
        "K6:K505" in item
        for item in conditional_ranges
    )


def test_evidence_sheet_has_filter_and_freeze_panes(
    generated_workbook: Workbook,
) -> None:
    worksheet = generated_workbook["Evidence"]

    assert worksheet.freeze_panes == "A6"

    assert worksheet.auto_filter.ref == (
        "A5:P505"
    )


def test_evidence_sheet_print_settings(
    generated_workbook: Workbook,
) -> None:
    worksheet = generated_workbook["Evidence"]

    assert worksheet.print_title_rows == (
        "$1:$5"
    )

    assert "$A$1:$P$505" in str(
        worksheet.print_area
    )

    assert worksheet.page_setup.orientation == (
        "landscape"
    )


def test_evidence_sheet_column_widths(
    generated_workbook: Workbook,
) -> None:
    worksheet = generated_workbook["Evidence"]

    assert (
        worksheet.column_dimensions["A"].width
        == 16
    )

    assert (
        worksheet.column_dimensions["B"].width
        == 34
    )

    assert (
        worksheet.column_dimensions["E"].width
        == 48
    )

    assert (
        worksheet.column_dimensions["P"].width
        == 45
    )


def test_evidence_cells_are_unlocked_for_input(
    generated_workbook: Workbook,
) -> None:
    worksheet = generated_workbook["Evidence"]

    for cell_reference in [
        "A6",
        "B6",
        "C6",
        "G6",
        "H6",
        "J6",
        "K6",
        "N6",
        "P6",
    ]:
        assert (
            worksheet[
                cell_reference
            ].protection.locked
            is False
        )