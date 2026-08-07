from __future__ import annotations

from openpyxl import Workbook

from src.workbook.workbook_pipeline import (
    WorkbookPipeline,
    WorkbookPipelineError,
    WorkbookPipelineResult,
)


def build_workbook():
    workbook = Workbook()

    assessment = workbook.active
    assessment.title = "Assessment"

    evidence = workbook.create_sheet("Evidence")
    poam = workbook.create_sheet("POA&M")
    dashboard = workbook.create_sheet("Dashboard")

    #
    # Assessment
    #
    assessment["A6"] = "AC"
    assessment["B6"] = "AC.L2-3.1.1"
    assessment["C6"] = "Access Control"
    assessment["G6"] = "Yes"
    assessment["I6"] = "NOT MET"
    assessment["J6"] = ""
    assessment["P6"] = "Security"

    assessment["A7"] = "AU"
    assessment["B7"] = "AU.L2-3.3.1"
    assessment["C7"] = "Audit Logging"
    assessment["G7"] = "Yes"
    assessment["I7"] = "MET"

    #
    # Evidence
    #
    evidence["A6"] = "EV-001"
    evidence["B6"] = "Access Policy"
    evidence["C6"] = "Policy"
    evidence["F6"] = "Security"
    evidence["G6"] = "Complete"
    evidence["L6"] = "AC.L2-3.1.1"

    return workbook


def test_pipeline_runs() -> None:
    workbook = build_workbook()

    result = WorkbookPipeline().run(workbook)

    assert isinstance(
        result,
        WorkbookPipelineResult,
    )


def test_assessment_records_loaded() -> None:
    workbook = build_workbook()

    result = WorkbookPipeline().run(workbook)

    assert result.assessment_count == 2


def test_evidence_records_loaded() -> None:
    workbook = build_workbook()

    result = WorkbookPipeline().run(workbook)

    assert result.evidence_count == 1


def test_evidence_coverage() -> None:
    workbook = build_workbook()

    result = WorkbookPipeline().run(workbook)

    assert (
        result.evidence_coverage.covered_requirements
        == 1
    )

    assert (
        result.evidence_coverage.uncovered_requirements
        == 1
    )


def test_poam_created() -> None:
    workbook = build_workbook()

    result = WorkbookPipeline().run(workbook)

    assert result.poam_count == 1

    poam = workbook["POA&M"]

    assert poam["B6"].value == "AC.L2-3.1.1"


def test_dashboard_written() -> None:
    workbook = build_workbook()

    WorkbookPipeline().run(workbook)

    dashboard = workbook["Dashboard"]

    assert dashboard["B18"].value == (
        "Synchronized Metrics"
    )

    assert dashboard["C19"].value == 2


def test_formulas_restored() -> None:
    workbook = build_workbook()

    WorkbookPipeline().run(workbook)

    assessment = workbook["Assessment"]
    poam = workbook["POA&M"]

    assert assessment["N6"].value.startswith("=IF(")
    assert poam["N6"].value.startswith("=IF(")
    assert poam["R6"].value.startswith("=IF(")
    assert poam["S6"].value.startswith("=IF(")


def test_recalculation_enabled() -> None:
    workbook = build_workbook()

    WorkbookPipeline().run(workbook)

    assert workbook.calculation.fullCalcOnLoad
    assert workbook.calculation.forceFullCalc
    assert workbook.calculation.calcMode == "auto"


def test_missing_sheet_raises_error() -> None:
    workbook = Workbook()

    workbook.active.title = "Assessment"

    workbook.create_sheet("Evidence")
    workbook.create_sheet("POA&M")

    try:
        WorkbookPipeline().run(workbook)
        assert False
    except WorkbookPipelineError:
        pass


def test_pipeline_is_idempotent() -> None:
    workbook = build_workbook()

    pipeline = WorkbookPipeline()

    first = pipeline.run(workbook)
    second = pipeline.run(workbook)

    assert (
        first.assessment_count
        == second.assessment_count
    )

    assert (
        first.poam_count
        == second.poam_count
    )

    assert workbook["POA&M"]["B6"].value == (
        "AC.L2-3.1.1"
    )


def test_dashboard_metrics_written() -> None:
    workbook = build_workbook()

    WorkbookPipeline().run(workbook)

    dashboard = workbook["Dashboard"]

    assert dashboard["C21"].value == 1
    assert dashboard["C22"].value == 1

    assert dashboard["F19"].value == 1

    assert dashboard["I19"].value == 1


def test_pipeline_alias() -> None:
    workbook = build_workbook()

    pipeline = WorkbookPipeline()

    assert (
        pipeline.run(workbook).assessment_count
        ==
        pipeline.synchronize(workbook).assessment_count
    )