from __future__ import annotations

from openpyxl import Workbook

from src.workbook.evidence_sync import (
    EvidenceCoverage,
    EvidenceRecord,
    EvidenceSynchronizer,
)


def build_evidence_sheet():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Evidence"

    headers = [
        "Evidence ID",
        "Evidence Title",
        "Evidence Type",
        "Description",
        "Storage Location / URL",
        "Document Owner",
        "Evidence Status",
        "Review Status",
        "Reviewer",
        "Review Date",
        "Expiration Date",
        "Requirement IDs",
        "Objective IDs",
        "Confidentiality",
        "Version",
        "Notes",
    ]

    for column, header in enumerate(
        headers,
        start=1,
    ):
        worksheet.cell(
            row=5,
            column=column,
            value=header,
        )

    rows = [
        {
            "evidence_id": "EV-001",
            "title": "Access Control Policy",
            "evidence_type": "Policy",
            "owner": "Security Manager",
            "status": "Complete",
            "requirements": (
                "AC.L2-3.1.1, AC.L2-3.1.2"
            ),
            "objectives": "a, b",
        },
        {
            "evidence_id": "EV-002",
            "title": "Audit Configuration",
            "evidence_type": "Configuration",
            "owner": "Audit Manager",
            "status": "In Progress",
            "requirements": "AU.L2-3.3.1",
            "objectives": "a",
        },
        {
            "evidence_id": "EV-003",
            "title": "Orphaned Evidence",
            "evidence_type": "Document",
            "owner": "Security Team",
            "status": "Not Started",
            "requirements": "",
            "objectives": "",
        },
    ]

    for row_number, row_data in enumerate(
        rows,
        start=6,
    ):
        worksheet.cell(
            row=row_number,
            column=1,
            value=row_data["evidence_id"],
        )

        worksheet.cell(
            row=row_number,
            column=2,
            value=row_data["title"],
        )

        worksheet.cell(
            row=row_number,
            column=3,
            value=row_data["evidence_type"],
        )

        worksheet.cell(
            row=row_number,
            column=6,
            value=row_data["owner"],
        )

        worksheet.cell(
            row=row_number,
            column=7,
            value=row_data["status"],
        )

        worksheet.cell(
            row=row_number,
            column=12,
            value=row_data["requirements"],
        )

        worksheet.cell(
            row=row_number,
            column=13,
            value=row_data["objectives"],
        )

    return worksheet


def test_synchronize_reads_all_evidence_rows() -> None:
    worksheet = build_evidence_sheet()

    records = EvidenceSynchronizer().synchronize(
        worksheet
    )

    assert len(records) == 3

    assert all(
        isinstance(record, EvidenceRecord)
        for record in records
    )


def test_first_record_values() -> None:
    worksheet = build_evidence_sheet()

    record = EvidenceSynchronizer().synchronize(
        worksheet
    )[0]

    assert record.evidence_id == "EV-001"
    assert record.title == "Access Control Policy"
    assert record.evidence_type == "Policy"
    assert record.status == "Complete"
    assert record.owner == "Security Manager"
    assert record.row_number == 6


def test_multiple_requirement_ids_are_parsed() -> None:
    worksheet = build_evidence_sheet()

    record = EvidenceSynchronizer().synchronize(
        worksheet
    )[0]

    assert record.requirement_ids == [
        "AC.L2-3.1.1",
        "AC.L2-3.1.2",
    ]


def test_objective_ids_are_parsed_and_normalized() -> None:
    worksheet = build_evidence_sheet()

    record = EvidenceSynchronizer().synchronize(
        worksheet
    )[0]

    assert record.objective_ids == [
        "A",
        "B",
    ]


def test_blank_requirement_and_objective_lists() -> None:
    worksheet = build_evidence_sheet()

    record = EvidenceSynchronizer().synchronize(
        worksheet
    )[2]

    assert record.requirement_ids == []
    assert record.objective_ids == []


def test_blank_optional_fields_become_empty_strings() -> None:
    worksheet = build_evidence_sheet()

    worksheet["B6"] = None
    worksheet["C6"] = None
    worksheet["F6"] = None
    worksheet["G6"] = None

    record = EvidenceSynchronizer().synchronize(
        worksheet
    )[0]

    assert record.title == ""
    assert record.evidence_type == ""
    assert record.owner == ""
    assert record.status == ""


def test_stops_at_first_blank_evidence_id() -> None:
    worksheet = build_evidence_sheet()

    worksheet["A7"] = None

    records = EvidenceSynchronizer().synchronize(
        worksheet
    )

    assert len(records) == 1
    assert records[0].evidence_id == "EV-001"


def test_empty_worksheet_returns_empty_list() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Evidence"

    records = EvidenceSynchronizer().synchronize(
        worksheet
    )

    assert records == []


def test_row_numbers_are_preserved() -> None:
    worksheet = build_evidence_sheet()

    records = EvidenceSynchronizer().synchronize(
        worksheet
    )

    assert [
        record.row_number
        for record in records
    ] == [
        6,
        7,
        8,
    ]


def test_coverage_calculation() -> None:
    worksheet = build_evidence_sheet()
    synchronizer = EvidenceSynchronizer()

    records = synchronizer.synchronize(
        worksheet
    )

    coverage = synchronizer.calculate_coverage(
        records,
        [
            "AC.L2-3.1.1",
            "AC.L2-3.1.2",
            "AU.L2-3.3.1",
            "IA.L2-3.5.1",
        ],
    )

    assert isinstance(
        coverage,
        EvidenceCoverage,
    )

    assert coverage.total_evidence == 3
    assert coverage.covered_requirements == 3
    assert coverage.uncovered_requirements == 1
    assert coverage.orphaned_evidence == 1
    assert coverage.duplicate_evidence == 0


def test_duplicate_evidence_ids_are_counted() -> None:
    worksheet = build_evidence_sheet()

    worksheet["A8"] = "EV-001"
    worksheet["L8"] = "IA.L2-3.5.1"

    synchronizer = EvidenceSynchronizer()

    records = synchronizer.synchronize(
        worksheet
    )

    coverage = synchronizer.calculate_coverage(
        records,
        [
            "AC.L2-3.1.1",
            "AC.L2-3.1.2",
            "AU.L2-3.3.1",
            "IA.L2-3.5.1",
        ],
    )

    assert coverage.duplicate_evidence == 1


def test_multiple_duplicate_occurrences_are_counted() -> None:
    worksheet = build_evidence_sheet()

    worksheet["A7"] = "EV-001"
    worksheet["A8"] = "EV-001"

    records = EvidenceSynchronizer().synchronize(
        worksheet
    )

    coverage = (
        EvidenceSynchronizer()
        .calculate_coverage(
            records,
            ["AC.L2-3.1.1"],
        )
    )

    assert coverage.duplicate_evidence == 2


def test_orphaned_evidence_is_counted() -> None:
    worksheet = build_evidence_sheet()

    records = EvidenceSynchronizer().synchronize(
        worksheet
    )

    coverage = (
        EvidenceSynchronizer()
        .calculate_coverage(
            records,
            [
                "AC.L2-3.1.1",
                "AC.L2-3.1.2",
                "AU.L2-3.3.1",
            ],
        )
    )

    assert coverage.orphaned_evidence == 1


def test_all_requirements_covered() -> None:
    worksheet = build_evidence_sheet()

    records = EvidenceSynchronizer().synchronize(
        worksheet
    )

    coverage = (
        EvidenceSynchronizer()
        .calculate_coverage(
            records,
            [
                "AC.L2-3.1.1",
                "AC.L2-3.1.2",
                "AU.L2-3.3.1",
            ],
        )
    )

    assert coverage.covered_requirements == 3
    assert coverage.uncovered_requirements == 0


def test_no_evidence_means_all_requirements_uncovered() -> None:
    coverage = (
        EvidenceSynchronizer()
        .calculate_coverage(
            [],
            [
                "AC.L2-3.1.1",
                "AU.L2-3.3.1",
            ],
        )
    )

    assert coverage.total_evidence == 0
    assert coverage.covered_requirements == 0
    assert coverage.uncovered_requirements == 2
    assert coverage.orphaned_evidence == 0
    assert coverage.duplicate_evidence == 0


def test_requirement_matching_is_case_insensitive() -> None:
    worksheet = build_evidence_sheet()

    worksheet["L6"] = (
        "ac.l2-3.1.1, ac.l2-3.1.2"
    )

    records = EvidenceSynchronizer().synchronize(
        worksheet
    )

    coverage = (
        EvidenceSynchronizer()
        .calculate_coverage(
            records,
            [
                "AC.L2-3.1.1",
                "AC.L2-3.1.2",
            ],
        )
    )

    assert coverage.covered_requirements == 2
    assert coverage.uncovered_requirements == 0


def test_unknown_requirement_references_do_not_count_as_covered() -> None:
    worksheet = build_evidence_sheet()

    worksheet["L6"] = "XX.L2-3.99.99"

    records = EvidenceSynchronizer().synchronize(
        worksheet
    )

    coverage = (
        EvidenceSynchronizer()
        .calculate_coverage(
            records,
            [
                "AC.L2-3.1.1",
                "AU.L2-3.3.1",
            ],
        )
    )

    assert coverage.covered_requirements == 1
    assert coverage.uncovered_requirements == 1


def test_duplicate_requirement_references_do_not_inflate_coverage() -> None:
    worksheet = build_evidence_sheet()

    worksheet["L6"] = (
        "AC.L2-3.1.1, AC.L2-3.1.1"
    )

    records = EvidenceSynchronizer().synchronize(
        worksheet
    )

    coverage = (
        EvidenceSynchronizer()
        .calculate_coverage(
            records,
            [
                "AC.L2-3.1.1",
                "AU.L2-3.3.1",
            ],
        )
    )

    assert coverage.covered_requirements == 2
    assert coverage.uncovered_requirements == 0


def test_calculate_coverage_accepts_iterables() -> None:
    worksheet = build_evidence_sheet()
    synchronizer = EvidenceSynchronizer()

    records = synchronizer.synchronize(
        worksheet
    )

    coverage = synchronizer.calculate_coverage(
        (
            record
            for record in records
        ),
        (
            requirement
            for requirement in [
                "AC.L2-3.1.1",
                "AC.L2-3.1.2",
                "AU.L2-3.3.1",
            ]
        ),
    )

    assert coverage.total_evidence == 3
    assert coverage.covered_requirements == 3