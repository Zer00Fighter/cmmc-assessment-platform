from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from src.evidence_requests.assessment_procedure_loader import (
    AssessmentProcedureDataset,
    AssessmentProcedureLoader,
    AssessmentProcedureLoaderError,
)


HEADERS = [
    "Family",
    "Identifier",
    "Sort-As",
    "Security Requirement",
    "Assessment Objective",
    "Potential Assessment Method and Objects: Examine",
    "Potential Assessment Method and Objects: Interview",
    "Potential Assessment Method and Objects: Test",
]


def build_test_workbook(
    path: Path,
) -> Path:
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "SP800-171A"

    worksheet.append(HEADERS)

    worksheet.append(
        [
            "AC",
            "3.1.1",
            "3.1.1",
            (
                "Limit system access to authorized users, "
                "processes acting on behalf of authorized "
                "users, and devices."
            ),
            "Determine if:",
            (
                "[SELECT FROM: "
                "access control policy; "
                "system security plan]"
            ),
            "system administrators",
            (
                "mechanisms implementing "
                "access controls"
            ),
        ]
    )

    worksheet.append(
        [
            "AC",
            "3.1.1[a]",
            "3.1.1[a]",
            "",
            (
                "authorized users are identified"
            ),
            "",
            "",
            "",
        ]
    )

    worksheet.append(
        [
            "AC",
            "3.1.1[b]",
            "3.1.1[b]",
            "",
            (
                "processes acting on behalf "
                "of authorized users are identified"
            ),
            "",
            "",
            "",
        ]
    )

    worksheet.append(
        [
            "IA",
            "3.5.3",
            "3.5.3",
            (
                "Use multifactor authentication "
                "for local and network access."
            ),
            "Determine if:",
            (
                "identification and "
                "authentication policy; "
                "system security plan"
            ),
            (
                "personnel with identification "
                "and authentication responsibilities"
            ),
            (
                "mechanisms implementing "
                "multifactor authentication"
            ),
        ]
    )

    worksheet.append(
        [
            "IA",
            "3.5.3[a]",
            "3.5.3[a]",
            "",
            (
                "multifactor authentication "
                "is implemented"
            ),
            "",
            "",
            "",
        ]
    )

    workbook.save(path)

    return path


@pytest.fixture
def loader() -> AssessmentProcedureLoader:
    return AssessmentProcedureLoader(
        framework_id="CMMC_L2",
        framework_name="CMMC Level 2",
        framework_version="2.0",
        source_document="NIST SP 800-171A",
        source_revision="Rev. 2",
    )


def test_dataset_requires_framework() -> None:
    with pytest.raises(
        AssessmentProcedureLoaderError
    ):
        AssessmentProcedureDataset(
            framework_id=""
        )


def test_loader_requires_framework() -> None:
    with pytest.raises(
        AssessmentProcedureLoaderError
    ):
        AssessmentProcedureLoader(
            framework_id=""
        )


def test_loader_requires_sheet_name() -> None:
    with pytest.raises(
        AssessmentProcedureLoaderError
    ):
        AssessmentProcedureLoader(
            framework_id="CMMC_L2",
            sheet_name="",
        )


def test_canonical_ac_requirement_id() -> None:
    assert (
        AssessmentProcedureLoader
        ._canonical_requirement_id(
            "AC",
            "3.1.1",
        )
        == "AC.L2-3.1.1"
    )


def test_canonical_ia_requirement_id() -> None:
    assert (
        AssessmentProcedureLoader
        ._canonical_requirement_id(
            "IA",
            "3.5.3",
        )
        == "IA.L2-3.5.3"
    )


def test_canonical_id_does_not_double_prefix() -> None:
    assert (
        AssessmentProcedureLoader
        ._canonical_requirement_id(
            "AC",
            "AC.L2-3.1.1",
        )
        == "AC.L2-3.1.1"
    )


def test_canonical_id_requires_family() -> None:
    with pytest.raises(
        AssessmentProcedureLoaderError
    ):
        (
            AssessmentProcedureLoader
            ._canonical_requirement_id(
                "",
                "3.1.1",
            )
        )


def test_parse_objective_identifier_with_extra_dot() -> None:
    result = (
        AssessmentProcedureLoader
        ._parse_identifier(
            "3.12.4.[h]"
        )
    )

    assert result == (
        "3.12.4",
        "h",
    )


def test_parse_objective_identifier_normalizes_parent() -> None:
    normal = (
        AssessmentProcedureLoader
        ._parse_identifier(
            "3.12.4[h]"
        )
    )

    dotted = (
        AssessmentProcedureLoader
        ._parse_identifier(
            "3.12.4.[h]"
        )
    )

    assert normal == dotted

    assert normal == (
        "3.12.4",
        "h",
    )

def test_parse_objective_identifier_normalizes_parent() -> None:
    normal = (
        AssessmentProcedureLoader
        ._parse_identifier(
            "3.12.4[h]"
        )
    )

    dotted = (
        AssessmentProcedureLoader
        ._parse_identifier(
            "3.12.4.[h]"
        )
    )

    assert normal == dotted

    assert normal == (
        "3.12.4",
        "h",
    )

def test_clean_determine_if_is_blank() -> None:
    assert (
        AssessmentProcedureLoader
        ._clean_objective_text(
            "Determine if:"
        )
        == ""
    )


def test_load_dataset(
    tmp_path: Path,
    loader: AssessmentProcedureLoader,
) -> None:
    path = build_test_workbook(
        tmp_path
        / "procedures.xlsx"
    )

    dataset = loader.load(
        path
    )

    assert isinstance(
        dataset,
        AssessmentProcedureDataset,
    )

    assert dataset.framework_id == "CMMC_L2"
    assert dataset.source_sheet == "SP800-171A"
    assert dataset.source_file == "procedures.xlsx"


def test_dataset_requirement_count(
    tmp_path: Path,
    loader: AssessmentProcedureLoader,
) -> None:
    dataset = loader.load(
        build_test_workbook(
            tmp_path
            / "procedures.xlsx"
        )
    )

    assert dataset.requirement_count == 2


def test_dataset_objective_count(
    tmp_path: Path,
    loader: AssessmentProcedureLoader,
) -> None:
    dataset = loader.load(
        build_test_workbook(
            tmp_path
            / "procedures.xlsx"
        )
    )

    assert dataset.objective_count == 3


def test_dataset_uses_canonical_control_ids(
    tmp_path: Path,
    loader: AssessmentProcedureLoader,
) -> None:
    dataset = loader.load(
        build_test_workbook(
            tmp_path
            / "procedures.xlsx"
        )
    )

    assert dataset.requirement_ids == (
        "AC.L2-3.1.1",
        "IA.L2-3.5.3",
    )


def test_base_requirement_row_created(
    tmp_path: Path,
    loader: AssessmentProcedureLoader,
) -> None:
    dataset = loader.load(
        build_test_workbook(
            tmp_path
            / "procedures.xlsx"
        )
    )

    rows = dataset.for_requirement(
        "AC.L2-3.1.1"
    )

    base = rows[0]

    assert base.requirement_id == (
        "AC.L2-3.1.1"
    )

    assert base.objective_id == ""

    assert "access control policy" in (
        base.examine
    )


def test_objective_rows_inherit_examine(
    tmp_path: Path,
    loader: AssessmentProcedureLoader,
) -> None:
    dataset = loader.load(
        build_test_workbook(
            tmp_path
            / "procedures.xlsx"
        )
    )

    rows = dataset.for_requirement(
        "AC.L2-3.1.1"
    )

    objective_rows = [
        row
        for row in rows
        if row.objective_id
    ]

    assert len(objective_rows) == 2

    for row in objective_rows:
        assert (
            "access control policy"
            in row.examine
        )

        assert (
            "system security plan"
            in row.examine
        )


def test_objective_rows_inherit_interview(
    tmp_path: Path,
    loader: AssessmentProcedureLoader,
) -> None:
    dataset = loader.load(
        build_test_workbook(
            tmp_path
            / "procedures.xlsx"
        )
    )

    rows = dataset.for_requirement(
        "AC.L2-3.1.1"
    )

    objective = rows[1]

    assert (
        objective.interview
        == "system administrators"
    )


def test_objective_rows_inherit_test(
    tmp_path: Path,
    loader: AssessmentProcedureLoader,
) -> None:
    dataset = loader.load(
        build_test_workbook(
            tmp_path
            / "procedures.xlsx"
        )
    )

    rows = dataset.for_requirement(
        "AC.L2-3.1.1"
    )

    objective = rows[1]

    assert (
        "mechanisms implementing access controls"
        in objective.test
    )


def test_for_requirement_is_case_insensitive(
    tmp_path: Path,
    loader: AssessmentProcedureLoader,
) -> None:
    dataset = loader.load(
        build_test_workbook(
            tmp_path
            / "procedures.xlsx"
        )
    )

    assert len(
        dataset.for_requirement(
            "ac.l2-3.1.1"
        )
    ) == 3


def test_source_location_preserved(
    tmp_path: Path,
    loader: AssessmentProcedureLoader,
) -> None:
    dataset = loader.load(
        build_test_workbook(
            tmp_path
            / "procedures.xlsx"
        )
    )

    first = dataset.rows[0]

    assert "SP800-171A!Row 2" in (
        first.source_location
    )

    assert "3.1.1" in (
        first.source_location
    )


def test_title_mapper(
    tmp_path: Path,
) -> None:
    loader = AssessmentProcedureLoader(
        framework_id="CMMC_L2",
        requirement_title_mapper=(
            lambda requirement_id,
            family,
            requirement_text:
            (
                "Authorized Access Control"
                if requirement_id
                == "AC.L2-3.1.1"
                else "Other Control"
            )
        ),
    )

    dataset = loader.load(
        build_test_workbook(
            tmp_path
            / "procedures.xlsx"
        )
    )

    row = dataset.for_requirement(
        "AC.L2-3.1.1"
    )[0]

    assert (
        row.requirement_title
        == "Authorized Access Control"
    )


def test_sprs_provider(
    tmp_path: Path,
) -> None:
    loader = AssessmentProcedureLoader(
        framework_id="CMMC_L2",
        sprs_weight_provider=(
            lambda requirement_id:
            (
                5
                if requirement_id
                == "AC.L2-3.1.1"
                else 3
            )
        ),
    )

    dataset = loader.load(
        build_test_workbook(
            tmp_path
            / "procedures.xlsx"
        )
    )

    assert (
        dataset
        .for_requirement(
            "AC.L2-3.1.1"
        )[0]
        .sprs_weight
        == 5
    )

    assert (
        dataset
        .for_requirement(
            "IA.L2-3.5.3"
        )[0]
        .sprs_weight
        == 3
    )


def test_missing_file_raises(
    loader: AssessmentProcedureLoader,
) -> None:
    with pytest.raises(
        AssessmentProcedureLoaderError
    ):
        loader.load(
            "does_not_exist.xlsx"
        )


def test_wrong_sheet_raises(
    tmp_path: Path,
) -> None:
    path = (
        tmp_path
        / "wrong_sheet.xlsx"
    )

    workbook = Workbook()
    workbook.active.title = "Wrong"
    workbook.save(path)

    loader = AssessmentProcedureLoader(
        framework_id="CMMC_L2"
    )

    with pytest.raises(
        AssessmentProcedureLoaderError
    ):
        loader.load(
            path
        )


def test_missing_required_header_raises(
    tmp_path: Path,
) -> None:
    path = (
        tmp_path
        / "missing_header.xlsx"
    )

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "SP800-171A"

    worksheet.append(
        [
            "Family",
            "Identifier",
            "Security Requirement",
        ]
    )

    workbook.save(path)

    loader = AssessmentProcedureLoader(
        framework_id="CMMC_L2"
    )

    with pytest.raises(
        AssessmentProcedureLoaderError
    ):
        loader.load(
            path
        )


def test_objective_before_requirement_raises(
    tmp_path: Path,
) -> None:
    path = (
        tmp_path
        / "bad_order.xlsx"
    )

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "SP800-171A"

    worksheet.append(
        HEADERS
    )

    worksheet.append(
        [
            "AC",
            "3.1.1[a]",
            "",
            "",
            "authorized users are identified",
            "",
            "",
            "",
        ]
    )

    workbook.save(path)

    loader = AssessmentProcedureLoader(
        framework_id="CMMC_L2"
    )

    with pytest.raises(
        AssessmentProcedureLoaderError
    ):
        loader.load(
            path
        )


def test_objective_must_match_current_requirement(
    tmp_path: Path,
) -> None:
    path = (
        tmp_path
        / "wrong_parent.xlsx"
    )

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "SP800-171A"

    worksheet.append(
        HEADERS
    )

    worksheet.append(
        [
            "AC",
            "3.1.1",
            "",
            "Requirement text.",
            "Determine if:",
            "access control policy",
            "",
            "",
        ]
    )

    worksheet.append(
        [
            "AC",
            "3.1.2[a]",
            "",
            "",
            "objective text",
            "",
            "",
            "",
        ]
    )

    workbook.save(path)

    loader = AssessmentProcedureLoader(
        framework_id="CMMC_L2"
    )

    with pytest.raises(
        AssessmentProcedureLoaderError
    ):
        loader.load(
            path
        )