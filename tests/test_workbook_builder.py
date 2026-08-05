from pathlib import Path

from openpyxl import load_workbook

from src.workbook import WorkbookBuilder


ROOT = Path(__file__).resolve().parents[1]


def test_workbook_builder_creates_file(
    tmp_path: Path,
) -> None:
    output_path = tmp_path / "CMMC_Assessment.xlsx"

    builder = WorkbookBuilder(
        project_root=ROOT,
        output_path=output_path,
    )

    result = builder.build()

    assert result == output_path
    assert output_path.exists()
    assert output_path.stat().st_size > 0


def test_workbook_contains_expected_sheets(
    tmp_path: Path,
) -> None:
    output_path = tmp_path / "CMMC_Assessment.xlsx"

    builder = WorkbookBuilder(
        project_root=ROOT,
        output_path=output_path,
    )

    builder.build()

    workbook = load_workbook(
        output_path,
        data_only=False,
    )

    expected_sheets = [
        "Cover",
        "Dashboard",
        "Assessment",
        "Domain Summary",
        "Evidence",
        "POA&M",
        "SSP Crosswalk",
        "Assessment History",
        "Executive Report",
        "Settings",
        "_Lists",
    ]

    assert workbook.sheetnames == expected_sheets


def test_assessment_contains_110_requirements(
    tmp_path: Path,
) -> None:
    output_path = tmp_path / "CMMC_Assessment.xlsx"

    builder = WorkbookBuilder(
        project_root=ROOT,
        output_path=output_path,
    )

    builder.build()

    workbook = load_workbook(
        output_path,
        data_only=False,
    )

    worksheet = workbook["Assessment"]

    requirement_ids = [
        worksheet.cell(
            row=row,
            column=2,
        ).value
        for row in range(6, 116)
    ]

    assert len(requirement_ids) == 110
    assert requirement_ids[0] == "AC.L2-3.1.1"
    assert requirement_ids[-1] == "SI.L2-3.14.7"
    assert len(set(requirement_ids)) == 110


def test_assessment_has_expected_formulas(
    tmp_path: Path,
) -> None:
    output_path = tmp_path / "CMMC_Assessment.xlsx"

    builder = WorkbookBuilder(
        project_root=ROOT,
        output_path=output_path,
    )

    builder.build()

    workbook = load_workbook(
        output_path,
        data_only=False,
    )

    worksheet = workbook["Assessment"]

    assert worksheet["J6"].value.startswith("=IF(")
    assert worksheet["O6"].value.startswith("=IF(")


def test_dashboard_has_score_formula(
    tmp_path: Path,
) -> None:
    output_path = tmp_path / "CMMC_Assessment.xlsx"

    builder = WorkbookBuilder(
        project_root=ROOT,
        output_path=output_path,
    )

    builder.build()

    workbook = load_workbook(
        output_path,
        data_only=False,
    )

    worksheet = workbook["Dashboard"]

    assert worksheet["B7"].value == (
        "=110-SUM(Assessment!J6:J115)"
    )


def test_hidden_lists_sheet_is_very_hidden(
    tmp_path: Path,
) -> None:
    output_path = tmp_path / "CMMC_Assessment.xlsx"

    builder = WorkbookBuilder(
        project_root=ROOT,
        output_path=output_path,
    )

    builder.build()

    workbook = load_workbook(
        output_path,
        data_only=False,
    )

    assert workbook["_Lists"].sheet_state == (
        "veryHidden"
    )


def test_assessment_data_validation_exists(
    tmp_path: Path,
) -> None:
    output_path = tmp_path / "CMMC_Assessment.xlsx"

    builder = WorkbookBuilder(
        project_root=ROOT,
        output_path=output_path,
    )

    builder.build()

    workbook = load_workbook(
        output_path,
        data_only=False,
    )

    worksheet = workbook["Assessment"]

    validations = list(
        worksheet.data_validations.dataValidation
    )

    assert len(validations) >= 3