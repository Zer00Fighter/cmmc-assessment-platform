from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.ssp_export import validate_ssp_readiness, write_readiness_report
from tests.test_ssp_word_exporter import _create_workbook


def _complete_demographics(workbook) -> None:
    cover = workbook["Cover"]
    values = {
        6: ("Organization Name", "Example Organization"),
        7: ("Assessment Name", "Enterprise System"),
        8: ("Assessment Scope", "CUI enclave"),
        9: ("CAGE Code", "1A2B3"),
        11: ("Assessment Start Date", "2026-08-01"),
        12: ("Assessment End Date", "2026-08-13"),
        13: ("Lead Assessor", "Lead Assessor"),
    }
    for row, (label, value) in values.items():
        cover.cell(row, 2, label)
        cover.cell(row, 3, value)


def test_unassessed_workbook_reports_blockers_and_warnings(tmp_path: Path) -> None:
    workbook_path = tmp_path / "omni.xlsx"
    _create_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    workbook["Assessment"]["I6"] = "NOT ASSESSED"
    workbook["Assessment"]["R6"] = ""
    workbook["SSP Crosswalk"]["D6"] = ""
    workbook["SSP Crosswalk"]["E6"] = ""
    workbook["SSP Crosswalk"]["G6"] = "Not Mapped"
    workbook.save(workbook_path)

    report = validate_ssp_readiness(workbook_path)

    assert not report.ready
    assert any(issue.field == "Assessment Status" for issue in report.blockers)
    assert any(issue.field == "Supporting Artifacts" for issue in report.warnings)
    output = write_readiness_report(report, tmp_path / "readiness.txt")
    assert "Status: NOT READY" in output.read_text(encoding="utf-8")


def test_completed_workbook_is_ready(tmp_path: Path) -> None:
    workbook_path = tmp_path / "omni.xlsx"
    _create_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    _complete_demographics(workbook)
    workbook["Assessment"]["I6"] = "MET"
    workbook["Assessment"]["Q6"] = "Control AC.L2-3.1.1"
    workbook["Assessment"]["R6"] = "The control operates as designed."
    workbook["SSP Crosswalk"]["D6"] = "Access Control Policy"
    workbook["SSP Crosswalk"]["E6"] = "Quarterly access review"
    workbook["SSP Crosswalk"]["G6"] = "Mapped"
    workbook.save(workbook_path)

    report = validate_ssp_readiness(workbook_path)

    assert report.ready
    assert not report.blockers
    assert not report.warnings
    assert report.completion_percent == 100.0
